import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from  matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from PyQt5 import QtCore, uic
from PyQt5.QtWidgets import QLineEdit,QDialog,QMessageBox,QMainWindow,QCompleter,QTableWidgetItem,QDialogButtonBox,QFormLayout,QApplication,QWidget
from PyQt5.QtGui import QPixmap , QIcon,QDesktopServices,QMovie,QColor,QTransform
from PyQt5.QtCore import Qt,QEasingCurve, QPropertyAnimation,QEvent,QStringListModel,QSize,QUrl,QTimer
import rpyc
import socket
from time import sleep, strftime
from datetime import datetime, timedelta,date 
import json
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch , mm
from reportlab.lib.colors import white, black
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from subprocess import Popen
from concurrent.futures import ThreadPoolExecutor
from random import randint

from app.modulos.helpers import Imagen, Funciones

class Vendedor(QMainWindow):
    ventana_login = 0
    def __init__(self):
        super( Vendedor,self).__init__()
        uic.loadUi('app/ui/vendedor.ui',self)
# ------ VARIABLES GLOBALES -----------
        self.datos_usuario = None
        self.conexion = None
        self.host = None
        self.puerto = None
        self.url_retiros = None #Para abrir app Retiros.
        self.url_descarga = None #Para nuevas versiones.
        self.carpeta = None
        self.dir_informes = None
        
        self.vendedores = None
        self.aux_tabla = None #tabla de respaldo, para usar filtros. usada par fitrar x vendedor, nulas y manuales incompletas
        self.bol_fact = None # boletas y facturas encontradas
        self.guias = None    # guias encontradas
        self.manuales = None # respaldo de solo ordenes manuales
        # -----------
        self.nro_doc = 0  
        self.nro_orden = 0 #NRO FOLIO DE LA ORDEN CREADA
        self.items = None
        self.vendedor = None
        self.tipo_doc = None
        self.fecha_venta = None
        self.inter = None #nro interno del doc
        self.tipo = None # tipo de orden (dim,elab,pall o carp)
        self.manual = None #Si la orden se hizo manual o no
        self.fecha_orden = None #FECHA en la que se creo la orden, formato DATE 
        self.nro_reingreso = 0 #folio del reingreso
        self.clave = None #usada para funciones de super usuario
        self.nro_reingreso = 0
        self.aux_vendedor = None
        self.anterior = None #Determina si al volver atras, vuelve a buscar ordenes o estadisticas -> buscar ordenes manuales
        self.anterior2 = None #Determina si volver atras vuelve a gestion de reingresos o busqueda de ordenes de trabajo.
        self.aux_detalle = None
        self.version = 0.0
        self.aux_version = 0.0 #v6.0 Usada al actualizar reingreso
        self.executor = ThreadPoolExecutor(max_workers=3) #Hilos para hacer Sondeos a la base de datos
        self.sondeo_params = {
            "estado_sondeo_manual" : False, 
            "intervalo_sondeo_manual" : 300, # 300 segundos x DEFECTO
            "omitir_uso_carpinteria" : False,
            "omitir_vinculo_exitoso" : False,
            "rango_dias": 7
        }
        self.datos_orden = {}  #Datos necesarios para crear,actualizar una orden de trabajo.
        self.datos_impresion = {} #Datos para impresion voucher.
#--------------- FUNCIONES GLOBALES ---------

        self.cargar_config()
        #LISTA PRODUCTOS
        self.completer_productos = QCompleter()
        self.lista_productos = []
        self.lista_codigos = []
        self.style_line_edit = ""
        self.aux_doc_venta = {} # Datos temporales de boleta,factura o guia. Usado para almacenar respectivos.
        #Login v6.0
        self.inicializar_login()
        self.btn_iniciar.clicked.connect(self.iniciar_session)
        #self.stackedWidget.setCurrentWidget(self.inicio)
        

        #buscar venta
        self.btn_buscar_1.clicked.connect(self.buscar_documento)
        self.comboBox_1.currentIndexChanged['QString'].connect(self.filtrar_vendedor)
        self.btn_crear_1.clicked.connect(self.inicializar_crear_orden)
        self.btn_atras_1.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.inicio))

        #crear orden
        self.btn_registrar_2.clicked.connect(self.registrar_orden)
        self.btn_vale_despacho.clicked.connect(lambda: self.registrar_vale_despacho(None) ) #Sin datos extras
        self.btn_ver_vale_despacho.clicked.connect(self.visualizar_vale_despacho)
        self.btn_atras_2.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.buscar_venta))
        self.btn_agregar.clicked.connect(self.agregar)
        self.btn_eliminar.clicked.connect(self.eliminar)
        self.btn_rellenar.clicked.connect(self.rellenar)

        self.completer = QCompleter()
        self.completer.activated.connect(self.rellenar_datos_cliente)
        self.nombres = []
        self.telefonos = []
        self.contactos = []

        self.tableWidget_2.setColumnWidth(0,80)
        self.tableWidget_2.setColumnWidth(1,430)
        self.tableWidget_2.setColumnWidth(2,85)
        
        #buscar orden
        self.btn_dimensionado.clicked.connect(self.buscar_dimensionado)
        self.btn_elaboracion.clicked.connect(lambda: self.busqueda_general('ELABORACION'))
        self.btn_carpinteria.clicked.connect(lambda: self.busqueda_general('CARPINTERIA'))
        self.btn_pallets.clicked.connect(lambda: self.busqueda_general('PALLETS'))
        self.btn_sin_transformacion.clicked.connect(lambda: self.busqueda_general('SIN_TRANSFORMACION'))

        self.btn_modificar_4.clicked.connect(self.inicializar_modificar_orden)
        self.btn_atras_4.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.inicio))

        #modificar orden
        self.detalle = None
        self.aux_obs = '' # respaldo de observacion
        self.datosTablaOrden = None # Respalda la tabla de la orden.
        self.btn_guardar_5.clicked.connect( self.actualizar_orden )
        self.btn_atras_5.clicked.connect(self.decidir_atras)
        self.btn_ver_5.clicked.connect(lambda: self.ver_pdf(self.tipo) )
        self.btn_agregar_.clicked.connect(self.agregar_2)
        self.btn_eliminar_.clicked.connect(self.eliminar_2)
        self.btn_anular_5.clicked.connect(self.anular)
        #v6.0 deshabilitados, hide en .ui
        #self.btn_copiar_detalle.clicked.connect(self.guardarDetalleOrden)
        #self.btn_pegar_detalle.clicked.connect(self.pegarDetalleOrden)
        # v6.1 clonar orden
        self.btn_duplicar_orden.clicked.connect(self.duplicar_orden)

        self.r_uso_interno_5.stateChanged.connect( lambda: self.cambiar_observacion('modificar') ) #v5.9
        self.r_separar_material_5.stateChanged.connect(lambda: self.cambiar_observacion('modificar') ) #v5.9
        self.r_despacho_5.stateChanged.connect( lambda: self.cambiar_observacion('modificar')) #v5.9
        self.r_facturar_5.stateChanged.connect( lambda: self.cambiar_observacion('modificar')) #v5.9
        self.r_enchape_5.stateChanged.connect( lambda: self.cambiar_observacion('modificar')) #v5.9

        # reingreso
        self.btn_reingreso_4.clicked.connect(self.inicializar_reingreso)
        self.btn_generar_reingreso.clicked.connect(self.registrar_reingreso)
        self.btn_agregar_2.clicked.connect(self.agregar_3)
        self.btn_eliminar_2.clicked.connect(self.eliminar_3)
        self.btn_atras_7.clicked.connect( lambda: self.stackedWidget.setCurrentWidget(self.buscar_orden) )

        # GESTION REINGRESO
        self.btn_buscar_reingreso.clicked.connect(self.buscar_reingreso_1)
        self.btn_mod_reingreso_1.clicked.connect(self.vista_modificar_reingreso)
        self.btn_anular_reingreso.clicked.connect(lambda:  self.anular_validar_reingreso('ANULAR') )
        self.btn_validar_reingreso.clicked.connect( lambda:  self.anular_validar_reingreso('VALIDAR')  )

        self.btn_guardar_6.clicked.connect(lambda: self.reingreso_manual('actualizar')  )
        self.btn_ver_pdf_reingreso.clicked.connect(self.verificar_pdf_reingreso)
        # INGRESO MANUAL
        #  ---- ORDEN MANUAL ----
        self.txt_descripcion_1.textChanged.connect(self.buscar_descripcion)
        self.txt_codigo_1.textChanged.connect(self.buscar_codigo)
        self.btn_add.clicked.connect(self.add_descripcion)
        
        self.r_uso_interno_1.stateChanged.connect( lambda: self.cambiar_observacion('manual') ) #v5.9
        self.r_separar_material_1.stateChanged.connect(lambda: self.cambiar_observacion('manual') ) #v5.9
        self.r_despacho_1.stateChanged.connect( lambda: self.cambiar_observacion('manual')) #v5.9
        self.r_facturar_1.stateChanged.connect( lambda: self.cambiar_observacion('manual')) #v5.9
        self.r_enchape_1.stateChanged.connect( lambda: self.cambiar_observacion('manual')) #v5.9
        

        self.btn_registrar_1.clicked.connect(self.registrar_orden_manual)
        self.btn_agregar_1.clicked.connect(self.agregar_4)
        self.btn_eliminar_1.clicked.connect(self.eliminar_4)

        self.completer_manual = QCompleter()
        self.completer_manual.activated.connect(self.rellenar_datos_cliente_manual)
        self.completer_reingreso_manual = QCompleter()
        self.completer_reingreso_manual.activated.connect(self.rellenar_datos_cliente_reingreso_manual)

        self.r_sin_trans_1.toggled.connect(self.habilitar_separar_material)

        #  ---- REINGRESO MANUAL ----
        self.btn_registrar_6.clicked.connect(lambda: self.reingreso_manual('registrar') )
        self.txt_descripcion_7.textChanged.connect(self.buscar_descripcion_2)
        self.btn_add_6.clicked.connect(self.add_descripcion_2)

        self.btn_agregar_6.clicked.connect(self.agregar_6)
        self.btn_eliminar_6.clicked.connect(self.eliminar_6)

        self.btn_atras_3.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.inicio))
        self.btn_atras_6.clicked.connect(self.decidir_atras2)
        #informes
        self.btn_generar_informe.clicked.connect(self.generar_informe)
        self.comboBox.currentIndexChanged['QString'].connect(self.vista_reingreso)
        self.btn_eliminar_exel.clicked.connect(self.eliminar_excel)
        self.btn_actualizar.clicked.connect(self.actualizar)
        self.btn_abrir.clicked.connect(self.abrir)
        #ESTADISTICAS
        self.btn_grafico.clicked.connect(self.crear_grafico)
        self.btn_buscar_manuales.clicked.connect(self.buscar_manuales)
        self.btn_modificar_5.clicked.connect(self.continuar_orden_manual)
        self.check_incompletas.stateChanged.connect(self.filtrar_solo_incompletas)
        self.btn_notificacion1.clicked.connect(self.easter_egg )
        #SONDEOS

        self.btn_iniciar_sondeo_manual.clicked.connect(self.iniciar_sondeo)
        self.btn_detener_sondeo_manual.clicked.connect(self.finalizar_sondeo)
        self.btn_actualizar_config_sondeo.clicked.connect(self.actualizar_sondeo_config)
        self.btn_obtener_sondeo.clicked.connect(lambda: self.obtener_datos_sondeo(True))
        
        #generar clave
        #self.btn_generar_clave.clicked.connect(self.generar_clave)

        #CONFIGURACION
        self.btn_configuracion.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.configuracion))
        self.btn_inicio.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.inicio))
        self.btn_retiros.clicked.connect( lambda: QDesktopServices.openUrl(QUrl( self.url_retiros )))

        #SIDE MENU BOTONES
        self.btn_buscar.clicked.connect(self.inicializar_buscar_venta)
        self.btn_modificar.clicked.connect(self.inicializar_buscar_orden)
        self.btn_orden_manual.clicked.connect(self.inicializar_ingreso_manual)
        self.btn_informe.clicked.connect(self.inicializar_informe)
        self.btn_atras.clicked.connect(self.cerrar_sesion)
        self.btn_conectar.clicked.connect(lambda: self.executor.submit(self.conectar) )
        self.btn_manual.clicked.connect(self.conectar_manual)
        self.btn_estadisticas.clicked.connect(self.inicializar_estadisticas)
        self.btn_reingreso.clicked.connect(self.inicializar_buscar_reingreso)
        self.btn_menu.clicked.connect(self.mostrar_menu)
        
        
    def cargar_config(self):
        print('|-- Cargando Config.json ')

        # Lee el archivo config.json
        with open('app/config.json', 'r') as f:
            config = json.load(f)

        try:
            # Obtén el entorno del archivo JSON
            environment = config["environment"]
            # Obtén los valores de host y port según el entorno
            self.host = config[environment]['host']
            self.puerto = config[environment]['port']
            self.url_retiros = config[environment]['url_retiros']
            print('|-- Creando Thread para crear conexion. ')
            self.executor.submit(self.conectar)
            print('|-- Threada para crear conexion creado. ')
            #self.conectar()
        except KeyError:
            print('No host - port in config ')
            pass
        #Datos usuario
        try:
            user = config['user']
            password = config['password']
            print(user , password)
            if user and password:
                self.txt_usuario.setText(user)
                self.txt_contra.setText(password)
        except KeyError:
            print('No user - password in config')
            pass
        #Datos sondeo params
        try:
            print('|_ Intervalo sondeo:',config["intervalo_sondeo_manual"])
            self.sondeo_params['intervalo_sondeo_manual'] = int(config["intervalo_sondeo_manual"])

            print('|_ Omitir_uso_carpinteria: ',config["omitir_uso_carpinteria"])
            if config["omitir_uso_carpinteria"]:
                self.sondeo_params["omitir_uso_carpinteria"] = config["omitir_uso_carpinteria"]
                self.label_75.setText('SI')
                self.ch_omitir_uso_carp.setChecked(True)

            
            print('|_ Omitir vinclo exitoso: ',config["omitir_vinculo_exitoso"])
            if config["omitir_vinculo_exitoso"]:
                self.sondeo_params["omitir_vinculo_exitoso"] = config["omitir_vinculo_exitoso"]
                self.label_76.setText('SI')
                self.ch_omitir_vinculo.setChecked(True)
            
            print('|_ Rango dias sondeo: ',config["rango_dias"])
            self.sondeo_params["rango_dias"] = int(config["rango_dias"])
            self.version = float(config["version"])
            self.aux_version = float(config["version"])
            self.lb_version.setText(str(self.version))

        except Exception as e:
            # Captura la excepción y muestra el mensaje de error
            print("Error:", str(e))
        
        try:
            self.datos_impresion["imprimir_voucher_en"] = config["imprimir_voucher_en"]
            self.datos_impresion["nombre_impresora_voucher"] = config["nombre_impresora_voucher"]
        except Exception as e:
            # Captura la excepción y muestra el mensaje de error
            print("Error:", str(e))
            

    def inicializar_login(self):
        print("|-- Mostrando Login ...")
        self.left_menu_container.hide()
        self.lb_vendedor.setText('Por Confirmar ...')
        self.txt_contra.setEchoMode(QLineEdit.Password)
        #self.txt_usuario.setFocus()
        logo = QPixmap('app/icono_imagen/madenco logo.png')
        self.lb_logo.setPixmap(logo)
        self.stackedWidget.setCurrentWidget(self.login)
        #Loagind de conexion
        movie = QMovie("app/icono_imagen/conn2.gif")  # Asegúrate de tener el archivo GIF en la misma carpeta
        self.lb_gif.setMovie(movie)
        movie.start()

        print('|-- Mostrando Loading')
        self.stackedWidgetLogin.setCurrentWidget(self.loading)
        #Loading
        movie = QMovie("app/icono_imagen/gif.gif")
        self.lb_loading.setMovie(movie)
        movie.start()
        # Retrasar la transición al input después de X segundos
        QTimer.singleShot(1000, lambda: (
            print('|-- Input'),
            self.stackedWidgetLogin.setCurrentWidget(self.input),
            self.btn_iniciar.setDefault(True)  # Establecer el botón como predeterminado para que se active al presionar Enter
        ))
        
    def iniciar_session(self):
        print("|-- Iniciando sesion ..." )
        usuario = self.txt_usuario.text()
        contra = self.txt_contra.text()
        if self.conexion:
            try:
                resultado = self.conexion.root.obtener_usuario_activo()
                encontrado = False
                for item in resultado:
                    if item[0] == usuario and item[1] == contra and item[6] == 'vendedor':
                        encontrado = True
                        self.datos_usuario = item
                        self.recordar_cuenta()
                        self.inicializar()
                if encontrado == False:
                    QMessageBox.about(self ,'ERROR', 'Usuario o contraseña invalidas')

            except EOFError:
                #QMessageBox.about(self ,'Conexion', 'El servidor no responde')
                self.conexion_perdida()
            except AttributeError:
                pass
        else:
            #QMessageBox.about(self ,'Conexion', 'SIN CONEXION\nIntente una nueva conexion.\n(Boton "Conectar" o "Conectar manual")')
            self.conexion_perdida()
        
    def recordar_cuenta(self):
        
        with open("app/config.json", "r") as f:
            config = json.load(f)

        if self.ch_recordar_cuenta.isChecked():
            print('|-----> recordando cuenta ...')
            config["user"] = self.txt_usuario.text()
            config["password"] = self.txt_contra.text()
            print('|-- Cuenta recordada con exito' )
        if self.ch_olvidar_cuenta.isChecked():
            print('|--- Olvidando cuenta ..')
            config["user"] = ''
            config["password"] = ''
            print('|-- Cuenta olvidada con exito' )

        with open("app/config.json", "w") as f:
            json.dump(config, f, indent=4)
        
            
    def inicializar(self):
        print('#'*55)
        print('|-- Inicializando SAGOT Vendedor ...')
        
        self.btn_atras.setIcon(QIcon('app/icono_imagen/logout.png'))

        self.btn_buscar.setIcon(QIcon('app/icono_imagen/buscar venta.png'))
        self.btn_modificar.setIcon(QIcon('app/icono_imagen/buscar orden.png'))
        self.btn_orden_manual.setIcon(QIcon('app/icono_imagen/orden manual.png'))
        #self.btn_generar_clave.setIcon(QIcon('app/icono_imagen/generar clave.png'))
        self.btn_informe.setIcon(QIcon('app/icono_imagen/informes.png'))
        self.btn_estadisticas.setIcon(QIcon('app/icono_imagen/estadisticas2.png'))
        self.btn_configuracion.setIcon(QIcon('app/icono_imagen/configuracion.png'))
        self.btn_reingreso.setIcon(QIcon('app/icono_imagen/reingreso.png'))
        self.btn_retiros.setIcon(QIcon('app/icono_imagen/order-delivery.png'))

        self.btn_generar_clave.hide() #v6.0 Se Oculta generar clave.
        self.style_line_edit = "QLineEdit { border: none; }" #Estilo para line edit de prediccion de productos en QTablewidget.
        self.btn_orden_manual.show()
        self.btn_informe.show()
        #v6.0
        self.left_menu_container.show()
        actual = os.path.abspath(os.getcwd())
        actual = actual.replace('\\' , '/')
        self.carpeta = actual.replace('\\' , '/')
        self.dir_informes = os.path.join(self.carpeta, 'app/informes')
        self.dir_ordenes = os.path.join(self.carpeta, 'app/ordenes')
        self.dir_reingreso = os.path.join(self.carpeta, 'app/reingresos')
        self.dir_formatos = os.path.join(self.carpeta, 'app/formatos')
        '''print('|--Carpeta root:' + self.carpeta)
        print('|--Carpeta informes: ' + self.dir_informes)
        print('|--Carpeta ordenes: ', self.dir_ordenes)
        print('|--Carpeta formatos: ',self.dir_formatos)
        print('|--Carpeta reingreso: ', self.dir_reingreso)
        print('|--Carpeta informes: ',self.dir_informes)'''
        
        self.btn_inicio.setIcon(QIcon('app/icono_imagen/enco_logov3.png'))

        menu = QPixmap('app/icono_imagen/menu_v4.png')
        self.btn_menu.setIcon(QIcon(menu))
        print(f'|-- Verificando usuario --> {self.datos_usuario}')
        if self.datos_usuario:  #Si existen los datos del usuario, x ende se inicio sesion correctamente...
            self.lb_vendedor.setText(self.datos_usuario[8]) #NOMBRE DEL VENDEDOR
            self.vendedor = self.datos_usuario[8]   #solucion al error cuando se modifica la orden y se crear el pdf interno con vendedor = none. Esto cierra el programa.

            if self.datos_usuario[4] == 'NO': #Si no es super usuario
                #self.btn_generar_clave.hide() #no puede generar claves
                detalle = json.loads(self.datos_usuario[7])
                funciones = detalle['vendedor']
                if not 'manual' in funciones:
                    self.btn_orden_manual.hide() #no puede generar ordenes manuales
                if not 'informes' in funciones:
                    self.btn_informe.hide() #no puede generar informes
        
        
        self.iniciar_sondeo()
        #se cargan los productos.
        self.cargar_productos()
        #Se busca actualizacion
        #self.buscar_actualizacion() v6.05 Se creo un launcher

        self.stackedWidget.setCurrentWidget(self.inicio)
        """# Se registran las conexiones de los botones
        button_connections = {
            self.btn_ver_pdf: self.ver_pdf,
            self.btn_crear_pdf: self.crear_pdf,
            # ... y así sucesivamente para todos los botones
        }

         for button, function in button_connections.items():
            button.clicked.connect(function) """
        
        # Crear un QTextEdit para mostrar el contenido del archivo
        print("|-- Leyendo historial ...")
        self.txt_historial.setReadOnly(True)

        # Leer el contenido del archivo y establecerlo en el QTextEdit
        with open( "Historial_de_versiones.txt", 'r') as file:
            content = file.read()
            print("|-- len content historial: ", len(content))
            self.txt_historial.setPlainText(content)

        print('#'*55)
    
    def conectar(self):
        print('|-- Obteniendo conexion ... ')
        self.btn_conectar.setEnabled(False)
        self.lb_conexion.hide()
        self.lb_gif.show()
        

        if self.conexion is not None:
            try:
                # Realiza un ping para verificar la conexión
                self.conexion.ping()
                self.lb_conexion.setText('CONECTADO')
            except (ConnectionResetError, EOFError):
                self.conexion = None
                self.lb_conexion.setText('CONEXIÓN PERDIDA')
        else:
            try:
                if self.host and self.puerto:
                    self.conexion = rpyc.connect(self.host, self.puerto)
                    self.lb_conexion.setText('CONECTADO')
                else:
                    self.lb_conexion.setText('Host y Puerto no encontrados')
            except ConnectionRefusedError:
                self.lb_conexion.setText('EL SERVIDOR NO RESPONDE')
            except socket.error:
                self.lb_conexion.setText('SERVIDOR FUERA DE RED')
        print("| (CONEXION) : ",self.conexion)
        self.lb_gif.hide()
        self.btn_conectar.setEnabled(True)
        self.lb_conexion.show()

        print('|-- Conexion finalizada. ')


    def conectar_manual(self):
        dialog = InputDialog('HOST:','PUERTO:','CONECTAR MANUAL', self)
        dialog.resize(250,100)   
        if dialog.exec():
                hostx , puertox = dialog.getInputs()
                try:
                    if hostx != '':
                        puertox = int(puertox)
                        self.host = hostx
                        self.puerto = puertox
                        self.conexion = rpyc.connect(hostx , puertox)
                        self.lb_conexion.setText('CONECTADO')
                    else: 
                        QMessageBox.about(self,'ERROR' ,'Ingrese un host antes de continuar')
                except ValueError:
                    QMessageBox.about(self,'ERROR' ,'Ingrese solo numeros en el PUERTO')
                except ConnectionRefusedError:
                    self.lb_conexion.setText('EL SERVIDOR NO RESPONDE')
                except socket.error:
                    self.lb_conexion.setText('SERVIDOR FUERA DE LA RED')
                    #QMessageBox.about(self,'ERROR' ,'No se puede establecer la conexion con el servidor')
# ------------   FUNCIONES BUSCAR VENTA  -----------------
    def inicializar_buscar_venta(self):
        print('------- INICIALIZAR BUSCAR VENTA ---------')
        self.cargar_clientes('normal')
        self.txt_interno_1.setText('')
        self.txt_cliente_1.setText('')
        self.tableWidget_1.setRowCount(0)
        self.radio2_1.setChecked(True)
        self.dateEdit_1.setCalendarPopup(True)
        self.dateEdit_1.setDate(datetime.now().date())
        if self.windowState() != Qt.WindowMaximized:
            print('|-- Ajustando tamaño de las tablas para ventana NORMAL')
            self.tableWidget_1.setColumnWidth(0,100) #interno
            self.tableWidget_1.setColumnWidth(1,100) #documento
            self.tableWidget_1.setColumnWidth(2,85) #nro doc
            self.tableWidget_1.setColumnWidth(3,100) #fecha venta
            self.tableWidget_1.setColumnWidth(4,150) #cliente
            self.tableWidget_1.setColumnWidth(5,100) #vendededor
            self.tableWidget_1.setColumnWidth(6,70) #total
            self.tableWidget_1.setColumnWidth(7,80) #despacho
        else:
            print('|-- VENTANA MAXIMIZADA, no es necesario ajustar el tamaño de las tablas')
        
        self.stackedWidget.setCurrentWidget(self.buscar_venta)
        print('----------------'*2)
    
    def buscar_documento(self):
        self.vendedores = []
        largo = self.comboBox_1.count()
        if largo > 1:
            for i in range(largo):
                #print("borranto: " + str(largo - 1) 
                self.comboBox_1.removeItem(1)

        if self.conexion:
            no_encontrados = True
            self.tableWidget_1.setRowCount(0)
            if self.radio1_1.isChecked():  #BUSCANDO POR NUMERO INTERNO
                inter = self.txt_interno_1.text()
                try:
                    inter = int(inter)
                    consulta = self.conexion.root.buscar_venta_interno(inter)
                    guia = self.conexion.root.obtener_guia_interno(inter)
                    print("(BUSQUEDA X INTERNO): ",consulta)
                    if consulta != None :
                        no_encontrados = False
                        print(consulta[1])
                        fila = self.tableWidget_1.rowCount()
                        self.tableWidget_1.insertRow(fila)

                        self.tableWidget_1.setItem(fila , 0 , QTableWidgetItem(str(consulta[0]))) #INTERNO
                        aux_tipo_doc = ""
                        aux_nro_doc = ""
                        if consulta[3] == 0 : #es boleta
                            aux_tipo_doc = "BOLETA"
                            aux_nro_doc = str(consulta[4])
                        elif consulta[4] == 0: #es factura
                            aux_tipo_doc = "FACTURA"
                            aux_nro_doc = str(consulta[3])

                        self.tableWidget_1.setItem(fila , 1 , QTableWidgetItem( aux_tipo_doc )) #TIPO DOCUMENTO
                        self.tableWidget_1.setItem(fila , 2 , QTableWidgetItem( aux_nro_doc ))      #NRO DOCUMENTO

                        self.tableWidget_1.setItem(fila , 3 , QTableWidgetItem(   str(consulta[1]  ))) #FECHA VENTA
                        self.tableWidget_1.setItem(fila , 4 , QTableWidgetItem( str(consulta[6] )   ))      #CLIENTE
                        self.tableWidget_1.setItem(fila , 5 , QTableWidgetItem(    consulta[2] ))             #VENDEDOR
                        
                        self.tableWidget_1.setItem(fila , 6 , QTableWidgetItem(   str(consulta[5]))   )      #TOTAL

                        self.tableWidget_1.setItem(fila , 7 , QTableWidgetItem(   str(consulta[7]))   )      #DESPACHO
                        self.tableWidget_1.setItem(fila , 8 , QTableWidgetItem(   str(consulta[8]))   )      #rut

                        if consulta[9]:
                            self.resaltar_fila_tabla(self.tableWidget_1,fila,2 , consulta[9] )
                            
                        
                    if guia != None:
                        no_encontrados = False
                        
                        consulta = guia
                        detalle = json.loads(consulta[2])

                        fila = self.tableWidget_1.rowCount()
                        self.tableWidget_1.insertRow(fila)

                        self.tableWidget_1.setItem(fila , 0 , QTableWidgetItem(str(consulta[1]))) #INTERNO
                        
                        self.tableWidget_1.setItem(fila , 1 , QTableWidgetItem( 'GUIA' )) #TIPO DOCUMENTO
                        self.tableWidget_1.setItem(fila , 2 , QTableWidgetItem( str( consulta[0])))      #NRO DOCUMENTO

                        self.tableWidget_1.setItem(fila , 3 , QTableWidgetItem(   str(consulta[4]  ))) #FECHA VENTA
                        self.tableWidget_1.setItem(fila , 4 , QTableWidgetItem( str(consulta[3]  )   ) )     #CLIENTE

                        self.tableWidget_1.setItem(fila , 5 , QTableWidgetItem( detalle['vendedor'] ))             #VENDEDOR
                        self.tableWidget_1.setItem(fila , 6 , QTableWidgetItem(   str(detalle['monto_final']))   )      #TOTAL
                        self.tableWidget_1.setItem(fila , 7 , QTableWidgetItem(   str(consulta[7]))   )      #DESPACHO
                        self.tableWidget_1.setItem(fila , 8 , QTableWidgetItem(    detalle['rut']  ) )      #rut
                        
                        if consulta[6]:
                            self.resaltar_fila_tabla(self.tableWidget_1,fila,2 , consulta[6] )
                        
                    if no_encontrados:
                        QMessageBox.about(self,'Busqueda' ,'Documento NO encontrado mediante el N° Interno.')
                    
                except ValueError:
                    QMessageBox.about(self,'ERROR' ,'Ingrese solo numeros')
                except EOFError:
                    self.conexion_perdida()

            elif self.radio2_1.isChecked(): #BUSCANDO POR FECHA ACTUAL
                lista_vendedores = []
                aux_lista = []
                date = self.dateEdit_1.date()
                aux = date.toPyDate()
                inicio = str(aux) + ' ' + '00:00:00'
                fin = str(aux) + ' ' + '23:59:59'
                try:
                    bol_fact = self.conexion.root.buscar_venta_fecha(inicio,fin)
                    guias = self.conexion.root.obtener_guia_fecha(inicio,fin)
                    if bol_fact != ():
                        no_encontrados = False
                        self.bol_fact = bol_fact
                        for consulta in bol_fact:
                            fila = self.tableWidget_1.rowCount()
                            self.tableWidget_1.insertRow(fila)
                            self.tableWidget_1.setItem(fila , 0 , QTableWidgetItem(str(consulta[0]))) #INTERNO
                            if consulta[3] == 0 : #es boleta
                                self.tableWidget_1.setItem(fila , 1 , QTableWidgetItem( 'BOLETA' )) #TIPO DOCUMENTO
                                self.tableWidget_1.setItem(fila , 2 , QTableWidgetItem( str(consulta[4]) ))      #NRO DOCUMENTO
                            elif consulta[4] == 0: #es factura
                                self.tableWidget_1.setItem(fila , 1 , QTableWidgetItem( 'FACTURA' )) #TIPO DOCUMENTO
                                self.tableWidget_1.setItem(fila , 2 , QTableWidgetItem( str( consulta[3])))      #NRO DOCUMENTO

                            self.tableWidget_1.setItem(fila , 3 , QTableWidgetItem(   str(consulta[1]  ))) #FECHA VENTA
                            self.tableWidget_1.setItem(fila , 4 , QTableWidgetItem( str(consulta[6] )   ))      #CLIENTE
                            self.tableWidget_1.setItem(fila , 5 , QTableWidgetItem(    consulta[2] ))             #VENDEDOR
                            
                            
                            aux = consulta[2]
                            aux = aux[0:10]
                            #print(aux)
                            
                            if aux not in aux_lista:
                                aux_lista.append(aux)
                                lista_vendedores.append( consulta[2] )
    
                            self.tableWidget_1.setItem(fila , 6 , QTableWidgetItem(   str(consulta[5]))   )      #TOTAL
                            self.tableWidget_1.setItem(fila , 7 , QTableWidgetItem(   str(consulta[7]))   )      #DESPACHO
                            self.tableWidget_1.setItem(fila , 8 , QTableWidgetItem(   str(consulta[8]))   )      #rut

                            if consulta[9]:
                                self.resaltar_fila_tabla(self.tableWidget_1,fila,2 , consulta[9] )
                    
                    if guias != ():
                        self.guias = guias
                        no_encontrados = False
                        for consulta in guias:
                            detalle = json.loads(consulta[2])

                            fila = self.tableWidget_1.rowCount()
                            self.tableWidget_1.insertRow(fila)

                            self.tableWidget_1.setItem(fila , 0 , QTableWidgetItem(str(consulta[1]))) #INTERNO
                           
                            self.tableWidget_1.setItem(fila , 1 , QTableWidgetItem( 'GUIA' )) #TIPO DOCUMENTO
                            self.tableWidget_1.setItem(fila , 2 , QTableWidgetItem( str( consulta[0])))      #NRO DOCUMENTO

                            self.tableWidget_1.setItem(fila , 3 , QTableWidgetItem(   str(consulta[4]  ))) #FECHA VENTA
                            
                            self.tableWidget_1.setItem(fila , 4 , QTableWidgetItem( str(consulta[3]  )   ) )     #CLIENTE

                            self.tableWidget_1.setItem(fila , 5 , QTableWidgetItem( detalle['vendedor'] ))             #VENDEDOR
                            aux = detalle['vendedor']
                            aux = aux[0:10]
                            #print(aux)
                            
                            if aux not in aux_lista:
                                aux_lista.append(aux)
                                lista_vendedores.append( detalle['vendedor'] )
    
                            self.tableWidget_1.setItem(fila , 6 , QTableWidgetItem(   str(detalle['monto_final']))   )      #TOTAL
                            self.tableWidget_1.setItem(fila , 7 , QTableWidgetItem(   str(consulta[7]))   )      #DESPACHO
                            self.tableWidget_1.setItem(fila , 8 , QTableWidgetItem(    detalle['rut']  ) )      #rut
                            if consulta[6]:
                                self.resaltar_fila_tabla(self.tableWidget_1,fila,2 , consulta[6] )

                    self.vendedores = lista_vendedores
                    for item in self.vendedores:
                        self.comboBox_1.addItem(item)

                    if no_encontrados:
                        QMessageBox.about(self,'Busqueda' ,'Documentos NO encontrados para la fecha indicada')

                except EOFError:
                    self.conexion_perdida()

            elif self.radio3_1.isChecked(): #BUSCANDO POR NOMBRE
                X = 0
                lista_vendedores = []
                aux_lista = []
                nombre = self.txt_cliente_1.text()
                try:
                    bol_fact = self.conexion.root.obtener_venta_nombre(nombre)
                    guias = self.conexion.root.obtener_guia_nombre(nombre)
                    if bol_fact != ():
                        no_encontrados = False
                        self.bol_fact = bol_fact
                        for consulta in bol_fact:
                            fila = self.tableWidget_1.rowCount()
                            self.tableWidget_1.insertRow(fila)
                            self.tableWidget_1.setItem(fila , 0 , QTableWidgetItem(str(consulta[0]))) #INTERNO
                            if consulta[3] == 0 : #es boleta
                                self.tableWidget_1.setItem(fila , 1 , QTableWidgetItem( 'BOLETA' )) #TIPO DOCUMENTO
                                self.tableWidget_1.setItem(fila , 2 , QTableWidgetItem( str(consulta[4]) ))      #NRO DOCUMENTO
                            elif consulta[4] == 0: #es factura
                                self.tableWidget_1.setItem(fila , 1 , QTableWidgetItem( 'FACTURA' )) #TIPO DOCUMENTO
                                self.tableWidget_1.setItem(fila , 2 , QTableWidgetItem( str( consulta[3])))      #NRO DOCUMENTO

                            self.tableWidget_1.setItem(fila , 3 , QTableWidgetItem(   str(consulta[1]  ))) #FECHA VENTA
                            self.tableWidget_1.setItem(fila , 4 , QTableWidgetItem( str(consulta[6] )   ))      #CLIENTE

                            self.tableWidget_1.setItem(fila , 5 , QTableWidgetItem(    consulta[2] ))             #VENDEDOR
                            aux = consulta[2]
                            aux = aux[0:10]
                            #print(aux)
                            
                            if aux not in aux_lista:
                                aux_lista.append(aux)
                                lista_vendedores.append( consulta[2] )
    
                            self.tableWidget_1.setItem(fila , 6 , QTableWidgetItem(   str(consulta[5]))   )      #TOTAL
                            self.tableWidget_1.setItem(fila , 7 , QTableWidgetItem(   str(consulta[7]))   )      #DESPACHO
                            self.tableWidget_1.setItem(fila , 8 , QTableWidgetItem(   str(consulta[8]))   )      #rut

                    if guias != ():
                        self.guias = guias
                        no_encontrados = False
                        for consulta in guias:
                            detalle = json.loads(consulta[2])

                            fila = self.tableWidget_1.rowCount()
                            self.tableWidget_1.insertRow(fila)

                            self.tableWidget_1.setItem(fila , 0 , QTableWidgetItem(str(consulta[1]))) #INTERNO
                           
                            self.tableWidget_1.setItem(fila , 1 , QTableWidgetItem( 'GUIA' )) #TIPO DOCUMENTO
                            self.tableWidget_1.setItem(fila , 2 , QTableWidgetItem( str( consulta[0])))      #NRO DOCUMENTO

                            self.tableWidget_1.setItem(fila , 3 , QTableWidgetItem(   str(consulta[4]  ))) #FECHA VENTA
                            
                            self.tableWidget_1.setItem(fila , 4 , QTableWidgetItem( str(consulta[3] )   ))      #CLIENTE

                            self.tableWidget_1.setItem(fila , 5 , QTableWidgetItem( detalle['vendedor'] ))             #VENDEDOR
                            aux = detalle['vendedor']
                            aux = aux[0:10]
                            #print(aux)
                            
                            if aux not in aux_lista:
                                aux_lista.append(aux)
                                lista_vendedores.append( detalle['vendedor'] )
    
                            self.tableWidget_1.setItem(fila , 6 , QTableWidgetItem(   str(detalle['monto_final']))   )      #TOTAL
                            self.tableWidget_1.setItem(fila , 7 , QTableWidgetItem(   str(consulta[7]))   )      #DESPACHO
                            self.tableWidget_1.setItem(fila , 8 , QTableWidgetItem(    detalle['rut']  ) )      #rut
                            if consulta[6]:
                                self.resaltar_fila_tabla(self.tableWidget_1,fila,2 , consulta[6] )

                    self.vendedores = lista_vendedores
                    for item in self.vendedores:
                        self.comboBox_1.addItem(item)

                    if no_encontrados:
                        QMessageBox.about(self,'Busqueda' ,'Documentos NO encontrados para la fecha indicada')

                except EOFError:
                    self.conexion_perdida()


            self.aux_tabla = self.tableWidget_1
        else:
            self.conexion_perdida()
    def resaltar_fila_tabla(self,tabla,fila,columna, data):
        vinculaciones = json.loads(data)
        try:
            if vinculaciones['ordenes']:
                tabla.item(fila, columna).setBackground(QColor(85, 247, 45))  # Cambia el color de fondo
                tabla.item(fila, columna).setForeground(QColor(0, 0, 0))  # Cambia el color del texto
        except KeyError:
            pass

    def rellenar_tabla(self):
        self.tableWidget_1.setRowCount(0)
        if self.bol_fact != None:
            for consulta in self.bol_fact:
                fila = self.tableWidget_1.rowCount()
                self.tableWidget_1.insertRow(fila)
                self.tableWidget_1.setItem(fila , 0 , QTableWidgetItem(str(consulta[0]))) #INTERNO
                if consulta[3] == 0 : #es boleta
                    self.tableWidget_1.setItem(fila , 1 , QTableWidgetItem( 'BOLETA' )) #TIPO DOCUMENTO
                    self.tableWidget_1.setItem(fila , 2 , QTableWidgetItem( str(consulta[4]) ))      #NRO DOCUMENTO
                elif consulta[4] == 0: #es factura
                    self.tableWidget_1.setItem(fila , 1 , QTableWidgetItem( 'FACTURA' )) #TIPO DOCUMENTO
                    self.tableWidget_1.setItem(fila , 2 , QTableWidgetItem( str( consulta[3])))      #NRO DOCUMENTO

                self.tableWidget_1.setItem(fila , 3 , QTableWidgetItem(   str(consulta[1]  ))) #FECHA VENTA
                self.tableWidget_1.setItem(fila , 4 , QTableWidgetItem( str(consulta[6] )   ))      #CLIENTE
                self.tableWidget_1.setItem(fila , 5 , QTableWidgetItem(    consulta[2] ))             #VENDEDOR
                self.tableWidget_1.setItem(fila , 6 , QTableWidgetItem(   str(consulta[5]))   )      #TOTAL
                self.tableWidget_1.setItem(fila , 7 , QTableWidgetItem(   str(consulta[7]))   )      #DESPACHO
                self.tableWidget_1.setItem(fila , 8 , QTableWidgetItem(   str(consulta[8]))   )      #rut

        if self.guias != None:
            for consulta in self.guias:
                detalle = json.loads(consulta[2])
                fila = self.tableWidget_1.rowCount()
                self.tableWidget_1.insertRow(fila)
                self.tableWidget_1.setItem(fila , 0 , QTableWidgetItem(str(consulta[1]))) #INTERNO
                
                self.tableWidget_1.setItem(fila , 1 , QTableWidgetItem( 'GUIA' )) #TIPO DOCUMENTO
                self.tableWidget_1.setItem(fila , 2 , QTableWidgetItem( str( consulta[0])))      #NRO DOCUMENTO

                self.tableWidget_1.setItem(fila , 3 , QTableWidgetItem(   str(consulta[4]  ))) #FECHA VENTA
                self.tableWidget_1.setItem(fila , 4 , QTableWidgetItem( str(consulta[3] )   ))      #CLIENTE
                self.tableWidget_1.setItem(fila , 5 , QTableWidgetItem( detalle['vendedor'] ))             #VENDEDOR
            
                self.tableWidget_1.setItem(fila , 6 , QTableWidgetItem(   str(detalle['monto_final']))   )      #TOTAL
                self.tableWidget_1.setItem(fila , 7 , QTableWidgetItem(   str(consulta[7]))   )      #DESPACHO
                self.tableWidget_1.setItem(fila , 8 , QTableWidgetItem(    detalle['rut']  ) )      #rut

    def filtrar_vendedor(self):
        vendedor = self.comboBox_1.currentText()
        
        if vendedor == 'TODOS':
            self.rellenar_tabla()
        else:
            self.rellenar_tabla()
            self.tableWidget_1 = self.aux_tabla
            remover = []
            vendedor = vendedor[0:10]
            print('solo: ' + vendedor)
            column = 5 #COLUMNA DEL CVENDEDOR EN LA TABLA
            # rowCount() This property holds the number of rows in the table
            for row in range(self.tableWidget_1.rowCount()): 
                # item(row, 0) Returns the item for the given row and column if one has been set; otherwise returns nullptr.
                _item = self.tableWidget_1.item(row, column) 
                if _item:            
                    item = self.tableWidget_1.item(row, column).text()
                    #print(f'row: {row}, column: {column}, item={item}')
                    aux_item = item[0:10]
                    if aux_item != vendedor:
                        remover.append(row)
            print(remover)
            k = 0
            for i in remover:
                self.tableWidget_1.removeRow(i - k)
                k += 1

# ------------ FUNCIONES PARA VISTA CREAR ORDEN ---------------------
    def inicializar_crear_orden(self) :
        seleccion = self.tableWidget_1.currentRow()
        if seleccion != -1:
            _item = self.tableWidget_1.item( seleccion, 0) 
            if _item:            
                interno = self.tableWidget_1.item(seleccion, 0 ).text()
                aux_tipo_doc = self.tableWidget_1.item(seleccion, 1 ).text()
                aux_nro_doc = self.tableWidget_1.item(seleccion, 2 ).text()
                
                nro_interno = int(interno)
                self.inter = nro_interno
                rut = self.tableWidget_1.item(seleccion, 8 ).text()
                nombre_cliente  = self.tableWidget_1.item(seleccion, 4 ).text()
                print('|-- Ventana crear orden ... para: '+ aux_tipo_doc+ ' - Folio: ' + aux_nro_doc + ' | Interno: ' + str(nro_interno) + ' | rut: ' + rut + ' | cliente: ' + nombre_cliente)

                fecha = datetime.now().date()
                self.fecha.setCalendarPopup(True)
                self.r_facturar_1.setChecked(False) # ASI NO MUESTRA EL "POR FACTURAR" 
                self.r_uso_interno_1.setChecked(False) # ASI NO MUESTRA EL "uso interno" 

                self.fecha.setDate(fecha)    #FECHA ESTIMADA DE ENTREGA
                self.lb_interno.setText(str(nro_interno)) #nro interno
                self.nombre_2.setText('')
                self.telefono_2.setText('')
                self.contacto_2.setText('')
                self.oce_2.setText('')
                self.r_enchape.setChecked(False)
                self.r_despacho.setChecked(False)
                self.r_dim.setChecked(False)
                self.r_carp.setChecked(False)
                self.r_elab.setChecked(False)
                self.r_pall.setChecked(False)

                self.tableWidget_2.setColumnWidth(0,100)
                self.tableWidget_2.setColumnWidth(1,650)
                self.tableWidget_2.setColumnWidth(2,100)

                #------- DATOS PARA VINCULAR VALE DESPACHO A VINCULACIONES DEL DOCUMENTO ------------------#
                self.aux_doc_venta = {} # Se limpia el contenido.
                self.aux_doc_venta["interno"] = nro_interno
                self.aux_doc_venta["tipo_doc"] = aux_tipo_doc
                self.btn_ver_vale_despacho.hide()
                self.btn_vale_despacho.show()
                #------- se rellena la ventana ------------------#
                try:
                    if aux_tipo_doc != 'GUIA' :
                        items = self.conexion.root.obtener_item_interno(nro_interno)
                        venta = self.conexion.root.obtener_venta_interno(nro_interno) #v5 nombre obtenido

                        self.aux_doc_venta["vinculaciones"] = venta[6]
                        if venta[6]: #Se verifica existencia de vinculaciones.
                            vinculaciones = json.loads(venta[6])
                            print("vinculaciones: ", vinculaciones)
                            try:
                                if vinculaciones['ordenes']:
                                    extra = Extra(vinculaciones,1 ,self)
                                    if extra.exec():
                                        print('|-- Dialog aceptado ... Se continua')
                                        pass
                                    else:
                                        print('|-- Dialog no aceptado ... se anula vista crear orden.')
                                        return
                            except KeyError:
                                print('|-- DOC SIN VINCULO(key) A ORDENES ... Continuando a vista crear orden...')
                                pass
                            # SE VERIFICA VALE DESPACHO ASOCIADO.
                            try:
                                if vinculaciones['vale_id']:
                                    self.aux_doc_venta["vale_id"] = vinculaciones['vale_id']
                                    print(f"(DOC_VENTA): TIENE VALE ASOCIADO -> {vinculaciones['vale_id']}.")
                                    self.btn_ver_vale_despacho.show()
                                    self.btn_vale_despacho.hide()
                                    
                            except KeyError:
                                print("(DOC_VENTA): SIN VALE ASOCIADO.")

                        aux = datetime.fromisoformat(str(venta[3] ) )
                        self.fecha_venta = aux
                        self.items = items
                        self.lb_fecha.setText(str(aux.strftime("%d-%m-%Y %H:%M:%S")))
                        if venta[1] == 0:
                            self.tipo_doc = 'BOLETA'
                            self.lb_doc.setText( self.tipo_doc )
                            self.lb_documento.setText(str(venta[2]))
                            self.nro_doc = venta[2]
                            
                        elif venta[2] == 0:
                            self.tipo_doc = 'FACTURA'
                            self.lb_doc.setText( self.tipo_doc )
                            self.lb_documento.setText(str(venta[1]))
                            self.nro_doc = venta[1]
               
                        self.vendedor = venta[4]  #VENDEDOR
                        if venta[5]:
                            self.nombre_2.setText(venta[5]) #nombre del cliente factura
                        

                    else:
                        guia = self.conexion.root.obtener_guia_interno(nro_interno)
                        self.aux_doc_venta["vinculaciones"] = guia[6]
                        if guia[6]:
                            vinculaciones = json.loads(guia[6])
                            print("vinculaciones: ", vinculaciones)
                            try:
                                if vinculaciones['ordenes']:
                                    extra = Extra(vinculaciones,1 ,self)
                                    if extra.exec():
                                        print('|-- Dialog aceptado ... Se continua')
                                        pass
                                    else:
                                        print('|-- Dialog no aceptado ... se anula vista crear orden.')
                                        return
                            except KeyError:
                                print('|-- DOC SIN VINCULO(key) A ORDENES ... Continuando a vista crear orden...')
                                pass
                            # SE VERIFICA VALE DESPACHO ASOCIADO.
                            try:
                                if vinculaciones['vale_id']:
                                    self.aux_doc_venta["vale_id"] = vinculaciones['vale_id']
                                    print(f"(DOC_VENTA): TIENE VALE ASOCIADO -> {vinculaciones['vale_id']}.")
                                    self.btn_ver_vale_despacho.show()
                                    self.btn_vale_despacho.hide()
                                    
                            except KeyError:
                                print("(DOC_VENTA): SIN VALE ASOCIADO.")
                                self.btn_ver_vale_despacho.hide()
                                self.btn_vale_despacho.show()

                        #print(guia)
                        detalle = json.loads(guia[2])
                        aux = datetime.fromisoformat(str(guia[4] ) ) #fecha de venta
                        self.fecha_venta = aux
                        self.lb_fecha.setText(str(aux.strftime("%d-%m-%Y %H:%M:%S")))
                        self.tipo_doc = aux_tipo_doc
                        self.lb_doc.setText( aux_tipo_doc )
                        self.lb_documento.setText(str(guia[0]))
                        self.nro_doc = guia[0]
                        self.nombre_2.setText(guia[3]) #nombre del cliente factura
                        self.vendedor = detalle['vendedor']  #VENDEDOR
                        i = 0
                        items = []
                        descripciones = detalle['descripciones']
                        cantidades = detalle['cantidades']
                        netos_totales = detalle['totales']
                        while i < len(cantidades) :
                            descr = descripciones[i]
                            item = (cantidades[i] , descr.strip() ,netos_totales[i]) # v5,9 REMOVIENDO espacios no necesarios.
                            items.append(item)
                            i += 1
                        self.items = items
                    print(items)
                    self.rellenar()
                    #----- new -----
                    datos_cliente = self.conexion.root.obtener_cliente(rut)
                    print(f"Rut: {rut} -> Datos cliente: {datos_cliente}")
                    if datos_cliente:
                        print('cliente encontrado --> se rellenan sus datos')
                        index = None
                        self.nombre_2.setText(datos_cliente[1]) #nombre del cliente
                        self.telefono_2.setText(datos_cliente[3]) #telefono  del cliente 
                        self.contacto_2.setText(datos_cliente[4]) #correo (o contacto) del cliente 

                        if datos_cliente[1] in self.nombres:
                            index = self.nombres.index(datos_cliente[1])
                            if datos_cliente[3] == '' and index:
                                telefono = self.telefonos[index]
                                print(f'telefono cliente: {datos_cliente[3]} | telefono orden: {telefono}')
                                self.telefono_2.setText(str(telefono)) #telefono  del cliente 
                            if datos_cliente[4] == '' and index:
                                contacto = self.contactos[index]
                                print(f'contacto cliente: {datos_cliente[4]} | contacto orden: {contacto}')
                                self.contacto_2.setText(contacto) #correo (o contacto) del cliente 
                            
                        else:
                            print('Cliente no encontrado en ordenes.')

                    
                    self.stackedWidget.setCurrentWidget(self.crear_orden)
                    #------------
                    #self.tabla_ordenes = self.conexion.root.obtener_clientes_de_ordenes('dimensionado')
                    #print('cantidad de datos: ' + str(len(self.tabla_ordenes)))
                    #ordenes = pd.DataFrame(self.tabla_ordenes)
                    #nombres = ordenes[1].tolist()
                    #print(ordenes.head())
                    #print(nombres[7])
                    #completer = QCompleter( nombres )
                    #self.nombre_2.setCompleter(completer)
                    

                except EOFError:
                    QMessageBox.about(self, 'ERROR', 'Se perdio la conexion con el servidor')
                
        else:
            QMessageBox.about(self,'ERROR', 'Seleccione una Fila antes de continuar')

     
    def rellenar(self):
        if self.items != ():
            self.tableWidget_2.setRowCount(0)
            for item in self.items:
                fila = self.tableWidget_2.rowCount()
                self.tableWidget_2.insertRow(fila)
                self.tableWidget_2.setItem(fila , 0 , QTableWidgetItem(str(item[0])) ) #CANTIDAD
                self.tableWidget_2.setItem(fila , 1 , QTableWidgetItem(str(item[1])) )  #DESCRIPCION
                if self.tipo_doc == 'BOLETA':
                    neto = ( item[2] / 1.19 )
                    neto = round(neto,2)  #maximo 2 decimales  continuar modificando el json de crear orden trabajo
                    self.tableWidget_2.setItem(fila , 2 , QTableWidgetItem( str(neto) ) )  #VALOR NETO BOLETAS VIENEN CON IVA 
                else: # GUIAS Y FACTURAS NETOS VIENEN SI IVA
                    self.tableWidget_2.setItem(fila , 2 , QTableWidgetItem(str(item[2] )) )  #VALOR NETO FACTURAS VIENEN SIN IVA

    def agregar(self):
        if self.tableWidget_2.rowCount() <=14 :
            fila = self.tableWidget_2.rowCount()
            self.tableWidget_2.insertRow(fila)
        else:
            QMessageBox.about(self, 'ERROR', 'Ha alcanzado el limite maximo de filas. Intente crear otra Orden para continuar agregando items.')

    def eliminar(self):
        fila = self.tableWidget_2.currentRow()  #FILA SELECCIONADA , retorna -1 si no se selecciona una fila
        if fila != -1:
            #print('Eliminando la fila ' + str(fila))
            self.tableWidget_2.removeRow(fila)
        else: 
            QMessageBox.about(self,'Consejo', 'Seleccione una fila para eliminar')

    def rellenar_datos_cliente(self):
        print('------- QCOMPLETE ACTIVADO --------------')
        nombre_cliente = self.nombre_2.text()
        print('Nombre: ' + nombre_cliente + ' | LEN: ' + str(len(nombre_cliente)) )
        try:
            index = self.nombres.index(nombre_cliente)
        except ValueError:
            index = -1

        if index >= 0:
            #nombre_cliente = nombre_cliente.rstrip()
            #print('Nombre: ' + nombre_cliente + ' | LEN: ' + str(len(nombre_cliente)) )
            self.telefono_2.setText(str(self.telefonos[index]))
            self.contacto_2.setText(str(self.contactos[index]))
        print('------- QCOMPLETE FIN --------------')
     

    def registrar_vale_despacho(self,data):
        print("|-- Visualizando VISTA CREAR VALE DESPACHO ...")
        if not data:
            print("|-- Datos extras no obtenidos --> Sin Orden de trabajo.")
            data = {
                    "nombre" : self.nombre_2.text().upper(),
                    "telefono" : str(self.telefono_2.text()) ,
                    "contacto" : self.contacto_2.text().upper()
                }

        config = {
            "model_nombres" : self.nombres,
            "aux_params" : data
        }
        dialog = ValeDespacho(config, self.conexion ,self )
        if dialog.exec():
            print("Dialog ACEPTADO.")
            datos = dialog.obtener_datos_validos()
            datos["vendedor"] = self.datos_usuario[8]
            datos["interno"] = self.inter
            print("Datos: ",datos)
            
            print("(END)")
            return 
        
            if self.conexion:
                try:
                    vale_id = self.conexion.root.registrar_vale_despacho(datos)
                    if vale_id:
                        # SE ASIGNA DESPACHO 'SI' A DOC VENTA.
                        try:
                            if self.conexion.root.actualizar_despacho(self.aux_doc_venta["tipo_doc"], self.aux_doc_venta["interno"], 'SI'):
                                print('ASIGNACION DEL DESPACHO EXITOSO')
                            else:
                                print('ASIGNACION DEL DESPACHO ERRONEA, intentelo mas tarde o contacte al soporte')
                                QMessageBox.warning(self,'PELIGRO', 'ASIGNACION DEL DESPACHO ERRONEA, intentelo mas tarde o contacte al soporte')
                        except EOFError:
                            self.conexion_perdida()

                        print("Vale creado exitosamente. id: ", vale_id)
                        print("(AUX_DOC_VENTA): ",self.aux_doc_venta)
                        try:
                            if self.aux_doc_venta["vinculaciones"]:
                                vinculaciones = json.loads(self.aux_doc_venta["vinculaciones"])
                            else:
                                vinculaciones = {}
                            print("VINCULACIONES ACTUAL: ",vinculaciones)
                            vinculaciones["vale_id"] = vale_id 
                            print("VINCULACIONES DESPUES: ",vinculaciones)
                            vinculaciones = json.dumps(vinculaciones)
                            self.aux_doc_venta["vinculaciones"] = vinculaciones

                            self.conexion.root.actualizar_vinculaciones(self.aux_doc_venta)



                        except KeyError:
                            print("Error de clave, no se encontro detalle en auxdocventa.")
                        """ datos.pop('interno')
                        datos.pop('vendedor')
                        Imagen.crear_vale_despacho(datos) """
                        

                        boton = QMessageBox.question(self,
                        "Vale registrado",
                        f"Se genero el vale Nro {vale_id} exitosamente.\nDesea imprimir el vale desspacho?")
                        if boton == QMessageBox.Yes:
                            #self.ver_pdf(datos_pdf["tipo"])
                            self.imprimir_voucher(vale_id)

                    else:
                        print("error al crear vale.")
                        QMessageBox.about(self,"ERROR",f"Se genero un error al generar el vale.")
                    
                except EOFError:
                    self.conexion_perdida()

            else:
                self.conexion_perdida()
        
        print("|-- FIN VISTA CREAR VALE DESPACHO. --")

    def visualizar_vale_despacho(self):
        vale_id = self.aux_doc_venta["vale_id"]
        print(f"(VISUALIZANDO): VALE DESPACHO ID: {vale_id}")

        resultado = self.conexion.root.obtener_vale_despacho_x_id(vale_id)
        params = dict(
        nombre = resultado[0],
        direccion = resultado[1],
        referencia = resultado[2],
        telefono = resultado[3],
        contacto = resultado[4],
        fecha_estimada = str(resultado[5])
        )
        Imagen.crear_voucher(params)

        dialog = VistaPrevia("imprimir", self)
        if dialog.exec():
            print("VISTA PREVIA ACEPTADA")
            self.imprimir_voucher(vale_id=vale_id)
            
        print("(VISUALIZANDO): FIN. ")

    def imprimir_voucher(self,vale_id):
        print(" ******** IMPRIMIENDO VALEEE ********")

        if self.datos_impresion["imprimir_voucher_en"] == "servidor":
            print("|-- IMPRIMIENDO VOUCHER EN EL SERVIDOR...")
            self.conexion.root.imprimir_vale_despacho(vale_id)
        elif self.datos_impresion["imprimir_voucher_en"] == "local":
            print("|-- IMPRIMIENDO VOUCHER LOCALMENTE ...")
            status, mensaje = Imagen.imprimir_voucher(self.datos_impresion["nombre_impresora_voucher"])
            QMessageBox.information(self, "IMPRESION", mensaje)
            
        self.tableWidget_1.setRowCount(0)
        self.stackedWidget.setCurrentWidget(self.buscar_venta)
        print("**************** FIN IMPRESION ****************")

    def registrar_orden(self):
        print('|-- INICIANDO REGISTRO DE ORDEN DE TRABAJO ...')
        nombre = self.nombre_2.text().upper() #v5.9
        telefono = self.telefono_2.text()
        cont = self.contacto_2.text().upper() #5.9
        oce = self.oce_2.text()
        fecha_orden = datetime.now().date() #FECHA DE CREACION DE ORDEN

        fecha = self.fecha.date()  #FECHA ESTIMADA
        fecha = fecha.toPyDate()
        lineas_totales = 0
        if nombre != '':
            if telefono != '':
                try:
                    telefono = int(telefono)
                    #continuar testeando con el analisis.py
                    cant = self.tableWidget_2.rowCount()
                    print(f'cantidad de datos: {cant}')
                    vacias = False
                    correcto = True
                    cantidades = []
                    descripciones = []
                    valores_neto = []
                    i = 0
                    while i< cant:
                        cantidad = self.tableWidget_2.item(i,0) #Collumna cantidades
                        descripcion = self.tableWidget_2.item(i,1) #Columna descripcion
                        neto = self.tableWidget_2.item(i,2) #Columna de valor neto
                        if cantidad != None and descripcion != None and neto != None :  #Checkea si se creo una fila, esta no este vacia.
                            if cantidad.text() != '' and descripcion.text() != '' and neto.text() != '' :  #Chekea si se modifico una fila, esta no este vacia
                                try: 
                                    nueva_cant = cantidad.text().replace(',','.',3)
                                    nuevo_neto = neto.text().replace(',','.',3)
                                    cantidades.append( float(nueva_cant) )
                                    descripciones.append( (descripcion.text()).upper() ) #v5.9
                                    
                                    lineas = self.separar(descripcion.text())
                                    lineas_totales = lineas_totales + len(lineas)
                                    valores_neto.append(float(nuevo_neto))

                                except ValueError:
                                    correcto = False
                            else:
                                vacias=True
                        else:
                            vacias = True
                        i+=1
                    print(descripciones)
                    if vacias:
                        QMessageBox.about(self, 'Alerta' ,'Una fila y/o columna esta vacia, rellenela para continuar' )
                    elif lineas_totales > 14:
                        QMessageBox.about(self, 'Alerta' ,'Filas totales: '+str(lineas_totales) + ' - El maximo soportado por el formato de la orden es de 14 filas.' )
                    elif correcto == False:
                        QMessageBox.about(self,'Alerta', 'Se encontro un error en una de las cantidades o Valores neto ingresados. Solo ingrese numeros en dichos campos')
                    else:
                        formato = {
                            "cantidades" : cantidades,
                            "descripciones" : descripciones,
                            "valores_neto": valores_neto,
                            "creado_por" : self.datos_usuario[8]
                        }
                        detalle = json.dumps(formato)
                        try:
                            despacho = 'NO'
                            enchape = 'NO'
                            datos_pdf = { "tipo": "" , "facturar" : "NO" }
                            watermarks = []
                            self.nro_orden = 0
                            # v6.07 Posible cambio a despachos.. 
                            # Solamente el vale despacho puede modificar el estado despacho de una nota_venta.
                            if self.r_despacho.isChecked():
                                despacho = 'SI'
                                watermarks.append('despacho') #v5.9
                                """ if self.conexion.root.actualizar_despacho(self.tipo_doc, self.inter , despacho):
                                    print('ORDEN CON DESPACHO -> CONFIRMA NOTA VENTA con DESPACHO.')
                                else:
                                    print('ORDEN CON DESPACHO ->Erro al actualizar DESPACHO A NOTA VENTA') """

                            # REGISTRO DE ORDEN DE TRABAJO
                            if self.r_dim.isChecked():
                                if self.r_enchape.isChecked():
                                    fecha = fecha + timedelta(days=2)
                                    enchape = 'SI'
                                self.nro_orden = self.conexion.root.registrar_orden_dimensionado( self.inter , str(self.fecha_venta), nombre , telefono, str(fecha) , detalle, self.tipo_doc, self.nro_doc,enchape,despacho,str(fecha_orden),cont,oce,self.vendedor)
                                datos_pdf["tipo"] = "dimensionado"
                                
                            elif self.r_elab.isChecked():
                                self.nro_orden = self.conexion.root.registrar_orden_general("elaboracion", nombre,telefono,str(fecha_orden), str(fecha),self.nro_doc,self.tipo_doc,cont,oce, despacho, self.inter ,detalle, str(self.fecha_venta), self.vendedor)                    
                                datos_pdf["tipo"] = "elaboracion"
                            elif self.r_carp.isChecked():
                                self.nro_orden =  self.conexion.root.registrar_orden_general("carpinteria",  nombre,telefono,str(fecha_orden), str(fecha),self.nro_doc,self.tipo_doc,cont,oce, despacho, self.inter ,detalle, str(self.fecha_venta), self.vendedor)
                                datos_pdf["tipo"] = "carpinteria"
                            elif self.r_pall.isChecked():
                                self.nro_orden =  self.conexion.root.registrar_orden_general("pallets",  nombre,telefono,str(fecha_orden), str(fecha),self.nro_doc,self.tipo_doc,cont,oce, despacho, self.inter ,detalle, str(self.fecha_venta), self.vendedor)
                                datos_pdf["tipo"] = "pallets"
                            #v6.0
                            elif self.r_sin_trans.isChecked():
                                self.nro_orden =  self.conexion.root.registrar_orden_general("sin_transformacion",  nombre,telefono,str(fecha_orden), str(fecha),self.nro_doc,self.tipo_doc,cont,oce, despacho, self.inter ,detalle, str(self.fecha_venta), self.vendedor)
                                datos_pdf["tipo"] = "sin_transformacion"
                                watermarks.append('material') 
                            else:
                                QMessageBox.about(self, 'ERROR', 'Seleccione un tipo de orden a generar, antes de proceder a registrar')    

                            if self.nro_orden != 0:
                                self.crear_vinculo(datos_pdf["tipo"])
                                print(f'Creando PDF {datos_pdf["tipo"]} ...')
                                datos = ( str(self.nro_orden) , str(fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones, enchape, cont,oce,self.vendedor)
                                self.crear_pdf(datos, datos_pdf ,watermarks)
                                
                                boton = QMessageBox.question(self, f'Orden de {datos_pdf["tipo"]} registrada correctamente', 'Desea abrir la Orden?')
                                if boton == QMessageBox.Yes:
                                    self.ver_pdf(datos_pdf["tipo"])
                                
                                if self.r_despacho.isChecked():
                                    print("|-- CHECK Despacho domicilio Detectado ...")
                                    print("|-- Procediendo a crear vale despacho manteniendo datos de orden...")

                                    data = {
                                        "nombre" : nombre,
                                        "telefono" : telefono,
                                        "contacto" : cont
                                    }
                                    self.registrar_vale_despacho(data)

                                    """ nombre 
                                    telefono 
                                    cont  """
                            print("|-- FIN REGISTRAR ORDEN DE TRABAJO.")
                            self.inicializar_buscar_venta()

                            
                        except EOFError:
                            self.conexion_perdida()
                        except AttributeError:
                            QMessageBox.about(self,'IMPORTANTE', 'Este mensaje se debe a que hubo un error al ingresar los datos a la base de datos. Contacte con el soporte')
                except ValueError:
                    QMessageBox.about(self, 'ERROR', 'Solo ingrese numeros en el campo "Telefono" ')          
            else:
                QMessageBox.about(self, 'Sugerencia', 'Ingrese un telefono antes de continuar')           
        else:
            QMessageBox.about(self, 'Sugerencia', 'Ingrese un nombre antes de continuar')
        
# ------ Funciones de VISTA buscar orden de trabajo -------------
    def inicializar_buscar_orden(self):
        self.anterior = None #SABER SI VUELVE A ESTADISTICAS O AQUI
        self.anterior2 = None #SABER SI VUELVE A GESTION REINGRESOS O AKI
        self.r_facturar_1.setChecked(False) # ASI NO MUESTRA EL "POR FACTURAR" 
        self.txt_orden.setText('')
        self.txt_cliente.setText('')
        self.dateEdit.setDate(datetime.now().date())
        self.dateEdit.setCalendarPopup(True)
        self.r_fecha.setChecked(True)
        self.tb_buscar_orden.setRowCount(0)

        self.tb_buscar_orden.setColumnWidth(0,100)
        self.tb_buscar_orden.setColumnWidth(1,100)
        self.tb_buscar_orden.setColumnWidth(2,120)
        self.tb_buscar_orden.setColumnWidth(3,200)
        self.tb_buscar_orden.setColumnWidth(4,120)
        self.tb_buscar_orden.setColumnWidth(5,120)
        self.tb_buscar_orden.setColumnWidth(6,100)

        self.stackedWidget.setCurrentWidget(self.buscar_orden)

    def buscar_dimensionado(self):
        self.ch_nulas.setChecked(False)
        self.lb_tipo_orden.setText('DIMENSIONADO')
        estado = 'Ninguno'
        if self.conexion:
            self.tb_buscar_orden.setRowCount(0)
            if self.r_orden.isChecked(): #Busqueda por numero interno
                try:
                    orden = int(self.txt_orden.text())
                    consulta = self.conexion.root.buscar_orden_dim_numero(orden)
                    if consulta != None :
                        items = []
                        if consulta[19]:
                            estado = 'ANULADA'
                        else:
                            estado = 'VALIDA'
                        #campos: (nro_orden, nro_interno,fecha_creacion,nombre_cliente,fecha_venta,fecha_estimada,estado del doc)
                        item = (str(consulta[0]), str(consulta[1]),  str(consulta[11]), consulta[3] , str(consulta[2]) ,str(consulta[5]), estado )
                        items.append(item)
                        self.mostrar_en_tabla(items)
                    else:
                        QMessageBox.about(self,'Busqueda' ,'Orden de dimensionado NO encontrada')
                except ValueError:
                    QMessageBox.about(self,'ERROR' ,'Ingrese solo numeros')
                except EOFError:
                    self.conexion_perdida()

            elif self.r_fecha.isChecked():
                date = self.dateEdit.date()
                aux = date.toPyDate() 
                try:
                    datos = self.conexion.root.buscar_orden_dim_fecha( str(aux) )
                    if datos != ():
                        items = []
                        for dato in datos:
                            if dato[6]:
                                estado = 'ANULADA'
                            else:
                                estado = 'VALIDA'
                            #campos: (nro_orden, nro_interno,fecha_creacion,nombre_cliente,fecha_venta,fecha_estimada,estado del doc)
                            item = (str(dato[0]), str(dato[1]),  str(dato[4]), dato[3] , str(dato[2]) ,str(dato[5]), estado )
                            items.append(item)
                        self.mostrar_en_tabla(items)
                    else:
                        QMessageBox.about(self ,'Resultado', 'No se encontraron Ordenes de dimensionado')
                except EOFError:
                    self.conexion_perdida()
            elif self.r_cliente.isChecked():
                nombre_cliente = self.txt_cliente.text()
                print(nombre_cliente)
                try:
                    datos = self.conexion.root.buscar_orden_nombre('dimensionado', nombre_cliente)
                    if datos != ():
                        self.mostrar_en_tabla(datos)
                    else:
                        QMessageBox.about(self ,'Resultado', 'No se encontraron Ordenes de dimensionado coincidentes')
                except EOFError:
                    self.conexion_perdida()
        else:
            self.conexion_perdida()

    def busqueda_general(self,tipo):
        self.ch_nulas.setChecked(False)
        self.lb_tipo_orden.setText(tipo)
        if self.conexion:
            self.tb_buscar_orden.setRowCount(0)
            if self.r_orden.isChecked(): #Busqueda por numero de orden
                try:
                    orden = int(self.txt_orden.text())
                    #BUSCA ORDEN GENERAL (elab,carp,pall,sin_trans)
                    consulta = self.conexion.root.buscar_orden_general_numero( tipo.lower() ,orden)

                    items = []
                    estado = 'Ninguno'
                    if consulta != None :
                        if consulta[16]:
                            estado = 'ANULADA'
                        else:
                            estado = 'VALIDA'
                        #campos: (nro_orden, nro_interno,fecha_creacion,nombre_cliente,fecha_venta,fecha_estimada,estado del doc)
                        item = (str(consulta[0]), str(consulta[10]),  str(consulta[3]), consulta[1] , str(consulta[12]) ,str(consulta[4]), estado )
                        items.append(item)
                        self.mostrar_en_tabla(items)
                    else:
                        QMessageBox.about(self,'Busqueda' ,'Orden de: '+ tipo+' NO encontrada')
                except ValueError:
                    QMessageBox.about(self,'ERROR' ,'Ingrese solo numeros')
                except EOFError:
                    self.conexion_perdida()

            elif self.r_fecha.isChecked(): #busqueda por fecha
                date = self.dateEdit.date()
                aux = date.toPyDate()
                #inicio = str(aux) + ' ' + '00:00:00'
                #fin = str(aux) + ' ' + '23:59:59'    
                try:
                    # Busca ordenes generales(elab,carp,pall,sin_trans)
                    datos = self.conexion.root.buscar_orden_general_fecha( tipo.lower() , str(aux))

                    items = []
                    estado = 'Ninguno'
                    if datos != ():
                        for dato in datos:
                            if dato[6]:
                                estado = 'ANULADA'
                            else:
                                estado = 'VALIDA'
                            #campos: (nro_orden, nro_interno,fecha_creacion,nombre_cliente,fecha_venta,fecha_estimada,estado del doc)
                            item = (str(dato[0]), str(dato[1]),  str(dato[2]), dato[3] , str(dato[4]) ,str(dato[5]), estado )
                            items.append(item)
                        self.mostrar_en_tabla(items)

                    else:
                        QMessageBox.about(self ,'Resultado', 'No se encontraron Ordenes de pallets')
                except EOFError:
                    self.conexion_perdida()

            elif self.r_cliente.isChecked():
                nombre_cliente = self.txt_cliente.text()
                try:
                    datos = self.conexion.root.buscar_orden_nombre( tipo.lower() , nombre_cliente) # SQLI detectada, ejemplo --> nombre = hola%' or 1=1 -- 
                    items = []
                    estado = 'Ninguno'
                    if datos != ():
                        for dato in datos:
                            if dato[6]:
                                estado = 'ANULADA'
                            else:
                                estado = 'VALIDA'
                            #campos: (nro_orden, nro_interno,fecha_creacion,nombre_cliente,fecha_venta,fecha_estimada,estado del doc)
                            item = (str(dato[0]), str(dato[1]),  str(dato[2]), dato[3] , str(dato[4]) ,str(dato[5]), estado )
                            items.append(item)
                        self.mostrar_en_tabla(items)
                    else:
                        QMessageBox.about(self ,'Resultado', 'No se encontraron Ordenes de: '+ tipo +' coincidentes')
                except EOFError:
                    self.conexion_perdida()
        else:
            self.conexion_perdida()

    def mostrar_en_tabla(self,datos):
        for dato in datos:
            fila = self.tb_buscar_orden.rowCount()
            self.tb_buscar_orden.insertRow(fila)
            self.tb_buscar_orden.setItem(fila , 0 , QTableWidgetItem( str(dato[0])) )  #nro orden
            self.tb_buscar_orden.setItem(fila , 1 , QTableWidgetItem( str(dato[1]) ))  #nro interno
            self.tb_buscar_orden.setItem(fila , 2 , QTableWidgetItem( str(dato[2]) ) ) #fecha creacion
            self.tb_buscar_orden.setItem(fila , 3 , QTableWidgetItem(dato[3]))       #nombre
            self.tb_buscar_orden.setItem(fila , 4 , QTableWidgetItem( str(dato[4]) ) )       #fecha venta
            self.tb_buscar_orden.setItem(fila , 5 , QTableWidgetItem( str(dato[5]) ) )       #fecha estimada
            self.tb_buscar_orden.setItem(fila , 6 , QTableWidgetItem(dato[6]) )       #estado
            
#------ Funciones para MODIFICAR ORDEN ----------- 

    def inicializar_modificar_orden(self):
        self.limpiar_varibles()
        self.txt_vendedor_5.setText('')
        self.contacto_5.setText('') 
        self.oce_5.setText('')
        self.txt_interno_5.setEnabled(True) #solo lectura
        self.txt_nro_doc_5.setEnabled(True)
        self.date_venta_5.setEnabled(True)
        self.comboBox_5.setEnabled(True)
        self.txt_vendedor_5.setEnabled(True)
        self.tipo_doc = None 
        self.txt_obs_2.clear() #v5.9

        seleccion = self.tb_buscar_orden.currentRow()
        if seleccion != -1:
            _item = self.tb_buscar_orden.item( seleccion, 0) 
            if _item:            
                orden = self.tb_buscar_orden.item(seleccion, 0 ).text()
                self.nro_orden = int(orden)
                tipo = self.lb_tipo_orden.text()
                print('----------- MODIFICAR ORDEN ... para: '+ tipo + ' -->' + str(self.nro_orden) + ' -------------')
                self.datos_orden["tipo_orden"] = tipo
                self.datos_orden["nro_orden"] = self.nro_orden

                self.tb_modificar_orden.setColumnWidth(0,100)
                self.tb_modificar_orden.setColumnWidth(1,650)
                self.tb_modificar_orden.setColumnWidth(2,100)
                #v5.9 se habilitan todos los checkboxes primero.
                self.r_despacho_5.setEnabled(True)
                self.r_enchape_5.setEnabled(True)
                self.r_facturar_5.setEnabled(True)
                self.r_separar_material_5.setEnabled(True)
                self.r_uso_interno_5.setEnabled(True)
                #v5.9 se desmarcan todos los checkboxes.
                self.r_despacho_5.setChecked(False)
                self.r_enchape_5.setChecked(False)
                self.r_facturar_5.setChecked(False)
                self.r_separar_material_5.setChecked(False)
                self.r_uso_interno_5.setChecked(False)

                self.rellenar_datos_orden(tipo)
                self.stackedWidget.setCurrentWidget(self.modificar_orden)
        
        else:
            QMessageBox.about(self,'ERROR', 'Seleccione un Nro interno antes de continuar')

    def rellenar_datos_orden(self,tipo):
        self.r_despacho_5.setChecked(False)
        self.date_venta_5.setCalendarPopup(True)
        self.fecha_5.setCalendarPopup(True)
        cantidades = None
        descripciones = None
        enchapado = 'NO'
        despacho = 'NO'
        self.date_venta_5.setDate( datetime.now())
        self.tipo = tipo # area de orden de trabajo
        self.tb_modificar_orden.setRowCount(0)
        self.comboBox_5.clear()
        self.manual = False
        self.aux_obs = ''
        self.r_uso_interno_5.show() #v5.9
        self.r_separar_material_5.show() #v5.9
        self.r_facturar_5.show() #v5.9
        self.frame_obs_2.show() # v5.9 Frame: campo de observacion
        datos_pdf = { "tipo" : tipo.lower() , 'facturar': 'NO'}
        watermarks = []

        if tipo == 'DIMENSIONADO':
            try:
                resultado = self.conexion.root.buscar_orden_dim_numero(self.nro_orden)
                if resultado != None :

                    if resultado[18]: #si es manual
                        print('ORDEN MANUAL DETECTADA')
                        self.aux_obs = resultado[ 18 ] #v5.9 respalda la observacion
                        self.manual = True
                        if resultado[8]:
                            self.txt_nro_doc_5.setText( str(resultado[8]) )       #NUMERO DOCUMENTO
                            self.nro_doc = int( resultado[8] )

                        else:
                            self.txt_nro_doc_5.setText( '0' )       #NUMERO DOCUMENTO

                        if resultado[7]:
                            if resultado[7] == 'BOLETA':
                                self.comboBox_5.addItem( resultado[7] )                 #TIPO DOCUMENTO
                                self.comboBox_5.addItem('FACTURA')
                                self.comboBox_5.addItem('GUIA')
                                self.tipo_doc = resultado[7]

                            elif resultado[7] == 'FACTURA':
                                self.comboBox_5.addItem( resultado[7] )                 #TIPO DOCUMENTO
                                self.comboBox_5.addItem('BOLETA')
                                self.comboBox_5.addItem('GUIA')
                                self.tipo_doc = resultado[7]
                            elif resultado[7] == 'GUIA':
                                self.comboBox_5.addItem(resultado[7])
                                self.comboBox_5.addItem('FACTURA')  
                                self.comboBox_5.addItem('BOLETA')
                                self.tipo_doc = resultado[7]
                            elif resultado[7] == 'NO ASIGNADO':
                                self.comboBox_5.addItem('NO ASIGNADO') 
                                self.comboBox_5.addItem('FACTURA')  
                                self.comboBox_5.addItem('BOLETA')
                                self.comboBox_5.addItem('GUIA')
                        else:
                            print('no existe su tipo doc') # else compatible con la version antgua donde tipo doc era null, actualmente tipo doc es 'no asignado'
                            self.comboBox_5.addItem('NO ASIGNADO') 
                            self.comboBox_5.addItem('FACTURA')  
                            self.comboBox_5.addItem('BOLETA')
                            self.comboBox_5.addItem('GUIA')
                        if resultado[2]:
                            aux3 = datetime.fromisoformat(str(resultado[2]) )
                            self.date_venta_5.setDate( aux3 )   #FECHA DE VENTA
                        if resultado[15]:
                            self.vendedor = resultado[15]   #VENDEDOR
                            self.txt_vendedor_5.setText( resultado[15] ) 

                        if resultado[18] == 'USO INTERNO':
                            self.r_uso_interno_5.setChecked(True)
                            watermarks.append('interno')
                        elif resultado[18] == 'SEPARAR MATERIAL':
                            self.r_separar_material_5.setChecked(True)
                            watermarks.append('material')
                        elif resultado[18] == 'USO INTERNO Y SEPARAR MATERIAL':
                            self.r_uso_interno_5.setChecked(True)
                            self.r_separar_material_5.setChecked(True)
                            watermarks.append('interno')
                            watermarks.append('material')
                        
                              
                    else: #SI NO ES MANUAL
                        print('ORDEN COMÚN DETECTADA')
                        self.manual = False
                        self.txt_interno_5.setEnabled(False) #solo lectura
                        self.txt_nro_doc_5.setEnabled(False)
                        self.date_venta_5.setEnabled(False)
                        self.comboBox_5.setEnabled(False)

                        self.comboBox_5.addItem( resultado[7] )    #tipo documento
                        self.tipo_doc = resultado[7]
                        aux1 = datetime.fromisoformat(str(resultado[2]) )
                        self.date_venta_5.setDate( aux1 )     #fecha de venta
                        self.txt_nro_doc_5.setText( str(resultado[8]) ) #numero de documento
                        self.nro_doc = resultado[8]
                        self.vendedor = resultado[15]    #vendedor
                        self.txt_vendedor_5.setText( resultado[15] ) #vendedor
                        # OCULTAR CHECKBOX: USO INTERNO - FACTURAR - SEPARAR MATERIAL
                        self.r_uso_interno_5.hide()
                        self.r_separar_material_5.hide()
                        self.r_facturar_5.hide()

                        self.frame_obs_2.hide()
                        


                    self.txt_vendedor_5.setEnabled(False)
                    self.txt_interno_5.setText(str( resultado[1] )) #INTERNO
                    self.inter = str( resultado[1] )
                    
                    self.nombre_5.setText( resultado[3] )   #nombre
                    self.telefono_5.setText( str(resultado[4]) ) #telefono
                    aux = datetime.fromisoformat(str(resultado[5]) )
                    self.fecha_5.setDate( aux )  #FECHA ESTIMADA     

                    self.detalle = json.loads(resultado[6])
                    if resultado[20]:
                        self.lb_vinculo_5.setText(resultado[20])
                    else:
                        self.lb_vinculo_5.setText('NO CREADO')

                    vinc = resultado[20]
                    print('Vinculo:')
                    print(vinc)
                    cantidades = self.detalle["cantidades"]
                    descripciones = self.detalle["descripciones"]
                    valores_neto = self.detalle["valores_neto"]
                    try:
                        self.aux_vendedor = self.detalle["creado_por"]
                    except:
                        print('orden antigua, sin creador asignado')
                        pass
                        
                    try:
                        if self.detalle["por_facturar"] == 'SI':
                            self.r_facturar_5.setChecked(True)      #por facturar
                            datos_pdf['facturar'] = 'SI' #v5.9
                    except:
                        print('version inferior a v5.9, por facturar NO registrado.')
                        pass

                    j = 0
                    while j < len( cantidades ):
                        fila = self.tb_modificar_orden.rowCount()
                        self.tb_modificar_orden.insertRow(fila)
                        self.tb_modificar_orden.setItem(fila , 0 , QTableWidgetItem( str( cantidades[j] )) ) 
                        self.tb_modificar_orden.setItem(fila , 1 , QTableWidgetItem( descripciones[j] ) )
                        self.tb_modificar_orden.setItem(fila , 2 , QTableWidgetItem( str( valores_neto[j] )) ) 
                        j+=1      
                    if resultado[9] == 'SI' :
                        self.r_enchape_5.setChecked(True)      #enchape
                        
                        enchapado = 'SI'
                    if resultado[10] == 'SI' :
                        self.r_despacho_5.setChecked(True)  #despacho
                        watermarks.append('despacho')

                    self.r_enchape_5.setVisible(True) #Se visualiza checkbox enchape para dimensionado. 
                    self.lb_fecha_orden_5.setText( str(resultado[11]))
                    self.fecha_orden = datetime.fromisoformat( str(resultado[11]) )  #Fecha orden
                    self.contacto_5.setText( resultado[12] ) #contacto
                    self.oce_5.setText( resultado[13] )      #orden comprar
                    

            except EOFError:
                self.conexion_perdida()

        else: 
            try:
                
                resultado = self.conexion.root.buscar_orden_general_numero( tipo.lower() , self.nro_orden)
                
                if resultado != None :
                    if resultado[ 15 ]: #si es manual
                        print('ORDEN MANUAL DETECTADA')
                        self.aux_obs = resultado[ 15 ] #v5.9 respalda la observacion
                        self.manual = True
                        if resultado[5]:
                            self.txt_nro_doc_5.setText( str(resultado[5]) )       #NUMERO DOCUMENTO
                            self.nro_doc = int( resultado[5] )
                        else:
                            self.txt_nro_doc_5.setText( '0' )       #NUMERO DOCUMENTO

                        if resultado[6]:
                            if resultado[6] == 'BOLETA':
                                self.comboBox_5.addItem( resultado[6] )                 #TIPO DOCUMENTO
                                self.comboBox_5.addItem('FACTURA')
                                self.comboBox_5.addItem('GUIA')
                                self.tipo_doc = resultado[6]

                            elif resultado[6] == 'FACTURA':
                                self.comboBox_5.addItem( resultado[6] )                 #TIPO DOCUMENTO
                                self.comboBox_5.addItem('BOLETA')
                                self.comboBox_5.addItem('GUIA')
                                self.tipo_doc = resultado[6]
                            elif resultado[6] == 'GUIA':
                                self.comboBox_5.addItem(resultado[6])
                                self.comboBox_5.addItem('FACTURA')  
                                self.comboBox_5.addItem('BOLETA')
                                self.tipo_doc = resultado[6]
                            elif resultado[6] == 'NO ASIGNADO':
                                self.comboBox_5.addItem('NO ASIGNADO') 
                                self.comboBox_5.addItem('FACTURA')  
                                self.comboBox_5.addItem('BOLETA')
                                self.comboBox_5.addItem('GUIA')
                        else:
                            print('no existe su tipo doc') # else compatible con la version antgua donde tipo doc era null, actualmente tipo doc es 'no asignado'
                            self.comboBox_5.addItem('NO ASIGNADO') 
                            self.comboBox_5.addItem('FACTURA')  
                            self.comboBox_5.addItem('BOLETA')
                            self.comboBox_5.addItem('GUIA')

                        if resultado[12]:
                            aux3 = datetime.fromisoformat(str(resultado[12]) )
                            self.date_venta_5.setDate( aux3 )   #FECHA DE VENTA
                        if resultado[14]:
                            self.vendedor = resultado[14]   #VENDEDOR
                            self.txt_vendedor_5.setText( resultado[14] )

                        if resultado[15] == 'USO INTERNO':
                            self.r_uso_interno_5.setChecked(True)
                            watermarks.append('interno')
                        elif resultado[15] == 'SEPARAR MATERIAL':
                            self.r_separar_material_5.setChecked(True)
                            watermarks.append('material')
                        elif resultado[15] == 'USO INTERNO Y SEPARAR MATERIAL':
                            self.r_uso_interno_5.setChecked(True)
                            self.r_separar_material_5.setChecked(True)
                            watermarks.append('interno')
                            watermarks.append('material')

                    else: #si no es manual
                        print('ORDEN COMÚN DETECTADA')
                        self.manual = False
                        self.txt_interno_5.setEnabled(False) #solo lectura
                        self.txt_vendedor_5.setEnabled(False)
                        self.txt_nro_doc_5.setEnabled(False)
                        self.date_venta_5.setEnabled(False)
                        self.comboBox_5.setEnabled(False)

                        self.txt_nro_doc_5.setText( str(resultado[5]) )       #NUMERO DOCUMENTO
                        self.nro_doc = resultado[5]
                        self.comboBox_5.addItem( resultado[6] )                 #TIPO DOCUMENTO
                        self.tipo_doc = resultado[6]
                        aux3 = datetime.fromisoformat(str(resultado[12]) )
                        self.date_venta_5.setDate( aux3 )   #FECHA DE VENTA
                        
                        self.vendedor = resultado[14]   #VENDEDOR
                        self.txt_vendedor_5.setText( resultado[14] )
                        # OCULTAR CHECKBOX: USO INTERNO - FACTURAR - SEPARAR MATERIAL
                        self.r_uso_interno_5.hide()
                        self.r_separar_material_5.hide()
                        self.r_facturar_5.hide()
                        self.frame_obs_2.hide() #ocultar frame observacion
                        

                    self.txt_interno_5.setText(str( resultado[10] ))        #NRO INTERNO
                    self.inter = str( resultado[10] )
                    self.r_enchape_5.setVisible(False)  # Se oculta checkbox enchape para ordenes elab,carp y pall
                    self.nombre_5.setText( resultado[1] )         #NOMBRE CLIENTE
                    self.telefono_5.setText( str(resultado[2]) )  #TELEFONO

                    self.fecha_orden = datetime.fromisoformat( str(resultado[3]) )  
                    self.lb_fecha_orden_5.setText( str(resultado[3]))   #FECHA DE ORDEN
                    
                    f_estimada = datetime.fromisoformat(str(resultado[4]) )
                    self.fecha_5.setDate( f_estimada )                   #FECHA ESTIMADA DE ENTREGA

                    
                    self.contacto_5.setText( resultado[7] )             #CONTACTO
                    self.oce_5.setText( resultado[8] )                   #ORDEN DE COMPRA
                              
                    if resultado[9] == 'SI' :
                        self.r_despacho_5.setChecked(True)               #DESPACHO
                        watermarks.append('despacho')  #v5.9 

                    if resultado[19]:
                        self.lb_vinculo_5.setText(resultado[19])
                    else:
                        self.lb_vinculo_5.setText('NO CREADO')

                    vinc = resultado[19]
                    print('Vinculo:')
                    print(vinc)

                    self.detalle = json.loads(resultado[11])                  #DETALLE
                    cantidades = self.detalle["cantidades"]
                    descripciones = self.detalle["descripciones"]
                    valores_neto = self.detalle["valores_neto"]
                    try:
                        self.aux_vendedor = self.detalle["creado_por"]
                    except:
                        print('orden antigua, sin creador asignado')
                        pass
                    
                    try:
                        if self.detalle["por_facturar"] == 'SI':
                            self.r_facturar_5.setChecked(True)      #por facturar
                            datos_pdf['facturar'] = 'SI' #v5.9
                    except:
                        print('version inferior a v5.9, por facturar NO registrado.')
                        pass

                    j = 0
                    while j < len( cantidades ):
                        fila = self.tb_modificar_orden.rowCount()
                        self.tb_modificar_orden.insertRow(fila)
                        self.tb_modificar_orden.setItem(fila , 0 , QTableWidgetItem( str( cantidades[j] )) ) 
                        self.tb_modificar_orden.setItem(fila , 1 , QTableWidgetItem( descripciones[j] ) )
                        self.tb_modificar_orden.setItem(fila , 2 , QTableWidgetItem( str( valores_neto[j] )) ) 
                        j+=1    
            except EOFError:
                self.conexion_perdida()

        # ------ comprobar si existe el PDF en los archivos locales -------------
        if tipo in ["DIMENSIONADO", "ELABORACION", "CARPINTERIA", "PALLETS","SIN_TRANSFORMACION"]:
            tipo_lower = tipo.lower()
            #v6.0
            abrir = os.path.join(self.dir_ordenes , f"{tipo_lower}_{self.nro_orden}.pdf" )
        else:
            # Manejo del caso por defecto o de errores
            abrir = None  # Puedes asignar un valor por defecto o manejar el error de otra manera

        print('|-- Verificando existencia del PDF ....')
        datos = [ str(self.nro_orden) ,str(self.fecha_orden.strftime("%d-%m-%Y")),self.nombre_5.text(),self.telefono_5.text(), str((self.fecha_5.date().toPyDate()).strftime("%d-%m-%Y")),cantidades,descripciones,enchapado ,self.contacto_5.text(),self.oce_5.text() ,self.vendedor ]
        
        

        self.datos_orden["interno"] = self.inter
        self.datos_orden["nro_doc"] = self.nro_doc
        self.datos_orden["tipo_doc"] = self.tipo_doc

        self.datos_orden["nombre"]= self.nombre_5.text()
        self.datos_orden["telefono"] = self.telefono_5.text()
        self.datos_orden["fecha_orden"] = self.lb_fecha_orden_5.text()
        self.datos_orden["fecha_venta"] = self.date_venta_5.date().toString("yyyy-MM-dd")
        self.datos_orden["fecha_estimada"] = self.fecha_5.date().toString("yyyy-MM-dd")
        self.datos_orden["oce"] = self.oce_5.text()

        self.datos_orden["enchape"] = 'NO'
        if self.r_enchape_5.isChecked(): #enchape
            self.datos_orden["despacho"] = 'SI'

        self.datos_orden["despacho"] = 'NO'
        if 'despacho' in watermarks:
            self.datos_orden["despacho"] = 'SI'

        self.datos_orden["vendedor"] = self.txt_vendedor_5.text()
        self.datos_orden["cont"] = self.contacto_5.text()#contacto
        self.datos_orden["detalle"] = self.detalle

        self.datos_orden["datos"] = datos
        self.datos_orden["datos_pdf"] = datos_pdf
        self.datos_orden["watermarks"] = watermarks

        if abrir and (not os.path.isfile(abrir)) :
            #v6.0 
            if tipo == 'SIN_TRANSFORMACION':
                watermarks.append('material')

            self.crear_pdf(datos, datos_pdf ,watermarks)
            print('|--Pdf no encontrado, pero se acaba de crear')
        else:
            print('|-- El pdf si existe localmente')

        print('----------------------------------------------------------------------')
    def agregar_2(self):
        if self.tb_modificar_orden.rowCount() < 14 :
            fila = self.tb_modificar_orden.rowCount()
            self.tb_modificar_orden.insertRow(fila)
        else:
            QMessageBox.about(self, 'ERROR', 'Ha alcanzado el limite maximo de filas. Intente crear otra Orden para continuar agregando items.')

    def eliminar_2(self):
        fila = self.tb_modificar_orden.currentRow()  #FILA SELECCIONADA , retorna -1 si no se selecciona una fila
        if fila != -1:
            #print('Eliminando la fila ' + str(fila))
            self.tb_modificar_orden.removeRow(fila)
        else: 
            QMessageBox.about(self,'Consejo', 'Seleccione una fila para eliminar')
    def actualizar_orden(self):
        nombre = self.nombre_5.text().upper() #v5.9
        telefono = self.telefono_5.text()
        contacto = self.contacto_5.text().upper() #v5.9
        oce = self.oce_5.text()
        fecha = self.fecha_5.date()  #fecha estimada
        fecha = fecha.toPyDate()
        enchapado = 'NO'
        despacho = 'NO'
        #v5.6
        aux_tipo = self.tipo_doc
        aux_nro = self.nro_doc

        self.tipo_doc = self.comboBox_5.currentText()
        self.vendedor = self.txt_vendedor_5.text()
        fecha_venta = str( self.date_venta_5.date().toPyDate() )
        observacion = self.txt_obs_2.toPlainText() #v5.9
        if self.manual:
            print('|-- Actualizando orden manual --> obs no None')
            if observacion == '':
                print('|-- Obs Vacia --> No se puede actualizar orden manual')
                QMessageBox.about(self, 'Alerta' ,'Debe ingresar una observacion. \nDebido a que la orden es manual.' )
                return 
        
        lineas_totales = 0 
        if nombre != '' and telefono != '':
            try:
                telefono = int(telefono)
                self.inter = int( self.txt_interno_5.text() )
                self.nro_doc = int( self.txt_nro_doc_5.text() )
                
                cant = self.tb_modificar_orden.rowCount() #cantidad de items de la tabla
                vacias = False
                correcto = True
                cantidades = []
                descripciones = []
                valores_neto = []
                i = 0
                while i< cant:
                    cantidad = self.tb_modificar_orden.item(i,0) #Collumna cantidades
                    descripcion = self.tb_modificar_orden.item(i,1) #Columna descripcion
                    neto = self.tb_modificar_orden.item(i,2) #Columna de valor neto
                    if cantidad != None and descripcion != None and neto != None :  #Checkea si se creo una fila, esta no este vacia.
                        if cantidad.text() != '' and descripcion.text() != '' and neto.text() != '' :  #Chekea si se modifico una fila, esta no este vacia
                            try: 
                                nueva_cant = cantidad.text().replace(',','.',3)
                                nuevo_neto = neto.text().replace(',','.',3)
                                cantidades.append( float(nueva_cant) )
                                descripciones.append((descripcion.text()).upper() ) #v5.9
                                lineas = self.separar(descripcion.text())
                                lineas_totales = lineas_totales + len(lineas)
                                valores_neto.append(float(nuevo_neto))

                            except ValueError:
                                correcto = False
                        else:
                            vacias=True
                    else:
                        vacias = True
                    i+=1
                if vacias:
                    QMessageBox.about(self, 'Alerta' ,'Una fila y/o columna esta vacia, rellenela para continuar' )
                elif lineas_totales > 14:
                    QMessageBox.about(self, 'Alerta' ,'Filas totales: '+str(lineas_totales) + ' - El maximo soportado por el formato de la orden es de 14 filas.' )
                elif correcto == False:
                    QMessageBox.about(self,'Alerta', 'Se encontro un error en una de las cantidades o Valores neto ingresados. Solo ingrese numeros en dichos campos')
                else:
                        
                    #print('---------------------------------------------------')
                    try:
                        watermarks = [] #v5.9 nuevas marcas de agua
                        datos_pdf = {
                            "tipo" : (self.tipo).lower(),
                            "facturar" : 'NO'
                        }
                        enchapado = 'NO'
                        despacho = 'NO'
                        if self.r_despacho_5.isChecked():
                            despacho = 'SI'
                            watermarks.append('despacho') #v5.9
                        if self.r_facturar_5.isChecked():
                            datos_pdf["facturar"] = 'SI' 
                        if self.r_uso_interno_5.isChecked():
                            watermarks.append('interno') #v5.9
                        if self.r_separar_material_5.isChecked() or (self.tipo == 'SIN_TRANSFORMACION'):
                            watermarks.append('material') #v5.9

                        self.detalle["cantidades"] = cantidades
                        self.detalle["descripciones"] = descripciones
                        self.detalle["valores_neto"] = valores_neto
                        try:
                            self.detalle['por_facturar'] = datos_pdf["facturar"]
                        except KeyError:
                            print('Orden normal no tiene "por_facturar" .keyerror')
                            pass
                        self.detalle = json.dumps(self.detalle)




                        if self.tipo == 'DIMENSIONADO':
                            if self.r_enchape_5.isChecked():
                                enchapado = 'SI'
                            datos = ( str(self.nro_orden) , str(self.fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones,enchapado, contacto,oce, self.vendedor)
                            self.crear_pdf(datos,datos_pdf,watermarks)
                            print('pdf actualizado')
                            if self.conexion.root.actualizar_orden_dim( self.manual,self.inter,fecha_venta,self.tipo_doc,self.nro_doc,self.vendedor, self.nro_orden,nombre,telefono,str(fecha),self.detalle,despacho,enchapado,contacto,oce):
                                if self.manual:
                                    self.conexion.root.actualizar_orden_dim_obser(observacion , self.nro_orden)
                                QMessageBox.about(self,'EXITO','La orden de dimensionado fue ACTUALIZADA correctamente')
                            else:
                                QMessageBox.about(self,'ERROR','La orden de dimensionado NO se actualizo, porque no se modificaron los datos que ya existian.')

                        else:
                            tipo_orden = (self.tipo).lower()
                            datos = ( str(self.nro_orden) , str(self.fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones,enchapado, contacto,oce, self.vendedor)
                            self.crear_pdf(datos,datos_pdf,watermarks)

                            if self.conexion.root.actualizar_orden_general( tipo_orden ,self.manual,self.inter,fecha_venta,self.tipo_doc,self.nro_doc,self.vendedor,  nombre,telefono, str(fecha), self.detalle, contacto, oce,despacho, self.nro_orden ):
                                if self.manual:
                                    self.conexion.root.actualizar_orden_general_obs(tipo_orden ,observacion , self.nro_orden)
                                QMessageBox.about(self,'EXITO',f'La orden de {tipo_orden} - Fue ACTUALIZADA correctamente')
                            else:
                                QMessageBox.about(self,'ERROR',f'La orden de {tipo_orden} - No se actualizo, porque no se modificaron los datos que ya existian.')
                        
                        if self.manual:
                            estado = self.lb_vinculo_5.text()
                            self.actualizar_vinculo_orden_manual(aux_tipo,aux_nro,self.tipo_doc,self.nro_doc,estado)
                        
                        if self.anterior == "ESTADISTICAS": #v5.9
                            self.stackedWidget.setCurrentWidget(self.estadisticas)
                        else:
                            self.stackedWidget.setCurrentWidget(self.buscar_orden)
                    except EOFError:
                        self.conexion_perdida()
                    except PermissionError:
                        QMessageBox.about(self,'ERROR' ,'La orden no se actualizo debido a que otro programa esta haciendo uso del archivo. Cierre dicho programa para continuar.')
                    #except AttributeError:
                        #   QMessageBox.about(self,'IMPORTANTE', 'Este mensaje se debe a que hubo un error al ingresar los datos a la base de datos. Contacte con el soporte')


            except ValueError:
                if self.manual:
                    QMessageBox.about(self, 'ERROR', 'Solo ingrese numeros en el campo "Telefono" , "Numero Interno" y "Numero Documento" ') 
                else:
                    QMessageBox.about(self, 'ERROR', 'Solo ingrese numeros en el campo "Telefono" ')                   
        else:
            QMessageBox.about(self, 'Sugerencia', 'Los campos nombre, telefono son obligatorios.')
    
    def duplicar_orden(self):

        datos_pdf = self.datos_orden["datos_pdf"]
        watermarks = self.datos_orden["watermarks"]
        
        interno = self.datos_orden["interno"]
        nro_doc = self.datos_orden["nro_doc"]
        tipo_doc = self.datos_orden["tipo_doc"]
        enchapado = self.datos_orden["enchape"]
        if self.r_enchape_5.isChecked():
            enchapado = 'SI'


        nombre= self.nombre_5.text().upper() #self.datos_orden["nombre"]
        telefono = self.telefono_5.text() #self.datos_orden["telefono"]
        fecha_orden = datetime.now().date()
        fecha_venta = self.datos_orden["fecha_venta"]
        fecha_estimada = self.fecha_5.date().toPyDate() # self.datos_orden["fecha_estimada"]
        oce =  self.oce_5.text() #self.datos_orden["oce"]

        despacho = self.datos_orden["despacho"]
        if self.r_despacho_5.isChecked():
            despacho = 'SI'
            watermarks.append('despacho') #v6.06

        vendedor = self.datos_orden["vendedor"]
        cont = self.contacto_5.text().upper() #self.datos_orden["cont"] #contacto
        detalle = self.datos_orden["detalle"]

        print("duplicando ...")

        cantidades,descripciones,valores_neto,vacias,correcto,lineas_totales = self.validar_item_orden(self.tb_modificar_orden)
        x = [cantidades,descripciones,valores_neto,vacias,correcto,lineas_totales]
        print(x)
        datos = [0,str(fecha_orden.strftime("%d-%m-%Y")),nombre,telefono,str((fecha_estimada).strftime("%d-%m-%Y")),cantidades,descripciones,enchapado,cont,oce,vendedor ]
        
        if vacias:
            QMessageBox.about(self, 'Alerta' ,'Una fila y/o columna esta vacia, rellenela para continuar' )
        elif lineas_totales > 14:
            QMessageBox.about(self, 'Alerta' ,'Filas totales: '+str(lineas_totales) + ' - El maximo soportado por el formato de la orden es de 14 filas.' )
        elif correcto == False:
            QMessageBox.about(self,'Alerta', 'Se encontro un error en una de las cantidades o Valores neto ingresados. Solo ingrese numeros en dichos campos')
        else:
            print("| (DUPLICAR ORDEN): TODOS LOS DATOS CORRECTOS")
            print(f"| DETALLE: {detalle} -> tipo: {type(detalle)}")

            print(f'| OLD CANTIDADES: {detalle["cantidades"]}')
            print(f"| OLD descripciones: {detalle['descripciones']}")
            print(f"| OLD netos: {detalle['valores_neto']}")

            print(f"| NEW CANTIDADES: {cantidades}")
            print(f"| NEW descripciones: {descripciones}")
            print(f"| NEW netos: {valores_neto}")

            detalle["cantidades"] = cantidades
            detalle["descripciones"] = descripciones
            detalle["valores_neto"] = valores_neto

            detalle = json.dumps(detalle)

            dialog = Duplicar(self)

            if dialog.exec():
                print("DIALOG ACEPTADO")
                datos_pdf["tipo"] = dialog.obtener_area()

                # DUPLICAR ORDEN DE TRABAJO
                self.nro_orden = 0
                if datos_pdf["tipo"] == "dimensionado":
                    self.nro_orden = self.conexion.root.registrar_orden_dimensionado( interno , str(fecha_venta), nombre , telefono, str(fecha_estimada) , detalle,tipo_doc, nro_doc,enchapado,despacho,str(fecha_orden),cont,oce,vendedor)
                    datos[0] = str(self.nro_orden)
                #v6.0
                elif datos_pdf["tipo"]in ["elaboracion","carpinteria", "pallets","sin_transformacion"]:

                    if datos_pdf["tipo"] == "sin_transformacion":
                        watermarks.append('material')

                    self.nro_orden = self.conexion.root.registrar_orden_general(datos_pdf["tipo"] , nombre,telefono,str(fecha_orden), str(fecha_estimada),nro_doc,tipo_doc,cont,oce, despacho, interno ,detalle, str(fecha_venta), vendedor)                    
                    datos[0] = str(self.nro_orden)
                else:
                    QMessageBox.about(self, 'ERROR', 'Seleccione un tipo de orden a generar, antes de proceder a registrar')  

                # Se crea vinculo , PDF y se muestra.
                if self.nro_orden != 0:
                    self.crear_vinculo(datos_pdf["tipo"])
                    print(f'(Duplicar orden) Creando PDF {datos_pdf["tipo"]} ...')
                    self.crear_pdf(datos, datos_pdf ,watermarks)
                    self.inicializar_buscar_venta()
                    boton = QMessageBox.question(self, f'Orden Duplicada correctamente', f'Nueva orden: {datos_pdf["tipo"]}| Nuevo Folio: {self.nro_orden} .\nDesea abrir el duplicado de Orden?')
                    if boton == QMessageBox.Yes:
                        self.ver_pdf(datos_pdf["tipo"]) 
            else:
                print("Dialog rechazado")
                return
            
            print("FIN DUPLICAR")


        
                

    def validar_manual_obs(self,obs):
        if self.manual:
            if obs != '':
                print('manual y obs valida')
                return True
            else:
                print('manual y obs no valida , error')
                return False
        else:
            print('no manual , no se cambia la observacion')
            return False

 #  Funcion para generar REINGRESO -------------
    def inicializar_reingreso(self):
        seleccion = self.tb_buscar_orden.currentRow()
        if seleccion != -1:
            _item = self.tb_buscar_orden.item( seleccion, 0) 
            if _item:            
                orden = self.tb_buscar_orden.item(seleccion, 0 ).text()
                self.nro_orden = int(orden)
                tipo = self.lb_tipo_orden.text()
                print('ventana  reingreso ... para: '+ tipo + ' -->' + str(self.nro_orden))
                self.lb_folio_orden.setText(str(self.nro_orden))
                self.tb_reingreso_2.setRowCount(0)
                self.tb_reingreso_2.setColumnWidth(1,100)
                self.tb_reingreso_2.setColumnWidth(0,600)
                self.tb_reingreso_2.setColumnWidth(2,100)
                self.combo_motivos.clear()
                self.combo_soluciones.clear()
                #self.txt_descripcion_2.clear()
                self.grupo_motivos.hide() #V5.8 
                self.txt_solucion_2.hide()#v5.8
                parametros_reingreso = self.conexion.root.obtener_parametros('parametros_reingreso')
                if parametros_reingreso:
                    if parametros_reingreso[1] != None:
                        detalle = json.loads(parametros_reingreso[1])
                        try:
                            lista_motivos = detalle["lista_motivos"]
                            for item in lista_motivos:
                                self.combo_motivos.addItem(item)
                        except:
                            pass
                        
                        try:
                            lista_soluciones = detalle['lista_soluciones']
                            for item in lista_soluciones:
                                self.combo_soluciones.addItem(item)
                        except:
                            pass
                
                self.rellenar_datos_reingreso()
                self.stackedWidget.setCurrentWidget(self.reingreso)
        else:
            QMessageBox.about(self,'ERROR', 'Seleccione un Nro interno antes de continuar')

    def registrar_reingreso(self):
    
        tipo_doc = self.lb_doc_2.text()
        try:
            nro_doc = int( self.lb_documento_2.text() )
        except ValueError:
            nro_doc = None
            
        fecha = datetime.now().date()
        motivo = self.combo_motivos.currentText() #v5.8
        
        proceso = self.lb_proceso_2.text()
        cliente = (self.lb_nombre_cliente.text()).upper() #v5.9
        descr = "DESHABILITADA" #self.txt_descripcion_2.toPlainText() #v5.9 Descripcion deshabilitada.

        solucion = self.combo_soluciones.currentText()  #v5.8
        lineas = 0

        if motivo != '' and descr != '' and solucion != '' :
            cant = self.tb_reingreso_2.rowCount()
            vacias = False #Determinna si existen campos vacios
            correcto = True #Determina si los datos estan escritos correctamente. campos cantidad y valor son numeros.
            cantidades = []
            descripciones = []
            valores_neto = []
            i = 0
            while i< cant:
                descripcion = self.tb_reingreso_2.item(i,0) #Collumna descripcion
                cantidad = self.tb_reingreso_2.item(i,1) #Columna cantidad
                neto = self.tb_reingreso_2.item(i,2) #Columna de valor neto
                if cantidad != None and descripcion != None and neto != None :  #Checkea si se creo una fila, esta no este vacia.
                    if cantidad.text() != '' and descripcion.text() != '' and neto.text() != '' :  #Chekea si se modifico una fila, esta no este vacia
                        try: 
                            nueva_cant = cantidad.text().replace(',','.',3)
                            nuevo_neto = neto.text().replace(',','.',3)
                            cantidades.append( float(nueva_cant) )
                            descripciones.append( (descripcion.text()).upper() ) #v5.9
                            print(descripcion.text())
                            linea = self.separar2( descripcion.text() , 60 )
                            lineas += len(linea)
                            print('total de lineas: '+ str(lineas))
                            valores_neto.append(float(nuevo_neto))

                        except ValueError:
                            correcto = False
                    else:
                        vacias=True
                else:
                    vacias = True
                i+=1
            if vacias:
                QMessageBox.about(self, 'Alerta' ,'Una fila y/o columna esta vacia, rellenela para continuar' )
            elif correcto == False:
                QMessageBox.about(self,'Alerta', 'Se encontro un error en una de las cantidades o Valores neto ingresados. Solo ingrese numeros en dichos campos')
            elif lineas > 4:
                QMessageBox.about(self, 'Alerta' ,'El maximo de filas por el formato de impresion es de 4.' )
            else:
                #print(cantidades)
                #print(valores_neto)
                formato = {
                        "cantidades" : cantidades,
                        "descripciones" : descripciones,
                        "valores_neto": valores_neto,
                        "creado_por" : self.datos_usuario[8],
                        "nombre_cliente"  : cliente,
                        "compatibilidad": f"version-{str(self.version)}"
                    }
                detalle = json.dumps(formato)
                print('procediendo a registrar reingreso ..')
                if self.conexion.root.registrar_reingreso( str(fecha), tipo_doc, nro_doc, self.nro_orden, motivo, descr, proceso, detalle,solucion):
                    resultado = self.conexion.root.obtener_max_reingreso()
                    self.nro_reingreso = resultado[0]
                    print('max nro reingreso: ' + str(resultado[0]) + ' de tipo: ' + str(type(resultado[0])))
                    datos = (resultado[0], str(fecha) , tipo_doc , nro_doc , motivo , descr , proceso , solucion, cantidades, descripciones, valores_neto,cliente)
                    self.crear_pdf_reingreso(datos)

                    boton = QMessageBox.question(self, 'Reingreso registrado correctamente', 'Desea ver el reingreso?')
                    if boton == QMessageBox.Yes:
                        self.stackedWidget.setCurrentWidget(self.inicio)
                        self.ver_pdf_reingreso()
                else:
                    QMessageBox.about(self,'ERROR','404 NOT FOUND. Contacte con Don Huber ...problemas al registrar')

        else:
            QMessageBox.about(self,'Datos incompletos','Los campos "descripcion" , "solucion" son obligatiorios. Como tambien si selecciona "OTROS" debe rellenar su campo')
    def rellenar_datos_reingreso(self):
        
            self.lb_proceso_2.setText(self.lb_tipo_orden.text())

            if self.lb_tipo_orden.text() == 'DIMENSIONADO':
                try:
                    resultado = self.conexion.root.buscar_orden_dim_numero(self.nro_orden)
                    if resultado != None :
                        self.lb_nombre_cliente.setText(resultado[3] ) #v 5.9

                        if resultado[18]: #si es manual
                            self.manual = True
                            if resultado[7]:
                                self.lb_doc_2.setText( resultado[7] ) #TIPO DOCUMENTO
                            else:
                                self.lb_doc_2.setText( 'No asignado' ) #TIPO DOCUMENTO
                            if resultado[8]:
                                self.lb_documento_2.setText( str(resultado[8]) ) #nro DOCUMENTO
                            else:
                                self.lb_documento_2.setText( 'No asignado' ) #nro DOCUMENTO
                        else:

                            self.lb_doc_2.setText( resultado[7] ) 
                            self.lb_documento_2.setText( str(resultado[8]) )

                        detalle = json.loads(resultado[6])

                        cantidades = detalle["cantidades"]
                        descripciones = detalle["descripciones"]
                        netos = detalle["valores_neto"]
                        j = 0
                        while j < len( cantidades ):
                            fila = self.tb_reingreso_2.rowCount()
                            self.tb_reingreso_2.insertRow(fila)
                            self.tb_reingreso_2.setItem(fila , 0 , QTableWidgetItem( descripciones[j] ) ) 
                            self.tb_reingreso_2.setItem(fila , 1 , QTableWidgetItem( str( cantidades[j] )  ) )
                            self.tb_reingreso_2.setItem(fila , 2 , QTableWidgetItem( str( netos[j] )  ) )
                            j+=1      
                                    
                except EOFError:
                    self.conexion_perdida()

            else:
                
                try:
                    resultado = None
                    tipo_orden = self.lb_tipo_orden.text()
                    if tipo_orden in ['ELABORACION','CARPINTERIA','PALLETS','SIN_TRANSFORMACION']:
                        resultado = self.conexion.root.buscar_orden_general_numero( tipo_orden.lower() , self.nro_orden)
                  
                    if resultado != None :
                        self.lb_nombre_cliente.setText(resultado[1] ) #v 5.9

                        if resultado[15]: #si es manual
                            self.manual = True

                            if resultado[6]:
                                self.lb_doc_2.setText( resultado[6] ) #TIPO DOCUMENTO
                            else:
                                self.lb_doc_2.setText( 'No asignado' ) #TIPO DOCUMENTO

                            if resultado[5]:
                                self.lb_documento_2.setText( str(resultado[5]) ) #nro DOCUMENTO
                            else:
                                self.lb_documento_2.setText( 'No asignado' ) #nro DOCUMENTO
                        else:                        
                            self.lb_doc_2.setText( resultado[6] )                 #TIPO DOCUMENTO
                            self.lb_documento_2.setText( str(resultado[5]) )       #NUMERO DOCUMENTO
                        
                                
                        detalle = json.loads(resultado[11])                  #DETALLE
                        cantidades = detalle["cantidades"]
                        descripciones = detalle["descripciones"]
                        netos = detalle["valores_neto"]
                        j = 0
                        while j < len( cantidades ):
                            fila = self.tb_reingreso_2.rowCount()
                            self.tb_reingreso_2.insertRow(fila)
                            self.tb_reingreso_2.setItem(fila , 0 , QTableWidgetItem( descripciones[j] ) ) 
                            self.tb_reingreso_2.setItem(fila , 1 , QTableWidgetItem( str( cantidades[j] )  ) )
                            self.tb_reingreso_2.setItem(fila , 2 , QTableWidgetItem( str( netos[j] )  ) )
                            j+=1     

                except EOFError:
                    self.conexion_perdida()

    def agregar_3(self):
        if self.tb_reingreso_2.rowCount() < 4:
            fila = self.tb_reingreso_2.rowCount()
            self.tb_reingreso_2.insertRow(fila)
        else:
            QMessageBox.about(self, 'ERROR', 'Ha alcanzado el limite maximo de filas (4). Intente crear otro REINGRESO para continuar agregando items.')

    def eliminar_3(self):
        fila = self.tb_reingreso_2.currentRow()  #FILA SELECCIONADA , retorna -1 si no se selecciona una fila
        if fila != -1:
            #print('Eliminando la fila ' + str(fila))
            self.tb_reingreso_2.removeRow(fila)
        else: 
            QMessageBox.about(self,'Consejo', 'Seleccione una fila para eliminar')
 # FUNCIONES PARA GESTION DEL REINGRESO 
    def inicializar_buscar_reingreso(self):
        self.anterior2 = True
        self.r_reingreso_2.setChecked(True)
        self.dateEdit_2.setCalendarPopup(True)
        self.dateEdit_2.setDate(datetime.now())
        self.stackedWidget.setCurrentWidget(self.buscar_reingreso)
        self.tb_buscar_reingreso.setRowCount(0)

    
    def buscar_reingreso_1(self):
        print('buscando reingreso')
        self.tb_buscar_reingreso.setRowCount(0)
        if self.r_reingreso_2.isChecked():
            print('uscando x numero reingreso')
            numero = self.txt_reingreso_2.text()
            if numero != '':
                numero = int(numero)
                print('buscando ', numero)
                consulta = self.conexion.root.obtener_reingreso_x_numero(numero)
                print(consulta)
                if consulta:
                    fila = self.tb_buscar_reingreso.rowCount()
                    self.tb_buscar_reingreso.insertRow(fila)
                    self.tb_buscar_reingreso.setItem(fila , 0 , QTableWidgetItem(str(consulta[1]))) #FECHA
                    self.tb_buscar_reingreso.setItem(fila , 1 , QTableWidgetItem(str(consulta[0]) )) #NRO REINGRESO
                    self.tb_buscar_reingreso.setItem(fila , 2 , QTableWidgetItem(str(consulta[2])))      #TIPO DOCUMENTO
                    self.tb_buscar_reingreso.setItem(fila , 3 , QTableWidgetItem(str(consulta[3]  ))) #NRO DOCUMENTO

                    self.tb_buscar_reingreso.setItem(fila , 4 , QTableWidgetItem(str(consulta[7] )   ))      #PROCESO
                    detalle = json.loads(consulta[8])
                    try:
                        creador = detalle["creado_por"]
                    except KeyError:
                        creador = 'No registrado'

                    self.tb_buscar_reingreso.setItem(fila , 5 , QTableWidgetItem( str(consulta[4] ) ))             # nro orden

                    try:
                        version = detalle["compatibilidad"]
                    except KeyError:
                        version = 'version-5.7'

                    self.tb_buscar_reingreso.setItem(fila , 6 , QTableWidgetItem( creador ))             #Vendedor
                    try:
                        estado = detalle["estado"]
                    except KeyError:
                        estado = 'VALIDA'
                    self.tb_buscar_reingreso.setItem(fila , 7 , QTableWidgetItem(  estado  ))  #estado

                    self.tb_buscar_reingreso.setItem(fila , 8 , QTableWidgetItem( version ))  #version de compatibilidad

        elif self.r_fecha_2.isChecked():
            
            print('uscando x fecha')
            fecha = self.dateEdit_2.date()
            print(fecha)
            fecha = fecha.toPyDate()
            print(fecha)

            lista_consulta = self.conexion.root.obtener_reingreso_x_fecha(str(fecha))
            if lista_consulta:
                for consulta in lista_consulta:
                    fila = self.tb_buscar_reingreso.rowCount()
                    self.tb_buscar_reingreso.insertRow(fila)
                    self.tb_buscar_reingreso.setItem(fila , 0 , QTableWidgetItem(str(consulta[1]))) #FECHA
                    self.tb_buscar_reingreso.setItem(fila , 1 , QTableWidgetItem(str(consulta[0]) )) #NRO REINGRESO
                    self.tb_buscar_reingreso.setItem(fila , 2 , QTableWidgetItem(str(consulta[2])))      #TIPO DOCUMENTO
                    self.tb_buscar_reingreso.setItem(fila , 3 , QTableWidgetItem(str(consulta[3]  ))) #NRO DOCUMENTO

                    self.tb_buscar_reingreso.setItem(fila , 4 , QTableWidgetItem(str(consulta[7] )   ))      #PROCESO
                    detalle = json.loads(consulta[8])
                    try:
                        creador = detalle["creado_por"]
                    except KeyError:
                        creador = 'No registrado'

                    self.tb_buscar_reingreso.setItem(fila , 5 , QTableWidgetItem( str(consulta[4] ) ))             # nro orden

                    try:
                        version = detalle["compatibilidad"]
                    except KeyError:
                        version = 'version-5.7'

                    self.tb_buscar_reingreso.setItem(fila , 6 , QTableWidgetItem( creador ))             #Vendedor
                    try:
                        estado = detalle["estado"]
                    except KeyError:
                        estado = 'VALIDA'
                    self.tb_buscar_reingreso.setItem(fila , 7 , QTableWidgetItem(  estado  ))  #estado

                    self.tb_buscar_reingreso.setItem(fila , 8 , QTableWidgetItem( version ))  #version de compatibilidad


        elif self.r_tipo_reingreso_2.isChecked():
            print('buscando x tipo reingreso')
            tipo = self.comboBox_2.currentText()
            print('buscando x ', tipo)

    def vista_modificar_reingreso(self):
        print('|-------- REDIRECCIONAR A VISTA MODIFICAR REINGRESO  ---------|')
        seleccion = self.tb_buscar_reingreso.currentRow()
        if seleccion != -1:
            _item = self.tb_buscar_reingreso.item( seleccion, 0) 
            if _item:            
                nro_reingreso = self.tb_buscar_reingreso.item( seleccion, 1).text() 
                
                print(self.version)
                estado = self.tb_buscar_reingreso.item( seleccion, 7).text() 
                print(estado)
                print('|-- Ventana  reingreso ... para:' , nro_reingreso )
                self.rellenar_datos_reingreso2(nro_reingreso) #ACCESIBLE DESDE EL GESTOR DE REINGRESOS
                self.tabWidget_2.setCurrentIndex(1)

                self.btn_registrar_6.hide() 

                self.btn_guardar_6.show()
                if estado == 'VALIDA':
                    self.btn_validar_reingreso.hide()
                    self.btn_anular_reingreso.show()
                elif estado == 'ANULADA':
                    self.btn_anular_reingreso.hide()
                    self.btn_validar_reingreso.show()

                self.label_65.show()
                self.lb_reingreso_6.show()
                self.lb_reingreso_6.setText(str(nro_reingreso))


                self.stackedWidget.setCurrentWidget(self.ingreso_manual)
        else:
            QMessageBox.about(self,'ERROR', 'Seleccione una fila antes de continuar')
    def rellenar_datos_reingreso2(self, id):

        self.rellenar_datos_manual()
        print('rellenando datos reingreso para: ', id)
        try:
            resultado = self.conexion.root.obtener_reingreso_x_numero(id) 
            self.nro_reingreso = id #GUARDA EL ID DEL REINGRESO PARA SU POSIBLE USO POSTERIOR (ANULAR, ACTUALIZAR, VER PDF)
            print('id reingreso copiado')
            tipo_doc = resultado[2] # 0 no asignado - 1 boleta - 2 factura - 3 guia
            if tipo_doc:
                if tipo_doc == 'BOLETA':
                    self.comboBox_6.setCurrentIndex(1)
                elif tipo_doc == 'FACTURA':
                    self.comboBox_6.setCurrentIndex(2)
                elif tipo_doc == 'GUIA':
                    self.comboBox_6.setCurrentIndex(3)

            nro_doc = resultado[3]
            self.txt_nro_doc_6.setText(str(nro_doc))
            nro_orden =  resultado[4]
            self.txt_orden_6.setText(str(nro_orden))
            motivo = resultado[5]

            if motivo == 'CAMBIO':
                self.r_cambio_6.setChecked(True)

            elif motivo == 'DEVOLUCION':
                self.r_devolucion_6.setChecked(True)
            else:
                self.r_otro_6.setChecked(True)
                self.txt_otro_6.setText(motivo)

            descripcion = resultado[6]
            #self.txt_descripcion_6.appendPlainText(descripcion)
            proceso = resultado[7]
            if proceso == 'DIMENSIONADO':
                self.r_d_6.setChecked(True)
            elif proceso == 'ELABORACION':
                self.r_e_6.setChecked(True)
            elif proceso == 'CARPINTERIA':
                self.r_c_6.setChecked(True)
            elif proceso == 'PALLETS':
                self.r_p_6.setChecked(True)

            detalle = json.loads(resultado[8])
            self.aux_detalle = detalle # COPIA DEL DETALLE, PARA LUEGO AÑADIR MAS OPCIONES SIN REALIZAR OTRA CONSULTA.

            cantidades = detalle["cantidades"]
            descripciones = detalle["descripciones"]
            netos = detalle["valores_neto"]
            j = 0
            while j < len( cantidades ):
                fila = self.tb_reingreso_manual.rowCount()
                self.tb_reingreso_manual.insertRow(fila)

                line_edit = QLineEdit( str(descripciones[j]) )
                line_edit.setStyleSheet(self.style_line_edit)
                line_edit.setCompleter(self.completer_productos)

                self.tb_reingreso_manual.setCellWidget(fila, 0, line_edit)
                self.tb_reingreso_manual.setItem(fila , 1 , QTableWidgetItem( str( cantidades[j] )  ) )
                self.tb_reingreso_manual.setItem(fila , 2 , QTableWidgetItem( str( netos[j] )  ) )
                j+=1 

            try:   # V5.9 añadido nombre cliente al reingreso.
                cliente = detalle['nombre_cliente']
                self.txt_nombre_cliente.setText(cliente)
            except KeyError:
                self.txt_nombre_cliente.setText('(no encontrado)')

            
            solucion = resultado[9]
            self.txt_solucion_6.appendPlainText(solucion)

            try:
                compatibilidad = detalle['compatibilidad']
                compatibilidad = compatibilidad.split('-')
                compatibilidad = float(compatibilidad[1])
                print('|-- Version reingreso detectada: ',str(compatibilidad))
                #self.version = compatibilidad
            except KeyError:
                #self.version = 'version-5.7'
                compatibilidad = 5.7
            self.aux_version = compatibilidad
            print(f'|-- App version: {str(self.version)}' )   
            if compatibilidad >= 5.8 :
                    print('|----- VERSION >= 5.8 COMPATIBLE')
                    self.grupo_motivos_2.hide()
                    self.txt_solucion_6.hide()
                    self.combo_motivos_2.show()
                    self.combo_soluciones_2.show()
                    aux = self.combo_motivos_2.findText(motivo)
                    print(aux)
                    if aux != -1:
                        self.combo_motivos_2.setCurrentIndex(aux)
                    
                    aux2 = self.combo_soluciones_2.findText(solucion)
                    print(aux2)
                    if aux2 != -1:
                        self.combo_soluciones_2.setCurrentIndex(aux2)
            else:
                print('|---- VERSION < 5.8 error de compatibilidad')
                #self.version = 'version-5.7'
                self.combo_motivos_2.hide()
                self.combo_soluciones_2.hide()
                self.grupo_motivos_2.show()
                self.txt_solucion_6.show()


                
        except EOFError:
            self.conexion_perdida()   
    def anular_validar_reingreso(self, nuevo_estado):
        if nuevo_estado == 'ANULAR':
            print('anulando reingreso: ',self.nro_reingreso)
            estado = 'ANULADA'
        elif nuevo_estado == 'VALIDAR':
            print('VALIDANDO reingreso: ',self.nro_reingreso)
            estado = 'VALIDA'

        print(self.aux_detalle)
        nuevo_detalle = self.aux_detalle
        nuevo_detalle['estado'] = estado
        fecha = datetime.now()
        fecha = fecha.strftime('%d-%m-%Y %H:%M')
        item = {
            'vendedor': self.datos_usuario[8],
            'fecha' : str(fecha) ,
            'estado': estado
        }
        print(item)
        itemx = json.dumps(item)

        historial = None
        try:
            historial = nuevo_detalle['historial']
            print('Tiene registro de cambios de estado.')
            historial.append(item)
            
        except KeyError:
            print('no tiene registro de cambios de estado.')
            historial = []
            historial.append(item)

        nuevo_detalle['historial'] = historial
        nuevo_detalle = json.dumps(nuevo_detalle)
        print(nuevo_detalle)
        try:
            if self.conexion.root.actualizar_detalle_reingreso(nuevo_detalle, self.nro_reingreso):
                self.tb_buscar_reingreso.setRowCount(0)
                self.aux_detalle = None
                self.stackedWidget.setCurrentWidget(self.buscar_reingreso)
                QMessageBox.about(self,'EXITO',f'Reingreso N° {self.nro_reingreso} {nuevo_estado} correctamente')
            else:
                QMessageBox.about(self,'ERROR','Problemas al {nuevo_estado} reingreso.\nPosiblemente no se hayan detectado cambios.\nContacte al Soporte.')
        except EOFError:
            self.conexion_perdida()


    def verificar_pdf_reingreso(self):
        seleccion = self.tb_buscar_reingreso.currentRow()
        if seleccion != -1:
            _item = self.tb_buscar_reingreso.item( seleccion, 0) 
            if _item:            
                nro_reingreso = self.tb_buscar_reingreso.item( seleccion, 1).text()
                self.nro_reingreso = nro_reingreso #Usada al abrir pdf.
                print(f'|------ Verificando existencia del pdf reingreso {nro_reingreso} ------------' )
                abrir = os.path.join( self.dir_reingreso , f'reingreso_{str(nro_reingreso)}.pdf')
                if  (not os.path.isfile(abrir)) :
                    print('|-- No existe PDF')

                    print('Obteniendo datos reales, creando PDF ...')
                    try:
                        resultado = self.conexion.root.obtener_reingreso_x_numero(self.nro_reingreso)
                        fecha = resultado[1]
                        tipo_doc = resultado[2]
                        nro_doc = resultado[3]
                        motivo = resultado[5]
                        descr = resultado[6]
                        proceso = resultado[7]
                        solucion = resultado[9]
                        detalle = json.loads(resultado[8])
                        cantidades = detalle['cantidades']
                        descripciones = detalle['descripciones']
                        valores_neto = detalle['valores_neto']
                        try:
                            cliente = (detalle['nombre_cliente']).upper()
                        except KeyError:
                            cliente = 'NO DEFINIDO'
                        datos = (nro_reingreso, str(fecha) , tipo_doc , nro_doc , motivo , descr , proceso , solucion, cantidades, descripciones, valores_neto,cliente)
                        self.crear_pdf_reingreso(datos)
                    except EOFError:
                        self.conexion_perdida()
                else:
                    print('|-- EXISTE PDF')
                try:
                    sleep(1)
                    self.ver_pdf_reingreso()
                except PermissionError:
                    QMessageBox.about(self,'ERROR', 'Otro programa tiene abierto el documento PDF. Intente cerrar el documento para poder volver a visualizarlo')
                

        else:
            QMessageBox.about(self,'ERROR', 'Seleccione una fila antes de continuar')

    def decidir_atras2(self):
        if self.anterior2: 
            self.stackedWidget.setCurrentWidget(self.buscar_reingreso)
        else:
            self.stackedWidget.setCurrentWidget(self.inicio)
# --------- Funciones de INGRESO MANUAL -----------------
    def inicializar_ingreso_manual(self):
        self.cargar_clientes('manual') #v5.9 se cargan los clientes antes.
        #v5.9 se habilitan todos los checkboxes primero.
        self.r_despacho_1.setEnabled(True)
        self.r_enchape_1.setEnabled(True)
        self.r_facturar_1.setEnabled(True)
        self.r_separar_material_1.setEnabled(True)
        self.r_uso_interno_1.setEnabled(True)
        #v5.9 se desmarcan todos los checkboxes.
        self.r_despacho_1.setChecked(False)
        self.r_enchape_1.setChecked(False)
        self.r_facturar_1.setChecked(False)
        self.r_separar_material_1.setChecked(False)
        self.r_uso_interno_1.setChecked(False)

        self.txt_obs_1.clear()
        self.anterior2 = False #para hacer que vuelva al inicio enves de gestion de reingresos
        
        self.btn_guardar_6.hide() #v.5.8
        self.btn_registrar_6.show() #v.5.8
        self.label_65.hide() #v.5.8 
        self.lb_reingreso_6.hide()#v.5.8

        self.grupo_motivos_2.hide() #v.5.8
        self.txt_solucion_6.hide() #v.5.8
        self.btn_anular_reingreso.hide()
        self.btn_validar_reingreso.hide()

        self.txt_nombre_cliente.setText('')
        self.combo_motivos_2.show()
        self.combo_soluciones_2.show()

        if self.datos_usuario[4] == 'SI' :
            self.rellenar_datos_manual()
            self.stackedWidget.setCurrentWidget(self.ingreso_manual)
        else:
            dialog = InputDialog2('CLAVE:', True ,'INGRESE CLAVE',self)
            if dialog.exec():
                clave = dialog.getInputs()
                try:
                    resultado = self.conexion.root.obtener_clave('dinamica') #v
                    end = 0
                    for item in resultado:
                        if clave == item[0]:    
                            self.clave = clave
                            self.rellenar_datos_manual()
                            self.stackedWidget.setCurrentWidget(self.ingreso_manual)
                            
                            end = 1
                    if end == 0:
                        QMessageBox.about(self,'ERROR' ,'CLAVE INVALIDA')

                except EOFError:
                    self.conexion_perdida()

    def rellenar_datos_manual(self):
        #-----datos de orden manual  ------
        self.r_uso_interno_1.setCheckState(False)
        self.r_facturar_1.setCheckState(False)
        self.r_enchape_1.setCheckState(False)
        self.r_despacho_1.setCheckState(False)
        self.nombre_1.setText('')
        self.telefono_1.setText('')
        self.contacto_1.setText('')
        self.oce_1.setText('')
        self.fecha_1.setDate( datetime.now())
        self.fecha_1.setCalendarPopup(True)
        self.tb_orden_manual.setRowCount(0)
        self.tb_orden_manual.setColumnWidth(0,100)
        self.tb_orden_manual.setColumnWidth(1,600)
        self.tb_orden_manual.setColumnWidth(2,100)
       
        #-----datos de reingreso manual  ------
        self.comboBox_6.clear()
        self.comboBox_6.addItem('No asignado')
        self.comboBox_6.addItem('BOLETA')
        self.comboBox_6.addItem('FACTURA')
        self.comboBox_6.addItem('GUIA')
        self.txt_orden_6.setText('0')
        self.txt_nro_doc_6.setText('0')
        self.txt_otro_6.setText('')
        #self.txt_descripcion_6.clear()
        self.txt_solucion_6.clear()
        self.txt_descripcion_7.setText('')

        self.tb_reingreso_manual.setRowCount(0)
        self.tb_reingreso_manual.setColumnWidth(1,100)
        self.tb_reingreso_manual.setColumnWidth(0,600)
        self.tb_reingreso_manual.setColumnWidth(2,100)

        self.combo_motivos_2.clear()
        self.combo_soluciones_2.clear()
        try:
            parametros_reingreso = self.conexion.root.obtener_parametros('parametros_reingreso')
            if parametros_reingreso:
                if parametros_reingreso[1] != None:
                    detalle = json.loads(parametros_reingreso[1])
                    #print(detalle)
                    try:
                        lista_motivos = detalle["lista_motivos"]
                        for item in lista_motivos:
                            self.combo_motivos_2.addItem(item)
                    except KeyError:
                        print('error key motivos')
                        lista_motivos = []
        
                    try:
                        lista_soluciones = detalle['lista_soluciones']
                        for item in lista_soluciones:
                            self.combo_soluciones_2.addItem(item)
                    except KeyError:
                        lista_soluciones = []
        except:
            pass

    def registrar_orden_manual(self):
        nombre = self.nombre_1.text().upper()     #NOMBRE CLIENTE #v5.9 mayuscula
        telefono = self.telefono_1.text() #TELEFONO
        fecha = self.fecha_1.date()  #FECHA ESTIMADA
        fecha = fecha.toPyDate()
        #self.tipo_doc = self.comboBox.currentText() #TIPO DE DOCUMENTO
        self.tipo_doc = None
        interno = 0 #NRO INTERNO
        self.nro_doc = None #NRO DOCUMENTO

        f_venta = None #FECHA DE  VENTA
        #f_venta = self.fecha_venta.dateTime() #FECHA DE  VENTA
        #f_venta = f_venta.toPyDateTime()
        #vendedor = self.txt_vendedor.text() #VENDEDOR
        vendedor = self.datos_usuario[8]
        observacion = self.txt_obs_1.toPlainText()
        lineas_totales = 0
        if nombre != '' and telefono != '' and observacion != '' :
            
            
            try:
                telefono = int(telefono)
                cant = self.tb_orden_manual.rowCount()
                vacias = False
                correcto = True
                cantidades = []
                descripciones = []
                valores_neto = []
                i = 0
                while i< cant:
                    cantidad = self.tb_orden_manual.item(i,0) #Columna cantidades
                    #v6.0
                    descripcion = self.tb_orden_manual.cellWidget(i,1)
                    neto = self.tb_orden_manual.item(i,2) #Columna de valor neto
                    if cantidad != None and descripcion != None and neto != None :  #Checkea si se creo una fila, esta no este vacia.
                        if cantidad.text() != '' and descripcion.text() != '' and neto.text() != '' :  #Chekea si se modifico una fila, esta no este vacia
                            try: 
                                nueva_cant = cantidad.text().replace(',','.',3)
                                nuevo_neto = neto.text().replace(',','.',3)
                                cantidades.append( float(nueva_cant) )
                                descripciones.append( (descripcion.text()).upper() ) #v5.9 mayuscula

                                lineas = self.separar(descripcion.text())
                                print(lineas)

                                lineas_totales = lineas_totales + len(lineas)
                                print(lineas_totales)
                                valores_neto.append(float(nuevo_neto))

                            except ValueError:
                                correcto = False
                        else:
                            vacias=True
                    else:
                        vacias = True
                    i+=1
                print('LINEAS TOTALES: ' + str(lineas_totales))
                if vacias:
                    QMessageBox.about(self, 'Alerta' ,'Una fila y/o columna esta vacia, rellenela para continuar' )
                elif lineas_totales > 14:
                    QMessageBox.about(self, 'Alerta' ,'Filas totales: '+str(lineas_totales) + ' - El maximo soportado por el formato de la orden es de 14 filas.' )
                elif correcto == False:
                    QMessageBox.about(self,'Alerta', 'Se encontro un error en una de las cantidades o Valores neto ingresados. Solo ingrese numeros en dichos campos')
                else:
                    
                    try:
                        watermarks = [] #v5.9 nuevas marcas de agua
                        datos_pdf = {
                            "tipo" : '',
                            "facturar" : 'NO'
                        }
                        enchape = 'NO'
                        despacho = 'NO'
                        self.nro_orden = 0 #v5.9
                        if self.r_despacho_1.isChecked():
                            despacho = 'SI'
                            watermarks.append('despacho') #v5.9
                        if self.r_facturar_1.isChecked():
                            datos_pdf['facturar'] = 'SI' 
                        if self.r_uso_interno_1.isChecked():
                            watermarks.append('interno') #v5.9
                        if self.r_separar_material_1.isChecked():
                            watermarks.append('material') #v5.9

                        
                        formato = {
                        "cantidades" : cantidades,
                        "descripciones" : descripciones,
                        "valores_neto": valores_neto,
                        "creado_por" : self.datos_usuario[8],
                        "por_facturar": datos_pdf["facturar"] #v5.9  (separar material y uso interno se almacenan como OBSERVACION)

                        }
                        detalle = json.dumps(formato)
                        
                        oce = self.oce_1.text()
                        fecha_orden = datetime.now().date()
                        cont = self.contacto_1.text().upper() #v5.9 mayuscula


                        if self.r_dim_1.isChecked():
                            datos_pdf['tipo'] = 'dimensionado'
                            if self.r_enchape_1.isChecked():
                                fecha = fecha + timedelta(days=2)
                                enchape = 'SI'
                            self.nro_orden =  self.conexion.root.registrar_orden_dimensionado( interno , f_venta , nombre , telefono, str(fecha) , detalle, self.tipo_doc, self.nro_doc ,enchape,despacho,str(fecha_orden),cont,oce, vendedor )
                            self.conexion.root.actualizar_orden_dim_obser(observacion , self.nro_orden)
                            
                        elif self.r_elab_1.isChecked():
                            datos_pdf['tipo'] = 'elaboracion'
                            self.nro_orden = self.conexion.root.registrar_orden_general('elaboracion',nombre,telefono,str(fecha_orden), str(fecha),self.nro_doc,self.tipo_doc,cont,oce, despacho, interno ,detalle,f_venta,vendedor)
                            self.conexion.root.actualizar_orden_general_obs('elaboracion',observacion , self.nro_orden)
                        elif self.r_carp_1.isChecked():
                            datos_pdf['tipo'] = 'carpinteria'
                            self.nro_orden = self.conexion.root.registrar_orden_general('carpinteria',nombre,telefono,str(fecha_orden), str(fecha),self.nro_doc,self.tipo_doc,cont,oce, despacho, interno ,detalle, f_venta ,vendedor)
                            self.conexion.root.actualizar_orden_general_obs('carpinteria',observacion , self.nro_orden)
                        elif self.r_pall_1.isChecked():
                            datos_pdf['tipo'] = 'pallets'
                            self.nro_orden = self.conexion.root.registrar_orden_general('pallets',nombre,telefono,str(fecha_orden), str(fecha),self.nro_doc,self.tipo_doc,cont,oce, despacho, interno ,detalle, f_venta,vendedor)
                            self.conexion.root.actualizar_orden_general_obs('pallets',observacion , self.nro_orden)
                        elif self.r_sin_trans_1.isChecked():
                            datos_pdf['tipo'] = 'sin_transformacion'
                            self.nro_orden = self.conexion.root.registrar_orden_general('sin_transformacion',nombre,telefono,str(fecha_orden), str(fecha),self.nro_doc,self.tipo_doc,cont,oce, despacho, interno ,detalle, f_venta,vendedor)
                            self.conexion.root.actualizar_orden_general_obs('sin_transformacion',observacion , self.nro_orden)
                        else:
                            QMessageBox.about(self, 'ERROR', 'Seleccione un tipo de orden a generar, antes de proceder a registrar')    
                        
                        if self.nro_orden != 0:
                            if self.clave:
                                self.conexion.root.eliminar_clave(self.clave,'dinamica')
                                self.clave = None
                            # CREAR PDF
                            datos = ( str(self.nro_orden) , str(fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones, enchape, cont,oce, vendedor)
                            self.crear_pdf(datos,datos_pdf, watermarks ) #v5.9
                            # SE ACTUALIZA EL VINCULO A VENTA. (NO_CREADO)
                            self.conexion.root.actualizar_vinculo_existente(datos_pdf['tipo'],self.nro_orden,'NO CREADO')
                            boton = QMessageBox.question(self, f'Orden de {datos_pdf["tipo"]} registrada correctamente', 'Desea abrir la Orden?')
                            if boton == QMessageBox.Yes:
                                self.ver_pdf(datos_pdf["tipo"])
                            self.stackedWidget.setCurrentWidget(self.inicio)
                        

                    except EOFError:
                        self.conexion_perdida()   
                    #except AttributeError:

                     #   QMessageBox.about(self,'IMPORTANTE', 'Este mensaje se debe a que hubo un error al ingresar los datos a la base de datos. Contacte con el soporte')


            except ValueError:
                QMessageBox.about(self, 'ERROR', 'Solo ingrese Numeros en los campos "Telefono", "Numero de documento" y "Numero interno" ')          
        else:
            QMessageBox.about(self, 'Sugerencia', 'Los campos "Nombre" , "Telefono" y "Observacion" son obligatorios')         
    def eliminar_clave(self):
        print('Eliminando clave dinamica...')

    def agregar_4(self):
        x = self.filas_estimadas_orden(self.tb_orden_manual)
        if self.tb_orden_manual.rowCount() < 14 and x < 14:
            fila = self.tb_orden_manual.rowCount()
            self.tb_orden_manual.insertRow(fila)

            line_edit = QLineEdit()
            line_edit.setStyleSheet(self.style_line_edit)
            line_edit.setCompleter(self.completer_productos)
            
            self.tb_orden_manual.setItem(fila, 0 , QTableWidgetItem( '0' ))
            self.tb_orden_manual.setCellWidget(fila, 1, line_edit)
            self.tb_orden_manual.setItem(fila, 2 , QTableWidgetItem( '0' ))

        else:
            QMessageBox.about(self, 'ERROR', f'Ha alcanzado el limite maximo de filas {14}. \nIntente crear otra Orden Manual para continuar agregando items.')

    def eliminar_4(self):
        fila = self.tb_orden_manual.currentRow()  #FILA SELECCIONADA , retorna -1 si no se selecciona una fila
        if fila != -1:
            #print('Eliminando la fila ' + str(fila))
            self.tb_orden_manual.removeRow(fila)
        else: 
            QMessageBox.about(self,'Consejo', 'Seleccione una fila para eliminar')
    def buscar_descripcion(self):
        self.productos.clear()
        descr = self.txt_descripcion_1.text()
        try:
            resultado = self.conexion.root.buscar_prod_descr(descr)
            for item in resultado:
                self.productos.addItem(item[1])
                #print(resultado)
        except EOFError:
            self.conexion_perdida()
        
    def buscar_codigo(self):
        self.productos.clear()
        codigo = self.txt_codigo_1.text()
        try:
            resultado = self.conexion.root.buscar_prod_cod(codigo)
            for item in resultado:
                self.productos.addItem(item[1])
        except EOFError:
            self.conexion_perdida()

    def add_descripcion(self):
        descripcion = self.productos.currentText()
        x = self.filas_estimadas_orden(self.tb_orden_manual)
        if self.tb_orden_manual.rowCount() < 14 and x < 14:
            fila = self.tb_orden_manual.rowCount()
            self.tb_orden_manual.insertRow(fila)

            line_edit = QLineEdit(str(descripcion))
            line_edit.setStyleSheet(self.style_line_edit)
            line_edit.setCompleter(self.completer_productos)
            
            self.tb_orden_manual.setItem(fila, 0 , QTableWidgetItem( '0' ))
            self.tb_orden_manual.setCellWidget(fila, 1, line_edit)
            self.tb_orden_manual.setItem(fila, 2 , QTableWidgetItem( '0' ))
        else:
            QMessageBox.about(self, 'ERROR', f'Ha alcanzado el limite maximo de filas {14}. \nIntente crear otra Orden para continuar agregando items.')
    def cambiar_observacion(self,vista): #v5.9

        '''self.r_despacho_1.setChecked(False)
        self.r_enchape_1.setChecked(False)
        self.r_facturar_1.setChecked(False)
        self.r_separar_material_1.setChecked(False)
        self.r_uso_interno_1.setChecked(False)'''
        if vista == 'manual': 
            # CHECKBOXES REGISTRAR ORDEN MANUAL
            self.r_despacho_1.setEnabled(True)
            self.r_enchape_1.setEnabled(True)
            self.r_facturar_1.setEnabled(True)
            self.r_separar_material_1.setEnabled(True)
            self.r_uso_interno_1.setEnabled(True)

            if self.r_enchape_1.isChecked():
                print('regla: enchape <> despacho,facturar | blockea los demas')
                self.r_separar_material_1.setEnabled(False)
                self.r_uso_interno_1.setEnabled(False)

            if self.r_uso_interno_1.isChecked():
                print('regla: uso innterno <> despacho, separar material | blockea los demas')
                self.r_enchape_1.setEnabled(False)
                self.r_facturar_1.setEnabled(False)

            if self.r_facturar_1.isChecked():
                print('regla: POR FACTURAR <> despacho,enchape,separar material | blockea los demas')
                self.r_uso_interno_1.setEnabled(False)

            if self.r_separar_material_1.isChecked():
                print('regla: separar material <> despacho,interno,facturar| blockea  enchape')
                self.r_enchape_1.setEnabled(False)
        


            if self.r_uso_interno_1.isChecked() and self.r_separar_material_1.isChecked():
                obs = "USO INTERNO Y SEPARAR MATERIAL"
                self.txt_obs_1.clear() 
                self.txt_obs_1.setEnabled(False) 
                self.txt_obs_1.appendPlainText(obs)
            elif self.r_uso_interno_1.isChecked():
                obs = "USO INTERNO"
                self.txt_obs_1.clear() 
                self.txt_obs_1.setEnabled(False) 
                self.txt_obs_1.appendPlainText(obs)
            elif self.r_separar_material_1.isChecked():
                obs = "SEPARAR MATERIAL"
                self.txt_obs_1.clear() 
                self.txt_obs_1.setEnabled(False) 
                self.txt_obs_1.appendPlainText(obs)
            else:
                self.txt_obs_1.clear() 
                self.txt_obs_1.setEnabled(True) 

            print('OBS: ',self.txt_obs_1.toPlainText() )

        elif vista == 'modificar':
            # CHECKBOXES MODIFICAR ORDEN
            self.r_despacho_5.setEnabled(True)
            self.r_enchape_5.setEnabled(True)
            self.r_facturar_5.setEnabled(True)
            self.r_separar_material_5.setEnabled(True)
            self.r_uso_interno_5.setEnabled(True)

            if self.r_enchape_5.isChecked():
                print('regla: enchape <> despacho,facturar | blockea los demas')
                self.r_separar_material_5.setEnabled(False)
                self.r_uso_interno_5.setEnabled(False)

            if self.r_uso_interno_5.isChecked():
                print('regla: uso innterno <> despacho, separar material | blockea los demas')
                self.r_enchape_5.setEnabled(False)
                self.r_facturar_5.setEnabled(False)

            if self.r_facturar_5.isChecked():
                print('regla: POR FACTURAR <> despacho,enchape,separar material | blockea los demas')
                self.r_uso_interno_5.setEnabled(False)

            if self.r_separar_material_5.isChecked():
                print('regla: separar material <> despacho,interno,facturar| blockea  enchape')
                self.r_enchape_5.setEnabled(False)

            if self.r_uso_interno_5.isChecked() and self.r_separar_material_5.isChecked():
                obs = "USO INTERNO Y SEPARAR MATERIAL"
                self.txt_obs_2.clear() 
                self.txt_obs_2.setEnabled(False) 
                self.txt_obs_2.appendPlainText(obs)
            elif self.r_uso_interno_5.isChecked():
                obs = "USO INTERNO"
                self.txt_obs_2.clear() 
                self.txt_obs_2.setEnabled(False) 
                self.txt_obs_2.appendPlainText(obs)
            elif self.r_separar_material_5.isChecked():
                obs = "SEPARAR MATERIAL"
                self.txt_obs_2.clear() 
                self.txt_obs_2.setEnabled(False) 
                self.txt_obs_2.appendPlainText(obs)
            else:
                self.txt_obs_2.clear() 
                self.txt_obs_2.setEnabled(True) 
                self.txt_obs_2.appendPlainText(self.aux_obs)
            print('OBS: ',self.txt_obs_2.toPlainText() )

        

    def rellenar_datos_cliente_manual(self):
        print('------- QCOMPLETE MANUAL ACTIVADO --------------')
        nombre_cliente = self.nombre_1.text()
        print('Nombre: ' + nombre_cliente + ' | LEN: ' + str(len(nombre_cliente)) )
        try:
            index = self.nombres.index(nombre_cliente)
        except ValueError:
            index = -1

        if index >= 0:
            #nombre_cliente = nombre_cliente.rstrip()
            #print('Nombre: ' + nombre_cliente + ' | LEN: ' + str(len(nombre_cliente)) )
            self.telefono_1.setText(str(self.telefonos[index]))
            self.contacto_1.setText(str(self.contactos[index]))
        print('------- QCOMPLETE MANUAL FIN --------------')

 #------ funciones reingreso manual
    def reingreso_manual(self, modo):
        print('f|-------- REINGRESO MANUAL: {modo}  --------|')
        nro_orden = self.txt_orden_6.text()           #NUMERO DE ORDEN
        tipo_doc = self.comboBox_6.currentText() #TIPO DE DOCUMENTO
        nro_doc = self.txt_nro_doc_6.text()        #NUMERO DE DOCUMENTO
        fecha = datetime.now().date()              #FECHA DE REINGRESO
        
        if self.aux_version >= 5.8:
            print('|-- Reingreso superior a 5.8: Obteniendo datos de combobox')
            motivo = self.combo_motivos_2.currentText()    #MOTIVO v5.8 O superior
            solucion = self.combo_soluciones_2.currentText()  #solucion v5.8 O superior
        else:
            print('|-- Reingreso inferior a 5.8: Obteniendo datos de qlineEdit')
            motivo = ''
            if self.r_cambio_6.isChecked():
                motivo = 'CAMBIO'
            elif self.r_devolucion_6.isChecked():
                motivo = 'DEVOLUCION'
            elif self.r_otro_6.isChecked():
                motivo = self.txt_otro_6.text()
            
            solucion = self.txt_solucion_6.toPlainText()   #solucion
        
        proceso = None
        if self.r_d_6.isChecked():
            proceso = 'DIMENSIONADO'
        elif self.r_e_6.isChecked():
            proceso = 'ELABORACION'
        elif self.r_c_6.isChecked():
            proceso = 'CARPINTERIA'
        elif self.r_p_6.isChecked():
            proceso = 'PALLETS'

        descr = "DESHABILITADA" #v5.9 self.txt_descripcion_6.toPlainText()   #descripcion 
        cliente = (self.txt_nombre_cliente.text()).upper() #v5.9 mayuscula 
        print('Nombre cliente reingreso: ' ,cliente)
        lineas = 0
        if motivo != '' and descr != '' and solucion != '' and cliente != '':
            cant = self.tb_reingreso_manual.rowCount()
            vacias = False #Determinna si existen campos vacios
            correcto = True #Determina si los datos estan escritos correctamente. campos cantidad y valor son numeros.
            cantidades = []
            descripciones = []
            valores_neto = []
            i = 0
            while i< cant:
                #v6.0
                descripcion_widget = self.tb_reingreso_manual.cellWidget(i, 0)  # Obtiene el widget de celda en la columna de descripción
                cantidad = self.tb_reingreso_manual.item(i,1) #Columna cantidad
                neto = self.tb_reingreso_manual.item(i,2) #Columna de valor neto
                if cantidad != None and descripcion_widget != None and neto != None :  #Checkea si se creo una fila, esta no este vacia.
                    if cantidad.text() != '' and descripcion_widget.text() != '' and neto.text() != '' :  #Chekea si se modifico una fila, esta no este vacia
                        try: 
                            nueva_cant = cantidad.text().replace(',','.',3)
                            nuevo_neto = neto.text().replace(',','.',3)
                            cantidades.append( float(nueva_cant) )
                            descripciones.append( (descripcion_widget.text()).upper() ) #v5.9 
                            linea = self.separar2(descripcion_widget.text(), 60 )
                            lineas += len(linea)
                            valores_neto.append(float(nuevo_neto))

                        except ValueError:
                            correcto = False
                    else:
                        vacias=True
                else:
                    vacias = True
                i+=1
            if vacias:
                QMessageBox.about(self, 'Alerta' ,'Una fila y/o columna esta vacia, rellenela para continuar' )
            elif correcto == False:
                QMessageBox.about(self,'Alerta', 'Se encontro un error en una de las cantidades o Valores neto ingresados. Solo ingrese numeros en dichos campos')
            elif lineas > 4:
                QMessageBox.about(self, 'Alerta' ,'El maximo de filas por el formato de impresion es de 4.' )
            elif lineas < 1:
                QMessageBox.about(self, 'Alerta' ,'Como minimo ingrese 1 item.' )
            elif proceso == None:
                QMessageBox.about(self, 'Alerta' ,'Seleccione el proceso de la orden de trabajo antes de continuar' )
            else:
                #print(cantidades)
                #print(valores_neto)
                formato = {
                        "cantidades" : cantidades,
                        "descripciones" : descripciones,
                        "valores_neto": valores_neto,
                        "nombre_cliente" : cliente, #v5.9 
                        "creado_por" : self.datos_usuario[8]
                    }
                if modo == 'registrar':
                    formato["compatibilidad"] = f'version-{ str(self.version) }'
                elif modo == 'actualizar':
                    formato["compatibilidad"] = f'version-{ str(self.aux_version) }'
                    
                detalle = json.dumps(formato)
                try:
                    nro_orden = int(nro_orden)
                    nro_doc = int(nro_doc)

                    #MODO PARA REGISTRAR LOS REINGRESOS
                    if modo == 'registrar':
                        print('|-- Registrando reingreso manual ')
                        if self.conexion.root.registrar_reingreso( str(fecha), tipo_doc, nro_doc, nro_orden, motivo, descr, proceso, detalle,solucion):
                            resultado = self.conexion.root.obtener_max_reingreso()
                            self.nro_reingreso = resultado[0]
                            print('max nro reingreso: ' + str(resultado[0]) + ' de tipo: ' + str(type(resultado[0])))
                            datos = (resultado[0], str(fecha) , tipo_doc , nro_doc , motivo , descr , proceso , solucion, cantidades, descripciones, valores_neto,cliente)
                            self.crear_pdf_reingreso(datos) 
                            if self.clave:
                                    self.conexion.root.eliminar_clave(self.clave,'dinamica')
                                    self.clave = None

                            boton = QMessageBox.question(self, 'Reingreso registrado correctamente', 'Desea ver el reingreso?')
                            if boton == QMessageBox.Yes:
                                self.ver_pdf_reingreso()
                            self.stackedWidget.setCurrentWidget(self.inicio)
                        else:
                            QMessageBox.about(self,'ERROR','404 NOT FOUND. Contacte con Don Huber ...problemas al registrar reingreso')
                    
                    #MODO PARA ACTUALIZAR LOS REINGRERSOS
                    elif modo == 'actualizar':
                        print('|-- Actualizando reingreso ')
                        id = self.lb_reingreso_6.text()
                        
                        if self.conexion.root.actualizar_reingreso( id , tipo_doc, nro_doc, nro_orden, motivo, descr, proceso, detalle,solucion):
                            print('|-- Actualizado correctamente')
                            self.tb_buscar_reingreso.setRowCount(0)
                            self.stackedWidget.setCurrentWidget(self.buscar_reingreso)
                            QMessageBox.about(self,'EXITO',f'Reingreso N° {id} actualizado')
                        else:
                            QMessageBox.about(self,'ERROR','Problemas al actualizar reingreso.\nPosiblemente no se hayan detectado cambios.\nIntente realizar un cambio y luego proceda a actualizar')

                except ValueError:
                    QMessageBox.about(self,'ERROR','Ingresar solo numeros en los campos: "NUMERO DE ORDEN" y "NUMERO DE DOCUMENTO" ')
                
                except EOFError:
                    self.conexion_perdida()
        else:
            QMessageBox.about(self,'Datos incompletos','Los campos "Nombre CLiente", "Motivo" y "solucion" son obligatiorios.')
    def agregar_6(self):
        if self.tb_reingreso_manual.rowCount() < 4:
            fila = self.tb_reingreso_manual.rowCount()
            self.tb_reingreso_manual.insertRow(fila)

            line_edit = QLineEdit()
            line_edit.setStyleSheet(self.style_line_edit)
            line_edit.setCompleter(self.completer_productos)

            self.tb_reingreso_manual.setCellWidget(fila, 0, line_edit)
            self.tb_reingreso_manual.setItem(fila, 1 , QTableWidgetItem( '0' ))
            self.tb_reingreso_manual.setItem(fila, 2 , QTableWidgetItem( '0' ))
        else:
            QMessageBox.about(self, 'ERROR', 'Ha alcanzado el limite maximo de filas (4). Intente crear otro REINGRESO para continuar agregando items.')

    def eliminar_6(self):
        fila = self.tb_reingreso_manual.currentRow()  #FILA SELECCIONADA , retorna -1 si no se selecciona una fila
        if fila != -1:
            #print('Eliminando la fila ' + str(fila))
            self.tb_reingreso_manual.removeRow(fila)
        else: 
            QMessageBox.about(self,'Consejo', 'Seleccione una fila para eliminar')
    def buscar_descripcion_2(self):
        self.productos_6.clear()
        if self.conexion: # verificar que exista conexion
            descr = self.txt_descripcion_7.text()
            try:
                resultado = self.conexion.root.buscar_prod_descr(descr)
                for item in resultado:
                    self.productos_6.addItem(item[1])
            except EOFError:
                self.conexion_perdida()
        
        
    def add_descripcion_2(self):
        descripcion = self.productos_6.currentText()

        if self.tb_reingreso_manual.rowCount() < 4 :
            fila = self.tb_reingreso_manual.rowCount()
            self.tb_reingreso_manual.insertRow(fila)

            line_edit = QLineEdit(str(descripcion))
            line_edit.setStyleSheet(self.style_line_edit)
            line_edit.setCompleter(self.completer_productos)

            self.tb_reingreso_manual.setCellWidget(fila, 0, line_edit)
            self.tb_reingreso_manual.setItem(fila, 1 , QTableWidgetItem( '0' ))
            self.tb_reingreso_manual.setItem(fila, 2 , QTableWidgetItem( '0' ))


        else:
            QMessageBox.about(self, 'ERROR', 'Ha alcanzado el limite maximo de filas soportado para la impresión de un reingreso. Intente hacer otro reingreso para los items faltantes ')
        
    def rellenar_datos_cliente_reingreso_manual(self):
        x  = 0
        
#---------- FUNCIONES DE INFORME ------------
    def inicializar_informe(self):
        self.tableWidget.setColumnWidth(0,371)
        self.vista_reingreso()
        self.actualizar()
        self.d_inicio.setDate( datetime.now().date() )
        self.d_inicio.setCalendarPopup(True)
        self.d_termino.setDate( datetime.now().date() )
        self.d_termino.setCalendarPopup(True)

        xd = {
            "tipo":"dimensionado",
            "facturar":"NO"
        }
        print(type(xd))
        print(xd)
        print(xd['tipo'])

        self.stackedWidget.setCurrentWidget(self.informes)

    def generar_informe(self):
        tipo_orden = self.comboBox.currentText()
        inicio = self.d_inicio.date()
        inicio = inicio.toPyDate()
        termino = self.d_termino.date()
        termino = termino.toPyDate()

        nombre = os.path.join(self.dir_informes , tipo_orden + '_'+ str( inicio.strftime("%d-%m-%Y")) + '_HASTA_' + str( termino.strftime("%d-%m-%Y")) + '.xlsx' )
        datos = None
        if self.r_orden_2.isChecked():
            try:
                if tipo_orden == 'DIMENSIONADO':
                    print('s')
                    datos = self.conexion.root.informe_dimensionado(str(inicio) , str(termino))
                    self.informe_dimensionado(datos,nombre)
                    
                elif tipo_orden == 'ELABORACION':
                    datos = self.conexion.root.informe_elaboracion(str(inicio) , str(termino))
                    self.informe_generico(datos,nombre)
                elif tipo_orden == 'CARPINTERIA':
                    datos = self.conexion.root.informe_carpinteria(str(inicio) , str(termino))
                    self.informe_generico(datos,nombre)
                elif tipo_orden == 'PALLETS':
                    datos = self.conexion.root.informe_pallets(str(inicio) , str(termino))
                    self.informe_generico(datos,nombre)
                elif tipo_orden == 'REINGRESO':
                    if self.r_d.isChecked() or self.r_e.isChecked() or self.r_c.isChecked() or self.r_p.isChecked() :
                        acept = []
                        nombre = tipo_orden #v6.0
                        entremedio = ''
                        if self.r_d.isChecked():
                            acept.append('DIMENSIONADO')
                            entremedio = entremedio + '_D'
                        if self.r_e.isChecked():
                            acept.append('ELABORACION')
                            entremedio = entremedio + '_E'
                        if self.r_c.isChecked():
                            acept.append('CARPINTERIA')
                            entremedio = entremedio + '_C'
                        if self.r_p.isChecked():
                            acept.append('PALLETS')
                            entremedio = entremedio + '_P'
                        fin = '_'+ str( inicio.strftime("%d-%m-%Y")) + '_HASTA_' + str( termino.strftime("%d-%m-%Y")) + '.xlsx' 
                        nombre = os.path.join(self.dir_informes , nombre + entremedio + fin )
                        datos = self.conexion.root.informe_reingreso(str(inicio), str(termino))
                        self.informe_reingreso(datos, acept, nombre)
                    else:
                        QMessageBox.about(self,'CONSEJO', 'SELECCIONE ALMENOS UN TIPO DE PROCESO PARA CCREAR EL INFORME DE REINGRESO')

                if datos:
                    self.actualizar()
                    QMessageBox.about(self,'EXITO', 'Informe generado correctamente')
                
            except PermissionError:
                QMessageBox.about(self,'ERROR',  'Otro programa tiene abierto el documento EXCEL. Intente cerrandolo y luego proceda a generar el EXCEL')
            except EOFError:
                self.conexion_perdida()

        elif self.r_venta.isChecked():
            QMessageBox.about(self,'EN DESARROLLO', 'Informes por fecha de venta no disponible en esta version')
            
    def informe_dimensionado(self,datos,nombre):
        if datos:
            wb = Workbook()
            ws = wb.active
            encabezado = ['TIPO DOCUMENTO','NRO DOCUMENTO','NOMBRE CLIENTE', 'NRO ORDEN', 'TELEFONO','DESCRIPCION','CANTIDAD','PRECIO NETO', 'DIMENSIONADOR','VENDEDOR','FECHA VENTA','FECHA ORDEN','FECHA INGRESO','FECHA ESTIMADA','FECHA REAL','ENCHAPADO','DESPACHO','CONTACTO','ORDEN COMPRA','OBSERVACION','ESTADO','MOTIVO','CREADO POR']

            ws.append(encabezado)
            f_ven = None

            for item in datos:
                if item[8]:
                    fecha_venta = datetime.fromisoformat( str( item[8] )) 
                    f_ven =  str( fecha_venta.strftime( "%d-%m-%Y %H:%M:%S" ) )  #FECHA DE VENTA
                f_ing = 'No asignada'
                f_real = 'No asignada'
                if item[4]:
                    ingreso = datetime.fromisoformat(str( item[4]) )
                    f_ing =  str(ingreso.strftime("%d-%m-%Y") ) #FECHA DE INGRESO
                estimada = datetime.fromisoformat(str( item[3]) )
                f_est =  str(estimada.strftime("%d-%m-%Y") )  #FECHA ESTIMADA DE ENTREGA
                if item[5]:
                    real = datetime.fromisoformat(str( item[5]) )
                    f_real =  str(real.strftime("%d-%m-%Y") ) #FECHA REAL DE ENTREGA
                fecha_orden = datetime.fromisoformat(str( item[14]) )
                f_ord =  str(fecha_orden.strftime("%d-%m-%Y") )  #FECHA DE ORDEN    
                detalle = json.loads( item[6])
                cant = detalle['cantidades']
                desc = detalle['descripciones']
                net = detalle["valores_neto"]
                try:
                    creador = detalle["creado_por"]
                except KeyError:
                    creador = 'No registrado'

                if item[19]:
                    extra = json.loads( item[19] )
                    estado = extra["estado"]
                    motivo = extra["motivo"]
                    
                else:
                    estado = 'VALIDA'
                    motivo = 'Ninguno'
                j = 0
                while j < len(cant):
                    fila = [item[11],item[10],item[1], item[0], item[2], desc[j],cant[j],net[j] ,item[9],item[17],f_ven,f_ord,f_ing,f_est,f_real,item[12],item[13],item[15],item[16],item[18],estado,motivo,creador]
                    ws.append(fila)
                    j +=1
            filas_total = ws.max_row
            tab = Table(displayName="tabla_dimensionado" , ref="A1:W"+ str(filas_total) )
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            dim_holder = DimensionHolder(worksheet=ws)
            for col in range(ws.min_column, ws.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)
            dim_holder['C'] = ColumnDimension(ws, min=3 ,max = 3, width=30)
            dim_holder['F'] = ColumnDimension(ws, min=6 ,max = 6, width=35)
            dim_holder['J'] = ColumnDimension(ws, min=10 ,max = 10, width=30)
            ws.column_dimensions = dim_holder
            ws.add_table(tab)
            wb.save(nombre)
        
        else:
            QMessageBox.about(self, 'ERROR', 'No se encontraron ordenes de trabajo para el rango de fechas definido')   
 
    def informe_generico(self,datos,nombre):
        if datos:
            wb = Workbook()
            ws = wb.active
            encabezado = ['TIPO DOCUMENTO','NRO DOCUMENTO','NOMBRE CLIENTE', 'NRO ORDEN', 'TELEFONO','DESCRIPCION','CANTIDAD','PRECIO NETO','VENDEDOR','FECHA VENTA','FECHA ORDEN','FECHA ESTIMADA','FECHA INGRESO','TRABAJADOR','FECHA REAL','DESPACHO','CONTACTO','ORDEN COMPRA','OBSERVACION','ESTADO','MOTIVO','CREADA POR']
            ws.append(encabezado)
            fv = None
            for item in datos:
                td =  item[7] #TIPO DOCUMENTO
                nd =  item[6]  #NUMERO DOCUMENTO  
                cl =  item[1]  #CLIENTE
                no =  item[0]  #NUMERO DE ORDEN
                tel =  item[2]  #TELEFONO
                detalle = json.loads( item[12]  )
                cant = detalle['cantidades']  #cantidades
                desc = detalle['descripciones'] #descripciones
                net = detalle["valores_neto"]  #valores neto
                try:
                    creador = detalle["creado_por"]
                except KeyError:
                    creador = 'No registrado'

                if item[13]:
                    fecha_venta = datetime.fromisoformat( str( item[13] ))
                    fv =  str( fecha_venta.strftime( "%d-%m-%Y %H:%M:%S" ) )  #FECHA DE VENTA
                estimada = datetime.fromisoformat(str( item[4]) )
                fe =  str(estimada.strftime("%d-%m-%Y") )  #FECHA ESTIMADA DE ENTREGA
                fr = 'No asignada'
                if item[5]:
                    real = datetime.fromisoformat(str( item[5]) )
                    fr = str(real.strftime("%d-%m-%Y") )   #FECHA REAL
                de =  item[10]  #DESPACHO
                co =  item[8]  #CONTACTO
                oc =  item[9]  #ORDEN DE COMPRA
                ve =  item[14]  #VENDEDOR
                fecha_orden = datetime.fromisoformat(str( item[3]) )
                fo =  str(fecha_orden.strftime("%d-%m-%Y") )  #FECHA DE ORDEN
                ob =  item[15] #OBSERVACION
                if item[16]: #ESTADO
                    extra = json.loads( item[16] )
                    estado = extra["estado"]
                    motivo = extra["motivo"]
                    
                else:
                    estado = 'VALIDA'
                    motivo = 'Ninguno'
                # version 5.5.1 informe actualizado compatible con personal area 5.3.1
                if item[17]:
                    trabajador = item[17]
                else:
                    trabajador = 'NO ASIGNADO'
                if item[18]:
                    ingreso = datetime.fromisoformat(str( item[18]) )
                    fi = str(ingreso.strftime("%d-%m-%Y") )   #FECHA REAL
                else:
                    fi = 'NO INGRESADA'
                #--------------------------------------
                j = 0
                while j < len(cant):
                    fila = [ td,nd,cl,no,tel,desc[j],cant[j],net[j],ve ,fv,fo,fe,fi,trabajador,fr,de,co,oc,ob,estado,motivo,creador]
                    ws.append(fila)
                    j +=1

            filas_total = ws.max_row
            tab = Table(displayName="tabla1" , ref="A1:V"+ str(filas_total) )
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            dim_holder = DimensionHolder(worksheet=ws)
            for col in range(ws.min_column, ws.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)
            dim_holder['C'] = ColumnDimension(ws, min=3 ,max = 3, width=30)
            dim_holder['F'] = ColumnDimension(ws, min=6 ,max = 6, width=35)
            dim_holder['I'] = ColumnDimension(ws, min=9 ,max = 9, width=30)
            ws.column_dimensions = dim_holder
            ws.add_table(tab)
            wb.save(nombre)
            
        else:
            QMessageBox.about(self, 'ERROR', 'No se encontraron ordenes de trabajo para el rango de fechas definido')

    def vista_reingreso(self):
        if self.comboBox.currentText() == 'REINGRESO':
            print('modo reingreso')
            self.groupBox.show()
            self.r_orden_2.hide()
            self.r_venta.hide()
            self.lb_buscar.show()
            self.lb_buscar.setText('FECHA DE CREACION')
        else:
            self.groupBox.hide()
            self.lb_buscar.hide()
            self.r_orden_2.show()
            self.r_venta.show()

    def informe_reingreso(self, datos,acept, nombre):
        if datos:
            wb = Workbook()
            ws = wb.active
            ws.title = 'REINGRESO_1'
            encabezado = ['NUMERO REINGRESO','FECHA REINGRESO','NUMERO DE ORDEN', 'TIPO DOCUMENTO', 'NUMERO DOCUMENTO', 'PROCESO','MERCADERIA','CANTIDAD','VALOR NETO','MOTIVO','DESCRIPCIÓN' ,'SOLUCIÓN','CREADO POR','ESTADO','HISTORIAL']
            ws.append(encabezado)
            for item in datos:
                nro_re =  str( item[0] ) #NRO REINGRESO INT
                reingreso = datetime.fromisoformat(str( item[1]) )
                fr =  str(reingreso.strftime("%d-%m-%Y") )  #FECHA REINGRESO
                td =  item[2] #TIPO DOCUMENTO   STR
                nd =  str( item[3] )  #NUMERO DOCUMENTO  INT
                no =  str( item[4] )  #NUMERO DE ORDEN    INT
                mot =  item[5]  #MOTIVO   STR
                descripcion =  item[6]  #DESCRIPCION  STR
                proc =  item[7]  #PROCESO   STR
                if proc in acept: #se filtra
                    detalle = json.loads( item[8]  ) #DETALLE
                    cant = detalle['cantidades']  #cantidades
                    merc = detalle['descripciones'] #mercaderia
                    net = detalle["valores_neto"]  #valores neto
                    try:
                        creador = detalle["creado_por"]
                    except KeyError:
                        creador = 'No registrado'

                    try:
                        estado = detalle["estado"]
                    except KeyError:
                        estado = 'VALIDA'
                    
                    try:
                        historial = json.dumps( detalle["historial"] )
                    except KeyError:
                        historial = 'SIN REGISTRO'

                    sol =  item[9] #SOLUCION  STR
                    j = 0
                    while j < len(cant):
                        fila = [ nro_re, fr, no, td, nd, proc, merc[j], cant[j], net[j], mot, descripcion ,sol, creador,estado,historial]
                        ws.append(fila)
                        j +=1

            filas_total = ws.max_row
            tab = Table(displayName="tabla1" , ref="A1:O"+ str(filas_total) )
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            dim_holder = DimensionHolder(worksheet=ws)
            for col in range(ws.min_column, ws.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=10)
            dim_holder['F'] = ColumnDimension(ws, min=6 ,max = 6, width=16)  #PROCESO
            dim_holder['G'] = ColumnDimension(ws, min=7 ,max = 7, width=45) #MERCADERIA 45px
            dim_holder['I'] = ColumnDimension(ws, min=9 ,max = 9, width=14) #VALOR NETO 10px
            dim_holder['J'] = ColumnDimension(ws, min=10 ,max = 10, width=14) #MOTIVO
            dim_holder['K'] = ColumnDimension(ws, min=11 ,max = 11, width=30)  #DESCRIPCION
            dim_holder['L'] = ColumnDimension(ws, min=12 ,max = 12, width=30)  #SOLUCION
            ws.column_dimensions = dim_holder
            ws.add_table(tab)
            wb.save(nombre)
        else: 
            QMessageBox.about(self,'ERROR', 'No se encontraron datos de reingreso para las fechas ingresadas.')

    def actualizar(self):
        self.tableWidget.setRowCount(0)
        informes = os.listdir(self.dir_informes)
        for item in informes:
            fila = self.tableWidget.rowCount()
            self.tableWidget.insertRow(fila)
            self.tableWidget.setItem(fila , 0 , QTableWidgetItem( item ) )  #nombre informe

    def abrir(self):
        seleccion = self.tableWidget.selectedItems()
        if seleccion != [] :
            nombre = seleccion[0].text()
            abrir = os.path.join(self.dir_informes , nombre)
            if os.path.isfile(abrir):
                Popen([abrir], shell=True)
            else:
                QMessageBox.about(self,'ERROR', 'El archivo no se encontro en la carpeta "informes" ')
        else:
            QMessageBox.about(self,'Sugerencia', 'Primero seleccione el nombre del informe para poder abrirlo ')

    def eliminar_excel(self):
        seleccion = self.tableWidget.selectedItems()
        if seleccion != [] :
            nombre = seleccion[0].text()
            ruta = os.path.join(self.dir_informes , nombre)
            if os.path.isfile(ruta):
                try:
                    os.remove(ruta)
                    self.actualizar()
                except PermissionError:
                    QMessageBox.about(self,'ERROR', 'No se pudo eliminar el informe, ya que archivo esta siendo utilizado por otro programa')
            else:
                QMessageBox.about(self,'ERROR', 'El archivo no se encontro')
        else:
            QMessageBox.about(self,'Sugerencia', 'Primero seleccione el nombre del informe para poder abrirlo ')
# ---------- FUNCIONES DE ESTADISTICAS ---------
    def inicializar_estadisticas(self):
        self.box_vendedores.clear()
        self.tb_orden_manual_2.setRowCount(0)
        self.tb_orden_manual_2.setColumnWidth(0,100)#tipo orden
        self.tb_orden_manual_2.setColumnWidth(1,100)#fecha orden
        self.tb_orden_manual_2.setColumnWidth(2,85)#nro orden
        self.tb_orden_manual_2.setColumnWidth(3,85)#interno
        self.tb_orden_manual_2.setColumnWidth(4,120)#documento
        self.tb_orden_manual_2.setColumnWidth(5,110)#nro_doc
        self.tb_orden_manual_2.setColumnWidth(6,150) #cliente

        self.stackedWidget.setCurrentWidget(self.estadisticas)
        if self.manuales:
            self.rellenar_tb_manuales()
        #cargar vendedores
        try:
            datos = self.conexion.root.obtener_vendedores_activos()
            #print(self.datos_usuario[8])
            self.box_vendedores.addItem('todos')
            self.box_vendedores.addItem(self.datos_usuario[8])
            if datos:
                for item in datos:
                    if item[10] != self.datos_usuario[8]:
                        self.box_vendedores.addItem(item[10])

            else:
                print('No se encontraron usuarios activos')
        except EOFError:
            self.conexion_perdida()
        
    def buscar_manuales(self):
        self.tb_orden_manual_2.setRowCount(0)
        tipo_orden = self.box_tipo_orden.currentText()
        vendedor = self.box_vendedores.currentText()
        print(tipo_orden)
        print(vendedor)
        self.manuales = None
        try:
            otro = {} # DETERMINA SI SE BUSCAN TODAS LAS MANUALES O SOLO INCOMPLETAS
            if self.ch_omitir_carp.isChecked():
                otro['omitir_uso_carpinteria'] = True
            if self.ch_omitir_exito.isChecked():
                otro['omitir_vinculo_exitoso'] = True

            otro['rango_dias'] = None #v5.9
                
            self.manuales = self.conexion.root.ordenes_manuales(tipo_orden,vendedor, otro) 
            if self.manuales:
                print('RELLENANDO TABLA CON ORDENES MANUALES PRICIPAL')
                for item in self.manuales:
                    fila = self.tb_orden_manual_2.rowCount()
                    self.tb_orden_manual_2.insertRow(fila)

                    date_string = str(item[1]) #fecha orden
                    date_object = datetime.strptime(date_string, "%Y-%m-%d")
                    aux = date_object.strftime('%d-%m-%Y')

                    self.tb_orden_manual_2.setItem(fila , 0, QTableWidgetItem( str(item[0]) ) ) #fecha orden
                    self.tb_orden_manual_2.setItem(fila , 1, QTableWidgetItem( str(aux) ) ) #fecha orden
                    self.tb_orden_manual_2.setItem(fila , 2, QTableWidgetItem( str(item[2]) ) ) #nro orden
                    if item[3] != None:
                        self.tb_orden_manual_2.setItem(fila , 3, QTableWidgetItem( str(item[3]) ) ) #NRO INTERNO
                    if item[4] != None:
                        self.tb_orden_manual_2.setItem(fila , 4, QTableWidgetItem( str(item[4]) ) ) # TIPO DOCUMENTO
                    if item[5] != None:
                        self.tb_orden_manual_2.setItem(fila , 5, QTableWidgetItem( str(item[5]) ) ) #NRO DOCUMENTO

                    self.tb_orden_manual_2.setItem(fila , 6, QTableWidgetItem( str(item[6]) ) ) # CLIENTE

                    if item[7] == None:
                        self.tb_orden_manual_2.setItem(fila , 7, QTableWidgetItem( 'NO CREADO' ) )
                    else:
                        self.tb_orden_manual_2.setItem(fila , 7, QTableWidgetItem( str(item[7]) ) ) #VINCULO

                    self.tb_orden_manual_2.setItem(fila , 8, QTableWidgetItem( str(item[8]) ) ) #VENDEDOR

            self.aux_tabla = self.tb_orden_manual_2

        except EOFError:
            self.conexion_perdida()

    def rellenar_tb_manuales(self):
        self.tb_orden_manual_2.setRowCount(0)
        if self.manuales:
                print('RELLENANDO TABLA CON ORDENES MANUALES PROCESO INTERNO')
                for item in self.manuales:
                    fila = self.tb_orden_manual_2.rowCount()
                    self.tb_orden_manual_2.insertRow(fila)

                    date_string = str(item[1]) #fecha orden
                    date_object = datetime.strptime(date_string, "%Y-%m-%d")
                    aux = date_object.strftime('%d-%m-%Y')

                    self.tb_orden_manual_2.setItem(fila , 0, QTableWidgetItem( str(item[0]) ) ) #fecha orden
                    self.tb_orden_manual_2.setItem(fila , 1, QTableWidgetItem( str(aux) ) ) #fecha orden
                    self.tb_orden_manual_2.setItem(fila , 2, QTableWidgetItem( str(item[2]) ) ) #nro orden
                    if item[3] != None:
                        self.tb_orden_manual_2.setItem(fila , 3, QTableWidgetItem( str(item[3]) ) ) #NRO INTERNO
                    if item[4] != None:
                        self.tb_orden_manual_2.setItem(fila , 4, QTableWidgetItem( str(item[4]) ) ) # TIPO DOCUMENTO
                    if item[5] != None:
                        self.tb_orden_manual_2.setItem(fila , 5, QTableWidgetItem( str(item[5]) ) ) #NRO DOCUMENTO

                    self.tb_orden_manual_2.setItem(fila , 6, QTableWidgetItem( str(item[6]) ) ) # CLIENTE

                    if item[7] == None:
                        self.tb_orden_manual_2.setItem(fila , 7, QTableWidgetItem( 'NO CREADO' ) )
                    else:
                        self.tb_orden_manual_2.setItem(fila , 7, QTableWidgetItem( str(item[7]) ) ) #VINCULO

                    self.tb_orden_manual_2.setItem(fila , 8, QTableWidgetItem( str(item[8]) ) ) #VENDEDOR
        

    def filtrar_solo_incompletas(self):
        
        if self.check_incompletas.isChecked():
            self.rellenar_tb_manuales()
            print('mostrando solo incompletas')
            if self.aux_tabla != None:
                self.tb_orden_manual_2 = self.aux_tabla
                remover = []
                column = 4 #COLUMNA DEL tipo documento
                # rowCount() This property holds the number of rows in the table
                for row in range(self.tb_orden_manual_2.rowCount()): 
                    # item(row, 0) Returns the item for the given row and column if one has been set; otherwise returns nullptr.
                    tipo_doc = self.tb_orden_manual_2.item(row, column) 
                    if tipo_doc:            
                        val_tipo_doc = self.tb_orden_manual_2.item(row, column).text()
                        val_nro_doc = self.tb_orden_manual_2.item(row, column + 1).text()
                        
                        vinculo = self.tb_orden_manual_2.item(row, 7 ).text() # VINCULO 
                        #print(vinculo)
                        if vinculo == 'CREADO' :
                            remover.append(row)
                print(remover)
                k = 0
                for i in remover:
                    self.tb_orden_manual_2.removeRow(i - k)
                    k += 1

        else:
            self.rellenar_tb_manuales()
            print('mostrando todas')
    
    def continuar_orden_manual(self):
        self.anterior = "ESTADISTICAS"
        self.limpiar_varibles()
        self.txt_vendedor_5.setText('')
        self.contacto_5.setText('') 
        self.oce_5.setText('')
        self.txt_interno_5.setEnabled(True) #solo lectura
        self.txt_nro_doc_5.setEnabled(True)
        self.date_venta_5.setEnabled(True)
        self.comboBox_5.setEnabled(True)
        self.txt_vendedor_5.setEnabled(True)
        # fecha_orden,nro_orden,interno,tipo_doc,nro_doc,nombre,vinc_existente 
        seleccion = self.tb_orden_manual_2.currentRow()
        if seleccion != -1:
            _item = self.tb_orden_manual_2.item( seleccion, 0) 
            if _item:            
                orden = self.tb_orden_manual_2.item(seleccion, 2 ).text() #nro orden
                self.nro_orden = int(orden)
                tipo = self.tb_orden_manual_2.item(seleccion, 0 ).text() #nro orden
                print('----------- MODIFICAR ORDEN MANUAL de estadisticas... para: '+ tipo + ' -->' + str(self.nro_orden) + ' -------------')
                self.tb_modificar_orden.setColumnWidth(0,100)
                self.tb_modificar_orden.setColumnWidth(1,650)
                self.tb_modificar_orden.setColumnWidth(2,100)
                #v5.9 se habilitan todos los checkboxes primero.
                self.r_despacho_5.setEnabled(True)
                self.r_enchape_5.setEnabled(True)
                self.r_facturar_5.setEnabled(True)
                self.r_separar_material_5.setEnabled(True)
                self.r_uso_interno_5.setEnabled(True)
                #v5.9 se desmarcan todos los checkboxes.
                self.r_despacho_5.setChecked(False)
                self.r_enchape_5.setChecked(False)
                self.r_facturar_5.setChecked(False)
                self.r_separar_material_5.setChecked(False)
                self.r_uso_interno_5.setChecked(False)

                self.rellenar_datos_orden(tipo)
                self.stackedWidget.setCurrentWidget(self.modificar_orden)
        
        else:
            QMessageBox.about(self,'ERROR', 'Seleccione una FILA antes de continuar')

    def crear_grafico(self):
        for i in reversed(range(self.box_grafico.count())): 
            self.box_grafico.itemAt(i).widget().setParent(None)

        tipo = self.tipo_estadistica.currentText()
        
        tipo_grafico = self.tipo_grafico.currentText()
        titulo = ''
        datos = None
        if tipo == 'GENERALES':
            titulo = 'CANTIDAD DE ORDENES POR ÁREA'
            datos = self.conexion.root.estadisticas_total_ordenes(False)
        elif tipo == 'ORDENES MANUALES':
            titulo = 'CANTIDAD DE ORDENES MANUALES POR ÁREA'
            datos = self.conexion.root.estadisticas_total_ordenes(True)
        elif tipo == 'ORDENES X VENDEDOR':
            
            tipo_orden = self.tipo_orden_trabajo.currentText()
            titulo = "ORDENES DE " + tipo_orden + ' X VENDEDOR'
            tipo_orden = tipo_orden.lower()
            datos = self.conexion.root.estadisticas_generales(tipo_orden)
        if self.tipo_grafico.currentText() == 'SCATTER PLOT':
            titulo = 'Grafico de visualizacion previa, en desarrollo...'
        if datos:
            x_list = []
            y_list = []
            for item in datos:
               x_list.append(item[0]) 
               y_list.append(item[1]) 

            grafico = Grafico(y_list,x_list,tipo_grafico,titulo)
            self.box_grafico.addWidget(grafico)

# Sondeos
    def iniciar_sondeo(self):
        print('|-- Iniciando Sondeo ...')
        if self.sondeo_params['estado_sondeo_manual'] == False:
            
            self.label_71.setText('ACTIVO')
            self.label_73.setText( str( self.sondeo_params['intervalo_sondeo_manual'] ))
            self.lb_rango_dias.setText(str(self.sondeo_params['rango_dias'] ))
            self.txt_rango_dias.setText(str(self.sondeo_params['rango_dias'] ))
            self.sondeo_params['estado_sondeo_manual'] = True
            self.btn_iniciar_sondeo_manual.setEnabled(False)
            self.btn_detener_sondeo_manual.setEnabled(True)
            self.btn_actualizar_config_sondeo.setEnabled(False)
            self.executor.submit(self.sondeo_manuales)
        else:
            print('|-- El sondeo ya fue iniciado')

    def sondeo_manuales(self):
        sleep(1)
        self.obtener_datos_sondeo(False)
        tiempo = 1
        while self.sondeo_params['estado_sondeo_manual']:
            if tiempo % self.sondeo_params['intervalo_sondeo_manual'] == 0 :
                fecha = datetime.now()
                print('|------ Sondeo realizado: ' + str(fecha) + '-------|') 
                self.obtener_datos_sondeo(False)

                tiempo = 1
            else:
                tiempo += 1
            sleep(1)
        print('|-- Sondeo ordenes manuales Finalizado')

    def obtener_datos_sondeo(self,actualizar_tabla):
        fecha_1 = datetime.now()
        fecha_1 =  fecha_1.strftime('%Y-%m-%d')
        otros_parametros = {
            "omitir_uso_carpinteria" :  self.sondeo_params['omitir_uso_carpinteria'] ,
            "omitir_vinculo_exitoso" : self.sondeo_params['omitir_vinculo_exitoso'] ,
            "rango_dias" : self.sondeo_params['rango_dias']

        }
        
        self.manuales = self.conexion.root.ordenes_manuales('todos','todos',otros_parametros)
        cantidad = str(len(self.manuales))
        self.btn_notificacion1.setText(cantidad)

        if actualizar_tabla:
            self.rellenar_tb_manuales()


    def actualizar_sondeo_config(self):
        intervalo = self.txt_intervalo_sondeo_manual.text()
        rango_dias = self.txt_rango_dias.text()
        print(intervalo)
        try:
            val = int(intervalo)
            if val > 30:
                self.sondeo_params['intervalo_sondeo_manual'] = val
                self.label_73.setText(intervalo)
            else:
                QMessageBox.about(self,'ERROR' ,'Intervalo minimo de 30 segundos')
            rango_dias = int(rango_dias)
            if rango_dias > 0:
                self.sondeo_params['rango_dias'] = rango_dias
                self.lb_rango_dias.setText(str(rango_dias))
            else:
                QMessageBox.about(self,'ERROR' ,'Rango dias minimo de 1 DIA')

        except:
            QMessageBox.about(self,'ERROR' ,'Ingrese solo numeros sobre el intervalo')

        if self.ch_omitir_uso_carp.isChecked():
            self.sondeo_params['omitir_uso_carpinteria'] = True
            self.label_75.setText('SI')
        else:
            self.sondeo_params['omitir_uso_carpinteria'] = False
            self.label_75.setText('NO')

        if self.ch_omitir_vinculo.isChecked():
            self.sondeo_params['omitir_vinculo_exitoso'] = True
            self.label_76.setText('SI')
        else:
            self.sondeo_params['omitir_vinculo_exitoso'] = False
            self.label_76.setText('NO')

    def finalizar_sondeo(self): 
        print('Finalizar sondeo')
        self.label_71.setText('INACTIVO')
        self.btn_iniciar_sondeo_manual.setEnabled(True)
        self.btn_detener_sondeo_manual.setEnabled(False)
        self.btn_actualizar_config_sondeo.setEnabled(True)
        self.sondeo_params['estado_sondeo_manual'] = False


    def easter_egg(self):
        r = lambda: randint(0,255)
        print('#%02X%02X%02X' % (r(),r(),r()))
        color = '#{:02x}{:02x}{:02x}'.format(r(), r(), r())
        self.btn_notificacion1.setStyleSheet("background-color: " + color  +"")

    def guardarDetalleOrden(self): #v5.9 
        print('GUARDANDO DETALLE ORDEN')
        datos = []
        for row in range(self.tb_modificar_orden.rowCount()):
            fila = []
            for col in range(self.tb_modificar_orden.columnCount()):
                fila.append(str(self.tb_modificar_orden.item(row, col).text()))
            datos.append(fila)
        self.datosTablaOrden = datos
        QMessageBox.about(self,'Exito' ,'Items copiados.')

    def pegarDetalleOrden(self):
        print('pegando detalle orden')
        if self.datosTablaOrden:
            self.tb_orden_manual.setRowCount(len(self.datosTablaOrden))
            self.tb_orden_manual.setColumnCount(len(self.datosTablaOrden[0]))
            for row, fila in enumerate(self.datosTablaOrden):
                for col, valor in enumerate(fila):
                    self.tb_orden_manual.setItem(row, col, QTableWidgetItem(valor))
            QMessageBox.about(self,'Exito' ,'Items pegados exitosamente..')
        else:
            QMessageBox.about(self,'Fallo' ,'No se detectaron items en el portapapeles.')

        

 # -------- funcion para generar clave ------------
    def generar_clave(self):
        dialog = InputDialog2('CLAVE:', False, 'REGISTRAR CLAVE',self)
        if dialog.exec():
            clave = dialog.getInputs()
            if clave != '':
                try:
                    if self.conexion.root.registrar_clave(clave,'dinamica'):
                        QMessageBox.about(self,'EXITO' ,'CLAVE: ' + clave +' REGISTRADA')
                    else:
                        QMessageBox.about(self,'ERROR' ,'ERROR AL REGISTRAR LA CLAVE')
                except EOFError:
                    self.conexion_perdida()
            else:
                QMessageBox.about(self,'ERROR' ,'Ingrese una clave antes de continuar')

#----- funciones generales -----------
#v6.0
    def habilitar_separar_material(self,checked ):
        print('habilitando separar material.')
        if checked:
            print("¡Se ha detectado un cambio en el QRadioButton!")
            self.r_separar_material_1.setChecked(True)
        else:
            print("no checkedado, quitando separar check-")
            self.r_separar_material_1.setChecked(False)

    def cargar_clientes(self,  tipo_estado):
        datos = []
        print('|-- Obteniendo Clientes de ordenes ' )
        try:
            datos = self.conexion.root.obtener_clientes_de_ordenes()
            print("(datos clientes ordenes type ): ", type(datos))
            #print(str(len(datos)))
            #print(str(type(datos)))
            #print('ordenes   | len: ' + str(len(datos)) + ' | type: '+ str(type(datos)))
            #self.txt_detalle.setText('orden: '+ tipo_orden + ' | len: ' + str(len(self.datos)) + ' | type: '+ str(type(self.datos)))
            ordenes = pd.DataFrame(datos) #([0]nombre, [1]telefono, [2]contacto)
            self.nombres = ordenes[0].tolist()
            self.telefonos = ordenes[1].tolist()
            self.contactos = ordenes[2].tolist()
            model = QStringListModel(self.nombres)
            #print(str(type(model)))
    
            if tipo_estado == 'normal':
                self.completer.setModel(model)
                self.completer.setMaxVisibleItems(7)
                self.completer.setCaseSensitivity(0) # 0: no es estricto con mayus | 1: es estricto con mayus
                #print(self.completer.filterMode()) #filtermode podria ser: coincidencias de inicio, fin o que basta que contenga.
                self.nombre_2.setCompleter(self.completer)
            elif tipo_estado == 'manual':
                # Completer cliente orden manual
                self.completer_manual.setModel(model)
                self.completer_manual.setMaxVisibleItems(7)
                self.completer_manual.setCaseSensitivity(0) # 0: no es estricto con mayus | 1: es estricto con mayus
                #print(self.completer_manual.filterMode()) #filtermode podria ser: coincidencias de inicio, fin o que basta que contenga.
                self.nombre_1.setCompleter(self.completer_manual)
                # Completer cliente Reingreso manual
                self.completer_reingreso_manual.setModel(model)
                self.completer_reingreso_manual.setMaxVisibleItems(7)
                self.completer_reingreso_manual.setCaseSensitivity(0) # 0: no es estricto con mayus | 1: es estricto con mayus
                #print(self.completer_reingreso_manual.filterMode()) #filtermode podria ser: coincidencias de inicio, fin o que basta que contenga.
                self.txt_nombre_cliente.setCompleter(self.completer_manual)
        except EOFError:
            self.conexion_perdida()
    def cargar_productos(self):
        datos = []
        print('|-- Extrayendo productos ... ---' )
        try:
            datos = self.conexion.root.obtener_todos_los_productos()
            #print(str(len(datos)))
            #print(str(type(datos)))
            #print('PRODUCTOS   | len: ' + str(len(datos)) + ' | type: '+ str(type(datos)))
            #self.txt_detalle.setText('orden: '+ tipo_orden + ' | len: ' + str(len(self.datos)) + ' | type: '+ str(type(self.datos)))
            productos = pd.DataFrame(datos) #([0]nombre, [1]telefono, [2]contacto)
            self.lista_codigos = productos[0].tolist() #codigo
            self.lista_productos = productos[1].tolist() #descripcion
            model = QStringListModel(self.lista_productos)
            #print(str(type(model)))
            self.completer_productos.setModel(model)
            self.completer_productos.setMaxVisibleItems(7)
            self.completer_productos.setCaseSensitivity(0) # 0: no es estricto con mayus | 1: es estricto con mayus
            #print(self.completer_productos.filterMode()) #filtermode podria ser: coincidencias de inicio, fin o que basta que contenga.
            print('|--- completer productos GENERADO ---')
        except EOFError:
            self.conexion_perdida()
        except:
            print('error cargar productos')
            pass
            
    def decidir_atras(self):
        if self.anterior: 
            self.stackedWidget.setCurrentWidget(self.estadisticas)
        else: 
            self.stackedWidget.setCurrentWidget(self.buscar_orden)

    def crear_vinculo(self,tipo):
        vinc = {
            "tipo" : tipo,
            "folio" : self.nro_orden }
        vinculo = json.dumps(vinc)
        # BUSCA SI EXISTEN VINCULOS.
        if self.conexion.root.añadir_vinculo_orden_a_venta(self.tipo_doc, vinculo,self.nro_doc):
            print('vinculo  de orden de '+ tipo +' a venta creado exitosamente')
        else:
            print('no se encontro el documento o sucedio un error')

    def crear_vinculo_v2(self,tipo):
        vinc = {
            "tipo" : tipo,
            "folio" : self.nro_orden }
        vinculo = json.dumps(vinc)
        # BUSCA SI EXISTEN VINCULOS.
        if self.conexion.root.añadir_vinculo_orden_a_venta(self.tipo_doc, vinculo,self.nro_doc):
            print('|-- Vinculo V2 de orden de trabajo '+tipo +' a venta creado exitosamente')
            self.lb_vinculo_5.setText('CREADO')
        else:
            print('|------ No se encontro el documento o sucedio un error')

    def actualizar_vinculo_orden_manual(self,aux_tipo,aux_nro,tipo_doc,nro_doc,estado):
        print('------------ ACTUALIZANDO VINCULO A ORDEN MANUAL -----------')
        print('|-- Estado vinculo: ' + estado +' - tipo orden: ' + self.tipo +' - nro_orden: '+ str(self.nro_orden))
        print(f'|-- ANTIGUOS: - tipo: {aux_tipo} , nro_doc : {str(aux_nro)} | NUEVOS: - tipo: {tipo_doc} , nro_doc: {str(nro_doc)}')
        if nro_doc != 0:
            print('|-- NRO ACEPTABLE...')
            if estado == 'NO CREADO':
                print('|-- Creando vinculo ......')
                new_tipo = self.tipo.lower()
                print(new_tipo)
                if tipo_doc != None or tipo_doc != 'NO ASIGNADO':
                    print('DOCUMENTO VALIDO: '+ tipo_doc)
                    if (aux_tipo == tipo_doc) and (aux_nro == nro_doc):
                        print('DOCUMENTO IDENTICO AL ANTERIOR, se cancela la vinculacion')
                    else:
                        print('DOCUMENTOS DISTINTO AL ANTERIOR, se procede a generar el vinculo')
                        self.crear_vinculo_v2(new_tipo)
                else:
                    print('TIPO DE DOCUMENTO NO VALIDO, VINCULO NO CREADO.')

            elif estado == 'CREADO':
                print('|-- creando vinculo ......')
                if tipo_doc != None or tipo_doc != 'NO ASIGNADO':
                    print('DOCUMENTO VALIDO: '+ tipo_doc)
                    if (aux_tipo == tipo_doc) and (aux_nro == nro_doc):
                        print('DOCUMENTO IDENTICO AL ANTERIOR, se cancela la vinculacion')
                    else:
                        print('DOCUMENTOS DISTINTO AL ANTERIOR, se procede a generar el vinculo')
                        new_tipo = self.tipo.lower()
                        print(new_tipo)
                        self.crear_vinculo_v2(new_tipo)
                else:
                    print('TIPO DE DOCUMENTO NO VALIDO, VINCULO NO CREADO.')

                print('ELIMINANDO vinculo anterior...')
        else:
            print("NRO DOC NO VALIDO = 0")
            new_tipo = self.tipo.lower()
        print('-------------------------------------------------------------')


    def crear_pdf(self, lista, datos, watermarks): #PDF DE ORDEN DE TRABAJO | watermarks v5.9 | datos: es un dicc| watermarks es array.
        print('------- CREANDO PDF -------')
        print('Datos: ',datos)
        print('Watermarks: ',watermarks)
        #6.0
        #ruta = ( self.carpeta +'app/ordenes/' + datos['tipo'] +'_' + lista[0] + '.pdf' )  #NRO DE ORDEN 
        ruta = os.path.join( self.dir_ordenes , datos['tipo'] +'_' + lista[0] + '.pdf' ) 
        print('1--PDF: ',ruta)
        #Formatos
        formato = os.path.join( self.dir_formatos , datos['tipo'] +".jpg")
        print('|-- Dir formato pdf: ',formato )
        agua = os.path.join( self.dir_formatos , 'new_despacho.png')
        uso_interno = os.path.join(self.dir_formatos , 'new_uso_interno.png')
        separar_material = os.path.join( self.dir_formatos ,"new_separar_material.png")

        hojas = 2  # Defecto -> especialmente para orden dimensionado.
        if 'interno' in watermarks or datos['tipo'] == 'sin_transformacion' or datos['tipo'] == 'carpinteria' or datos['tipo'] == 'pallets' or datos['tipo'] == 'elaboracion': # o uso interno
            hojas = 1
        try:
            documento = canvas.Canvas(ruta)

            for pagina in range(hojas):
                
                documento.setPageSize(( 216 * mm , 279 * mm))
                documento.drawImage( formato, 0* mm , 2 * mm , 216 *mm ,279 *mm )
                documento.setFillAlpha(1)
                documento.drawString( 0 * mm, 139.5 * mm ,'------------------------------------------------------------------------------------------')
                documento.drawString( 105 * mm, 139.5 * mm ,'----------------------------------------------------------------------------------------------')

                

                documento.setFont('Helvetica',10)

                k = 2.5 #constante
                salto = 0
                for i in range(2):
                    documento.rotate(90)        
                    if datos['facturar'] == 'SI':
                        documento.setFont('Helvetica',11)
                        documento.setFillAlpha(0.6)
                        documento.drawString( (53 + k + salto) *mm , -20.5 * mm , '"POR FACTURAR"' )  #por facturar
                    documento.setFillAlpha(1)


                    documento.setFont('Helvetica',9)
                    documento.drawString( (28 + k + salto) *mm , -59.5 * mm , lista[2] )  #NOMBRE

                    documento.drawString( (100 + k + salto) *mm , -66 * mm , str(lista[3]) )   #TELEFONO
                    documento.drawString( (106 + k +salto) *mm , -59.5 * mm , lista[1]  )    #FECHA DE ORDEN

                    
                    if self.tipo_doc: # si existe el tipo de documento
                        if self.tipo_doc == 'FACTURA':
                            documento.drawString( (45 + k + salto) *mm , -85 * mm , str(self.nro_doc) )     #NRO FACTURA FOLIO
                        elif self.tipo_doc == 'BOLETA':
                            documento.drawString( (15 + k + salto) *mm , -85 * mm ,  str(self.nro_doc) )      #NRO BOLETA FOLIO
                        elif self.tipo_doc == 'GUIA':
                            documento.drawString( (78 + k + salto) *mm , -85 * mm , str(self.nro_doc) )     #NRO GUIA  V5.4 VENDEDOR
                        

                    if datos['tipo'] == 'dimensionado':
                        documento.drawString( (88 + k + salto) *mm , -94 * mm ,  lista[7] )   #ENCHAPE

                    documento.drawString( (110 + k + salto) *mm , -85 * mm ,   lista[9]  )       #ORDEN DE COMPRA

                    documento.drawString( (32 + k + salto) *mm , -66 * mm ,  lista[8] )   #CONTACTO
                    if lista[10] != None:
                        documento.drawString( (33+ salto) *mm , -205 * mm , lista[10].upper() ) #NOMBRE VENDEDOR
                    
                    documento.setFont('Helvetica-Bold',12)
                    documento.drawString( (106 + k + salto) *mm , -44.5 * mm , lista[4]  ) #FECHA ESTIMADA
                    documento.drawString( (106 + k + salto) *mm , -20.5 * mm , lista[0]  ) #NRO DE ORDEN

                    documento.setFillColor(white)
                    documento.setStrokeColor(white)
                    documento.rect((5+ k + salto )* mm, -200 * mm ,80*mm,3*mm, fill=1 )
                    documento.setFillColor(black)
                    documento.setStrokeColor(black)
                    documento.setFont('Helvetica-Bold',10)

                    if pagina == 0:
                        if k == 2.5:
                            documento.drawString( (10+ salto) *mm , -200 * mm , 'ORIGINAL' )
                        else:
                            documento.drawString( (10+ salto) *mm , -200 * mm , 'COPIA ' + datos['tipo'].upper() )
                    else:
                        if k == 2.5:
                            documento.drawString( (10+ salto) *mm , -200 * mm , 'COPIA CLIENTE' )
                        else:
                            documento.drawString( (10+ salto) *mm , -200 * mm , 'COPIA BODEGA' )
                    documento.drawString( (10+ salto) *mm , -205 * mm , 'VENDEDOR:' )   

                    documento.rotate(-90)
                    # MARCAS DE AGUA 
                    documento.setFillAlpha(0.7)
                    margin = 15   # : 1watermark -> 41mm en adelante |2: -> 21mm
                    if len(watermarks) == 1:
                        margin += 41
                    elif len(watermarks) == 2:
                        margin += 21
                    x_pos = 0
                    for item in watermarks:
                        if item == 'despacho':
                            documento.drawImage( agua , 155* mm , (margin + x_pos + salto)* mm , 15*mm ,33*mm , mask= 'auto')
                        if item == 'interno':
                            documento.drawImage( uso_interno , 155* mm , (margin + x_pos + salto)* mm , 15*mm ,33*mm , mask= 'auto')
                        if item == 'material':   
                            documento.drawImage( separar_material , 155* mm ,(margin + x_pos + salto)*mm , 15*mm ,33*mm , mask= 'auto')
                        x_pos += 40
                    salto += 139.5
                    k = 0 



                documento.rotate(90)
                documento.setFillAlpha(1)
                #items
                cons = 108
                documento.setFont('Helvetica',9)
                align_cant_1 = 12
                align_descr_1 = 22
                align_cant_2 = 151
                align_descr_2 = 161

                cantidades = self.normalizar_cantidades(lista[5]) #arreglado los int se ven como int y float como float en la impresion
                descripciones = lista[6]
                i = 0
                while i < len(cantidades):
                    documento.drawString(align_cant_1 *mm , -cons* mm , str(cantidades[i]) )  #cantidad
                    
                    documento.drawString(align_cant_2 *mm , -cons * mm , str(cantidades[i]) )

                    cadenas = self.separar(descripciones[i])

                    for cadena in cadenas:
                        documento.drawString(align_descr_1 *mm , -cons* mm , cadena )  #descripcion

                        documento.drawString(align_descr_2 *mm , -cons* mm , cadena ) #54 LONG MAXIMA
                        cons += 5
                    i +=1
               
                documento.showPage()
            #documento.drawString( (10+ salto) *mm , -205 * mm , 'VENDEDOR:' )
            documento.save()
            print('------- CREANDO PDF FINALIZADO -------')
            sleep(1)
        except PermissionError:
            QMessageBox.about(self,'ERROR', 'Otro programa esta modificando este archivo por lo cual no se puede modificar actualmente.')
    
    def crear_pdf_reingreso(self, datos): #pdf de reingreso de mercaderias
            print('--------- CREANDO PDF REINGRESO -----------')
            ruta = os.path.join( self.dir_reingreso ,  f'reingreso_{str(datos[0])}.pdf' )
            print('|-- PDF: ' ,ruta)
            documento = canvas.Canvas(ruta)
            imagen =  os.path.join(self.dir_formatos ,"reingreso_solo.jpg" )
            print('|-- PDF format: ' ,imagen)
            documento.setPageSize(( 216 * mm , 279 * mm))
            documento.drawImage( imagen, 0* mm , 0 * mm , 216 *mm , 139.5 *mm )
            documento.drawImage( imagen, 0* mm , 139.5 * mm , 216 *mm , 139.5 *mm )
            documento.drawString( 0*mm , 139.5 *mm , '------------------------------------------------------------------------------')
            documento.drawString( 108*mm , 139.5 *mm , '----------------------------------------------------------------------------')
            salto = 0
        
            for i in range(2):
                documento.setFont('Helvetica',9)
                if datos[2] == 'FACTURA':
                    documento.drawString(131*mm, (99.5 + salto )*mm , str(datos[3])  )   #NUMERO DOCUMENTO , FACTURA
                elif datos[2] == 'BOLETA':
                    documento.drawString(53* mm, (99.5+ salto )*mm , str(datos[3])  )   #NUMERO DOCUMENTO , BOLETA
                elif datos[2] == 'GUIA':
                    documento.drawString(95* mm, (99.5+ salto )*mm , str(datos[3]) )   #NUMERO DOCUMENTO , guia
                '''k = 0 
                j = 0 
                for item in lista:
                    documento.drawString(20* mm, (92.5 + salto - k)*mm , lista[j] )  #descripcion del problema
                    #documento.drawString(20* mm, (92.5 + salto - k )*mm , descr )  #descripcion del problema
                    k += 6
                    j += 1'''
                
                documento.drawString(53*mm, (106 + salto )*mm , datos[11] ) #v5.9 NOMBRE CLIENTE
                documento.drawString(40*mm, (35 + salto )*mm , datos[7] )  #SOLUCION
                    

                

                cants = datos[8]
                descrs = datos[9]
                netos = datos[10]
                p = 0
                q = 0
                for item in cants:
                    documento.drawString(150*mm, (65 + salto - q )*mm , str(cants[p]) )  # PRODUCTO cantidad
                    documento.drawString(170*mm, (65 + salto - q )*mm , str(netos[p]) )  # PRODUCTO neto
                    cadenas = self.separar2(descrs[p] , 60 ) 
                    for cadena in cadenas:
                        documento.drawString(20*mm, (65 + salto -q )*mm , cadena)  # PRODUCTO descripcion
                        q += 5
                    p +=1

                documento.setFillAlpha(0.6)
                documento.drawString(18*mm, (6+ salto)*mm , 'Vendedor:' )  # VENDEDOR
                if self.vendedor: #si no existe es orden manual antigua. ordenes manuales acutales tienen vendedor asociado al usuario sagot.
                    documento.drawString(38*mm, (6 + salto  )*mm , (self.datos_usuario[8]).upper()  )  #nombre vendedor
                
                documento.setFillAlpha(1)
                
                documento.drawString(43*mm, (92.8 + salto )*mm , datos[4] )  # Motivo

                documento.setFont('Helvetica-Bold', 9 )

                documento.drawString(177*mm, (120 + salto )*mm ,  str(datos[0]) )  #NRO DE REINGRESO
                documento.drawString(177*mm, (106 + salto ) *mm , datos[1] )    #FECHA DEL REINGRESO

                if datos[6] == 'DIMENSIONADO':
                    documento.drawString(73*mm, (82 + salto )*mm , 'X' )  #PROCESO DIMENSIONADO
                elif datos[6] == 'ELABORACION':    
                    documento.drawString(150*mm, (82 + salto )*mm , 'X' )  #PROCESO ELABORACION
                elif datos[6] == 'CARPINTERIA':
                    documento.drawString(111*mm, (82 + salto )*mm , 'X' )  #PROCESO CARPINTERIA
                elif datos[6] == 'PALLETS':
                    documento.drawString(178.5*mm, (82 + salto )*mm , 'X' )  #PROCESO PALLETS

                salto += 139.5 
            documento.save()
            print('---------'*3)
    def anular(self):
        print(self.datos_usuario)

        if self.datos_usuario[4] == 'SI':
            dialog = InputDialog2('MOTIVO:' , False ,'INGRESE UN MOTIVO DE ANULACIÓN' ,self)
            dialog.resize(330,80)
            if dialog.exec():
                motivo  = dialog.getInputs()
                print(motivo)
                formato = {
                            "estado" : 'ANULADA',
                            "motivo" : motivo,
                            "usuario": self.datos_usuario[8]
                        }
                extra = json.dumps(formato)
                
                tipo = self.lb_tipo_orden.text()
                try:
                    self.conexion.root.anular_orden( tipo.lower(), extra, self.nro_orden )
                    QMessageBox.about(self,'EXITO' ,'Orden nro: ' + str(self.nro_orden) + ' Anulada')
                    self.inicializar_buscar_orden()
                except EOFError:
                    self.conexion_perdida()
        else:
            dialog = InputDialog('CLAVE:' ,'MOTIVO:', 'INGRESE UNA CLAVE Y MOTIVO DE ANULACION',self)
            dialog.resize(400,100)
            if dialog.exec():
                aux_clave , motivo = dialog.getInputs()
                claves = self.conexion.root.obtener_clave('dinamica')
                clave = (aux_clave ,)
                if clave in claves:
                    print('clave valida')
                    formato = {
                            "estado" : 'ANULADA',
                            "motivo" : motivo,
                            "usuario": self.datos_usuario[0]
                        }
                    extra = json.dumps(formato)
                    
                    tipo = self.lb_tipo_orden.text()
                    
                    try:
                        self.conexion.root.anular_orden( tipo.lower(), extra, self.nro_orden )
                        self.conexion.root.eliminar_clave(aux_clave,'dinamica')
                        QMessageBox.about(self,'EXITO' ,'Orden nro: ' + str(self.nro_orden) + ' Anulada')
                        self.inicializar_buscar_orden()
                    except EOFError:
                        self.conexion_perdida()
                else:
                    #print('clave invalida')
                    QMessageBox.about(self,'ERROR' ,'La clave ingresada no es valida')
    
    def validar_item_orden(self,tabla):
        lineas_totales = 0 #Para validar el total de lineas a mostrar en el PDF. (max 14 lineas)
        cant = tabla.rowCount() #cantidad de items de la tabla
        vacias = False
        correcto = True
        cantidades = []
        descripciones = []
        valores_neto = []
        i = 0
        while i< cant:
            cantidad = tabla.item(i,0) #Collumna cantidades
            descripcion = tabla.item(i,1) #Columna descripcion
            neto = tabla.item(i,2) #Columna de valor neto
            if cantidad != None and descripcion != None and neto != None :  #Checkea si se creo una fila, esta no este vacia.
                if cantidad.text() != '' and descripcion.text() != '' and neto.text() != '' :  #Chekea si se modifico una fila, esta no este vacia
                    try: 
                        nueva_cant = cantidad.text().replace(',','.',3)
                        nuevo_neto = neto.text().replace(',','.',3)
                        cantidades.append( float(nueva_cant) )
                        descripciones.append((descripcion.text()).upper() ) #v5.9
                        lineas = self.separar(descripcion.text())
                        lineas_totales = lineas_totales + len(lineas)
                        valores_neto.append(float(nuevo_neto))

                    except ValueError:
                        correcto = False
                else:
                    vacias=True
            else:
                vacias = True
            i+=1
        return cantidades,descripciones,valores_neto,vacias,correcto,lineas_totales
    
    def buscar_nro_orden(self,tupla):
        mayor = 0
        for item in tupla:
            if item[0] > mayor :
                mayor = item[0]
        return mayor
    def ver_pdf_reingreso(self):
        #ruta = self.carpeta + '/reingresos/reingreso_' + str(self.nro_reingreso) + '.pdf'
        ruta = os.path.join(self.dir_reingreso , f"reingreso_{ str(self.nro_reingreso) }.pdf")
        Popen([ruta], shell=True)

    def ver_pdf(self, tipo):
        #v6.0
        #abrir = self.carpeta+ '/ordenes/' + tipo.lower() +'_' +str(self.nro_orden) + '.pdf'
        abrir = os.path.join(self.dir_ordenes , f"{ tipo.lower() }_{ str(self.nro_orden) }.pdf" )
        Popen([abrir], shell=True)
    def filas_estimadas_orden(self,table):
        descripciones = []
        filas_estimadas = 0
        for row in range(table.rowCount()):
            descripcion_item = table.cellWidget(row, 1)
            #print(type(descripcion_item))
            if descripcion_item is not None:
                descripcion = descripcion_item.text()
                filas_usadas = self.separar(descripcion)
                filas_estimadas += len(filas_usadas)
                descripciones.append(descripcion)
        #print('Filas estimadas: ',filas_estimadas)
        return filas_estimadas


    def separar(self,cadena):
        lista = []
        iter = len(cadena)/54
        iter = int(iter) + 1 #cantidad de items a escribir
        #print('----------------------------------------------------')
       # print(cadena)
        #print('espacios necesarios: ' + str(iter))
        i = 0
        while len(cadena)> 54:
        
            #print('long > 54:')
            aux = cadena[0:54]
            index = aux[::-1].find(' ')
        
            aux = aux[:(54-index)]
           # print('Iteracion: '+str(i)+ ': '+ aux)
            lista.append(aux)
            cadena = cadena[54 - index :]
            i += 1

        if len(cadena) > 0 :
            vacias = cadena.count(' ')
            if vacias == len(cadena):
                print('item O CADENA vacio')
                #print('----------------------------------------------------')
            else:
               # print('LONG < 54: ' + cadena)
                lista.append(cadena)
                #print('----------------------------------------------------')

        return lista

    def separar2(self,cadena, long): #(TEXTO, cantidad de caracteres por linea) : 54 para pdf de ordenes y 85 para reingresos
        lista = []
        iter = len(cadena)/long
        iter = int(iter) + 1 #cantidad de items a escribir
        #print('----------------------------------------------------')
        #print(cadena)
        #print('espacios necesarios: ' + str(iter))
        i = 0
        while len(cadena)> long:
        
            #print('long > '+ str(long) +':')
            aux = cadena[0:long]
            index = aux[::-1].find(' ')
        
            aux = aux[:(long-index)]
            #print('Iteracion: '+str(i)+ ': '+ aux)
            lista.append(aux)
            cadena = cadena[long - index :]
            i += 1

        if len(cadena) > 0 :
            vacias = cadena.count(' ')
            if vacias == len(cadena):
                z  = 0 
                #print('item vacio')
                #print('----------------------------------------------------')
            else:
                #print('fin long < '+ str(long) +':' + cadena)
                lista.append(cadena)
                #print('----------------------------------------------------')

        return lista

    def normalizar_cantidades(self, lista):
        aux = []
        for i in lista:
            number = str(i)
            n = number.split('.')
            if n[1] == '0':
                aux.append(int(i))
            else:
                aux.append(i)
        return aux

    def limpiar_varibles(self):
        self.vendedores = None
        self.aux_tabla = None
        self.bol_fact = None # boletas y facturas encontradas
        self.guias = None    # guias encontradas
        # -----------
        self.nro_doc = 0  
        self.nro_orden = 0 #NRO FOLIO DE LA ORDEN CREADA
        self.items = None
        self.vendedor = None
        self.tipo_doc = None
        self.fecha_venta = None
        self.inter = None #nro interno del doc
        self.tipo = None # tipo de orden (dim,elab,pall o carp)
        self.manual = None #Si la orden se hizo manual o no
        self.fecha_orden = None #FECHA en la que se creo la orden, formato DATE 
        self.nro_reingreso = 0 #folio del reingreso
        self.aux_vendedor = None
        self.datos_orden = {}

# ------ Funciones del menu -----------
    def cerrar_sesion(self):
        self.inicializar_login()

    def mostrar_menu(self):   
        if True:
            ancho = self.left_menu_container.width()
            normal = 70
            if ancho == 70:
                extender = 150
                self.btn_inicio.show()
                self.btn_buscar.setText(' Notas de venta')
                self.btn_modificar.setText(' Ordenes\nde trabajo')
                self.btn_orden_manual.setText(' Ingreso manual')
                #self.btn_generar_clave.setText('Generar clave')
                self.btn_informe.setText(' Generar informe')
                self.btn_atras.setText(' Cerrar sesión')
                self.btn_estadisticas.setText(' Estadisticas')
                self.btn_configuracion.setText(' Configuración')
                self.btn_reingreso.setText(' Reingreso')
                self.btn_retiros.setText(' Retiros')
            else:
                self.btn_inicio.hide()
                # Define una hoja de estilo para mostrar solo el icono en los botones
                icon_only_style = """
                QPushButton {
                    text-align: left;
                }

                QPushButton::text {
                    color:transparent;
                }
                """

                # Establece la hoja de estilo en cada botón
                self.btn_buscar.setText('')
                self.btn_buscar.setIconSize(QSize(40, 40)) 

                self.btn_modificar.setText('')
                self.btn_modificar.setIconSize(QSize(40, 40)) 
                
                self.btn_reingreso.setText('')
                self.btn_reingreso.setIconSize(QSize(40, 40)) 

                self.btn_orden_manual.setText('')
                self.btn_orden_manual.setIconSize(QSize(40, 40)) 

                #self.btn_generar_clave.setText('')
                #self.btn_generar_clave.setIconSize(QSize(40, 40)) 
                
                self.btn_informe.setText('')
                self.btn_informe.setIconSize(QSize(40, 40)) 

                self.btn_estadisticas.setText('')
                self.btn_estadisticas.setIconSize(QSize(40, 40)) 

                self.btn_configuracion.setIconSize(QSize(40, 40)) 
                self.btn_configuracion.setStyleSheet(icon_only_style)

                self.btn_atras.setText('')
                self.btn_atras.setIconSize(QSize(40, 40)) 
                self.btn_retiros.setText('')
                
                
                                
                extender = normal
            
            self.animation = QPropertyAnimation(self.left_menu_container, b"maximumWidth" )
            self.animation.setDuration(500)
            self.animation.setEasingCurve(QEasingCurve.InOutQuad)
            self.animation.setStartValue(ancho)
            self.animation.setEndValue(extender)
            
            self.animation.start()

    def conexion_perdida(self):
        self.conexion = None
        self.lb_conexion.setText('DESCONECTADO')
        QMessageBox.about(self,'ERROR','Se perdio la conexion con el servidor')

    def changeEvent(self, event):
        if event.type() == QEvent.WindowStateChange:
            print(event.oldState())
            if event.oldState() and Qt.WindowMinimized:
                
                print('MINIMIZADA ')
                print("Window restored (to normal or maximized state)!")

            elif event.oldState() == Qt.WindowNoState and self.windowState() == Qt.WindowMaximized:
                print("Window Maximized!")
                # ----- SE ADAPTAN TODAS LAS TABLAS  -------
                #Buscar venta
                self.tableWidget_1.setColumnWidth(0,80) #interno
                self.tableWidget_1.setColumnWidth(1,80) #documento
                self.tableWidget_1.setColumnWidth(2,80) #nro doc
                self.tableWidget_1.setColumnWidth(3,125) #fecha venta
                self.tableWidget_1.setColumnWidth(4,200) #cliente
                self.tableWidget_1.setColumnWidth(5,200) #vendedor
                self.tableWidget_1.setColumnWidth(6,100) #total
                self.tableWidget_1.setColumnWidth(7,120) #despacho domicilio
                #Crear venta
                #Buscar orden
                #Modificar orden
                #reingreso
                #ORDEN MANUAL
                #REINGRESO MANUAL
                #estadisticas
                self.tb_orden_manual_2.setColumnWidth(5,320)

    def closeEvent(self, event):
        """Generate 'question' dialog on clicking 'X' button in title bar.

        Reimplement the closeEvent() event handler to include a 'Question'
        dialog with options on how to proceed - Save, Close, Cancel buttons
        """
        reply = QMessageBox.question(
            self, "Salir",
            "¿Deseas salir de la aplicación?",
            QMessageBox.Close | QMessageBox.Cancel)

        if reply == QMessageBox.Close:
            self.finalizar_sondeo()
            print('Cerrando aplicacion')
            event.accept()
        else:
            event.ignore()
        

class InputDialog(QDialog):
    def __init__(self,label1,label2,title ,parent=None):
        super().__init__(parent)

        self.setWindowTitle(title)
        self.txt1 = QLineEdit(self)
        self.txt2 = QLineEdit(self)
        buttonBox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self);
       
        layout = QFormLayout(self)
        layout.addRow(label1, self.txt1)
        layout.addRow(label2, self.txt2)
        layout.addWidget(buttonBox)

        buttonBox.accepted.connect(self.accept)
        buttonBox.rejected.connect(self.reject)
    
    def getInputs(self):
        return self.txt1.text(), self.txt2.text()

class InputDialog2(QDialog):
    def __init__(self, label1, hide , titulo , parent=None ):
        super().__init__(parent)
        self.hide = hide
        self.setWindowTitle(titulo)
        self.txt1 = QLineEdit(self)
        self.inicializar()
        buttonBox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)

        layout = QFormLayout(self)
        layout.addRow(label1, self.txt1)
        layout.addWidget(buttonBox)

        buttonBox.accepted.connect(self.accept)
        buttonBox.rejected.connect(self.reject)
    
    def inicializar(self):
        if self.hide == True:
            self.txt1.setEchoMode(QLineEdit.Password)

    def getInputs(self):
        return self.txt1.text()

class InputDialog3(QDialog):
    def __init__(self,tipo_doc, folio, estado , parent=None ):
        super().__init__(parent)
        uic.loadUi('app/ui/upd_domicilio.ui',self)
        self.btn_guardar.clicked.connect(self.guardar)
        self.btn_cancelar.clicked.connect(self.cancelar)
        self.lb_folio.setText(folio)
        self.lb_doc.setText(tipo_doc)
        self.lb_estado.setText(estado)
        self.setMinimumHeight(300)  # Establecer una altura mínima

    def guardar(self):
        self.accept()
    def cancelar(self):
        self.reject()
    def getInputs(self):
        return self.comboBox.currentText()

class Extra(QDialog):
    def __init__(self, data , opciones, parent = None):
        super().__init__(parent)
        uic.loadUi('app/ui/extra.ui', self)
        self.data = data # Vinculaciones
        self.btn_continuar.clicked.connect(self.continuar)
        self.btn_cancelar.clicked.connect(self.cancelar)
        self.tableWidget.setColumnWidth(1,350)
        self.fr_header_info.hide()
        self.inicializar()

    def inicializar(self):
        print('|-- INICIALIZANDO QDIALOG EXTRA VINCULACIONES...')
        try:
            grouped_data = {}  # Diccionario para agrupar los valores
            print(f"|----- Ordenes: {self.data['ordenes']}")
            for item in self.data['ordenes']:
                item = json.loads(item)
                tipo = item["tipo"]
                folio = item["folio"]
                if tipo not in grouped_data:
                    grouped_data[tipo] = [folio]
                else:
                    grouped_data[tipo].append(folio)
            print(f"|----- grouped data: {grouped_data}")
            row = 0
            for tipo, valores in grouped_data.items():
                self.tableWidget.insertRow(row)
                self.tableWidget.setItem(row, 0, QTableWidgetItem(tipo.upper()))
                self.tableWidget.setItem(row, 1, QTableWidgetItem("Folio: "+str(valores)))
                row += 1
        except KeyError:
            pass
    def continuar(self):
        self.accept()
    def cancelar(self):
        self.reject()

class ValeDespacho(QDialog):
    def __init__(self , config ,conn, parent = None):
        super().__init__(parent)
        uic.loadUi("app/ui/vale_despacho.ui", self)
        self.config = config
        self.conexion = conn

        self.nombres = [] #nombres
        self.direcciones = [] #codigo
        self.referencias = [] #descripcion
        self.telefonos = [] #descripcion
        self.contactos= [] #descripcion

        self.completer_nombre = QCompleter()
        self.completer_nombre.activated.connect(self.rellenar_datos_cliente)

        self.completer_direccion = QCompleter()
        self.completer_referencia = QCompleter()
        self.completer_telefono = QCompleter()
        self.completer_contacto = QCompleter()

        self.limites_vale = [0,0,0,0,0]
        self.btn_crear_vale.clicked.connect(self.validar_datos) # Se acepta el dialog.
        self.btn_cancelar.clicked.connect(self.reject)
        self.inicializar()

    def inicializar(self):
        logo = QPixmap("app/icono_imagen/madenco logo.png")
        self.lb_logo.setPixmap(logo)

        self.dt_fecha_estimada.setCalendarPopup(True)
        fecha_actual = datetime.now()

        # Añadir 3 días a la fecha actual
        fecha_actual = fecha_actual + timedelta(days=3)

        self.dt_fecha_estimada.setDate( fecha_actual.date() )
        self.dt_fecha_estimada.setTime( fecha_actual.time() )

        if self.config["aux_params"]:
            print(f"|(Vale despacho): Aux_params -> |{self.config['aux_params']}| type: {type(self.config['aux_params'] )} ")
            self.txt_nombre.setText(self.config['aux_params']['nombre'])
            self.txt_telefono.setText(str(self.config['aux_params']['telefono']))
            self.txt_contacto.setText(self.config['aux_params']['contacto'])

        """ print(type(self.config["model_nombres"]))
        model = QStringListModel(self.config["model_nombres"])
        
        self.completer_nombre.setModel(model)
        self.completer_nombre.setMaxVisibleItems(7)
        self.completer_nombre.setCaseSensitivity(0)  # 0: no es estricto con mayus | 1: es estricto con mayus
        #print(self.completer.filterMode()) #filtermode podria ser: coincidencias de inicio, fin o que basta que contenga.
        self.txt_nombre.setCompleter(self.completer_nombre)"""
        #Se cargan COMPLETERS DE direccion ,referencia, telefono,contacto
        if self.conexion:
            try:
                print("(VALEDESPACHO): OBTENIENDO ESTADISTICAS DE TODOS LOS VALES.")
                resultado = self.conexion.root.obtener_vale_despacho_completer()
                print("(RESULTADO TYPE)  : " ,type(resultado))
                if resultado:
                    datos = pd.DataFrame(resultado) #(nombre,direccion,referencia,telefono,contacto,fecha_estimada,fecha_real,vendedor,interno)
                    self.nombres = datos[0].tolist() #nombres
                    self.direcciones = datos[1].tolist() #codigo
                    self.referencias = datos[2].tolist() #descripcion
                    self.telefonos = datos[3].tolist() #descripcion
                    self.contactos= datos[4].tolist() #descripcion

                    model = QStringListModel(self.nombres)
                    self.completer_nombre.setModel(model)
                    self.completer_nombre.setMaxVisibleItems(7)
                    self.completer_nombre.setCaseSensitivity(0) # 0: no es estricto con mayus | 1: es estricto con mayus
                    self.txt_nombre.setCompleter(self.completer_nombre)

    
                    model = QStringListModel( list(set(self.direcciones)) ) #v6.07
                    self.completer_direccion.setModel(model)
                    self.completer_direccion.setMaxVisibleItems(7)
                    self.completer_direccion.setCaseSensitivity(0) # 0: no es estricto con mayus | 1: es estricto con mayus
                    self.txt_direccion.setCompleter(self.completer_direccion)

                    model = QStringListModel( list(set(self.referencias)) ) #v6.07
                    self.completer_referencia.setModel(model)
                    self.completer_referencia.setMaxVisibleItems(7)
                    self.completer_referencia.setCaseSensitivity(0) # 0: no es estricto con mayus | 1: es estricto con mayus
                    self.txt_referencia.setCompleter(self.completer_referencia)

                    model = QStringListModel( list(set(self.telefonos)) ) #v6.07
                    self.completer_telefono.setModel(model)
                    self.completer_telefono.setMaxVisibleItems(7)
                    self.completer_telefono.setCaseSensitivity(0) # 0: no es estricto con mayus | 1: es estricto con mayus
                    self.txt_telefono.setCompleter(self.completer_telefono)

                    model = QStringListModel( list(set(self.contactos)) ) #v6.07
                    self.completer_contacto.setModel(model)
                    self.completer_contacto.setMaxVisibleItems(7)
                    self.completer_contacto.setCaseSensitivity(0) # 0: no es estricto con mayus | 1: es estricto con mayus
                    self.txt_contacto.setCompleter(self.completer_contacto)

            except EOFError:
                QMessageBox(self,"ERROR DE CONEXION", "Se perdio la conexion con el servidor.")
                pass
            except Exception as e:
                print("(ERROR): OTRO ERROR DETECTADO -> ", str(e))
                
        # Se cargan config limites.
        with open("app/config.json", "r") as f:
            config = json.load(f)
        try:
            self.limites_vale = config["limites_params_vale"]
        except KeyError:
            print("(VALE DESPACHO): ERROR AL CARGAR LIMITES DE CONFIG.")
            pass
    def rellenar_datos_cliente(self):
         
        print('------- QCOMPLETE VALE DESPACHO ACTIVADO --------------')
        nombre_cliente = self.txt_nombre.text()
        print('Nombre: ' + nombre_cliente + ' | LEN: ' + str(len(nombre_cliente)) )
        try:
            index = self.nombres.index(nombre_cliente)
        except ValueError:
            index = -1

        if index >= 0:
            #nombre_cliente = nombre_cliente.rstrip()
            #print('Nombre: ' + nombre_cliente + ' | LEN: ' + str(len(nombre_cliente)) )
            self.txt_direccion.setText(str(self.direcciones[index]))
            self.txt_referencia.setText(str(self.referencias[index]))
            self.txt_telefono.setText(str(self.telefonos[index]))
            self.txt_contacto.setText(str(self.contactos[index]))

        print('------- QCOMPLETE VALE DESPACHO FIN --------------')

    def validar_datos(self):
        limites_min = self.limites_vale #[nombre,direccion,referencia,telefono,contacto]
        print("(VALE DESPACHO): Limites -> ", self.limites_vale )
        params = dict(
            nombre = self.txt_nombre.text() ,
            direccion = self.txt_direccion.text(),
            referencia = self.txt_referencia.text(),
            telefono = self.txt_telefono.text(),
            contacto = self.txt_contacto.text()
            )
        incompletos = "Cantidades Minimas: \n "
        estado = True
        for key,value in params.items():
            if len(value) >= limites_min[key]:
                print(f"|(QDialog)| Key: {key} cumple el limite {limites_min[key]}.")
            else:
                print(f"|(QDialog)| Key: {key} no cumple el limite {limites_min[key]}.")
                incompletos = incompletos + f"| Min  {str(key)} : {str(limites_min[key])} | " 
                estado = False

        if estado:
            # VALIDAR NUMERO TELEFONICO
            pais_code = self.txt_code_telefono.text()
            telefono = self.txt_telefono.text()
            status , mensaje = Funciones.validar_telefono(pais_code, telefono)
            if not status:
                print(mensaje)
                QMessageBox.about(self,"Numero de telefono", mensaje )
                return
                
            new_params = self.obtener_datos_validos()
            Imagen.crear_voucher(new_params)

            dialog = VistaPrevia("generar",self)
            if dialog.exec():
                print("VISTA PREVIA ACEPTADA")
                self.accept() 
            else:
                print("vista previa rechazada.")
        else:
            QMessageBox.about(self,"Datos Incompletos", str(incompletos) )

        

    def obtener_datos_validos(self):
        print("(QDialog) Obteniendo datos...")

        """ Parametros voucher Despacho """
        params = dict(
            nombre = self.txt_nombre.text() ,
            direccion = self.txt_direccion.text(),
            referencia = self.txt_referencia.text(),
            telefono = self.txt_code_telefono.text().strip() + self.txt_telefono.text().strip(),
            contacto = self.txt_contacto.text(),
            fecha_estimada = str(self.dt_fecha_estimada.dateTime().toPyDateTime())
            )
        return params

class Duplicar(QDialog):
    def __init__(self,parent = None):
        super().__init__(parent)
        uic.loadUi("app/ui/duplicar.ui",self)
        self.area = None
        self.btn_aceptar.clicked.connect(self.actualizar_area)
        self.btn_cancelar.clicked.connect(self.reject)
    
    def actualizar_area(self):
        if self.r_dim.isChecked():
            self.area = "dimensionado"
        elif self.r_elab.isChecked():
            self.area = "elaboracion"
        elif self.r_carp.isChecked():
            self.area = "carpinteria"
        elif self.r_pall.isChecked():
            self.area = "pallets"
        elif self.r_sin_trans.isChecked():
            self.area = "sin_transformacion"
        else:
            QMessageBox.warning(self,"FALTAN DATOS","Porfavor Seleccione el Area de la ORDEN DE TRABAJO\nAntes de continuar.")
            return

        self.accept()
    def obtener_area(self):
        return self.area #Devuelve el area de la nueva orden.
class VistaPrevia(QDialog):
    def __init__(self,modo ,parent = None):
        super().__init__(parent)
        uic.loadUi("app/ui/vista_previa_voucher.ui",self)
        self.modo = modo
        self.rotateImageAndSetLabel("app/icono_imagen/test.png", -90 )
        self.btn_aceptar.clicked.connect(self.accept)
        self.btn_cancelar.clicked.connect(self.reject)
        self.inicializar()
        
    def inicializar(self):
        if self.modo == "generar":
            self.btn_aceptar.setText("GENERAR VALE DESPACHO")
        elif self.modo == "imprimir":
            self.btn_aceptar.setText("IMPRIMIR VALE DESPACHO")

    def rotateImageAndSetLabel(self, image_path, angle):
        pixmap = QPixmap(image_path)
        if not pixmap.isNull():
            # Rotar la imagen
            transform = QTransform().rotate(angle)
            rotated_pixmap = pixmap.transformed(transform, mode=Qt.SmoothTransformation)
            
            # Establecer la imagen rotada en el QLabel
            self.lb_voucher.setPixmap(rotated_pixmap)
        else:
            self.lb_voucher.setText("No se pudo cargar la vista previa. \nContacte al soporte.")

    
class Canvas(FigureCanvas):
    def __init__(self,parent,x,y, tipo_grafico,titulo):
        self.fig , self.ax = plt.subplots(figsize=(6, 4), constrained_layout=True )
        
        super().__init__(self.fig)
        self.setParent(parent)
        self.x = x
        self.y = y
        print(x) #labels
        print(y) #cantidades
        # MATPLOTLIB SCRIPT
        self.ax.set_title(titulo, fontsize=12)
        self.dibujar_grafico(tipo_grafico)

    def dibujar_grafico(self,tipo):
        if tipo == 'BARRAS':
            print('DIBUJANDO GRAFICO DE BARRAS')
            self.ax.barh(self.x, self.y)
        elif tipo == 'CIRCULAR':
            print('DIBUJANDO GRAFICO CIRCULAR')
            self.ax.pie(self.y ,  labels=self.x, autopct='%1.1f%%',shadow=True, startangle=90)
            self.ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
        elif tipo == 'LINEAL':
            print('DIBUJANDO GRAFICO LINEAL')
            self.ax.plot(self.x,self.y)
            plt.xticks(rotation=90)            
        elif tipo == 'STEP':
            self.ax.stem(self.x, self.y) 
            plt.xticks(rotation=90) 
        elif tipo == "SCATTER PLOT":
            np.random.seed(19680801)

            # Compute areas and colors
            N = 150
            r = 2 * np.random.rand(N)
            theta = 2 * np.pi * np.random.rand(N)
            area = 200 * r**2
            colors = theta

            self.ax = self.fig.add_subplot(projection='polar')
            c = self.ax.scatter(theta, r, c=colors, s=area, cmap='hsv', alpha=0.75) 
        
class Grafico(QWidget):
    def __init__(self, x, y, tipo_grafico,title):
        super().__init__()
        chart = Canvas(self, x , y,tipo_grafico,title)


    
